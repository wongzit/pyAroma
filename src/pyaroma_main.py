# ===================================================================
#                       * pyaroma_main.py *
#                    Main Script of py.Aroma 4
# -------------------------------------------------------------------
#
#     This script is the main file of py.Aroma 4, could be used for
#  running py.Aroma 4 from source code mode, just do:
#                  'python pyaroma_main.py'
#
#     Make sure all libraries have been installed:
#              'pip install pyqt6 openpyxl numpy matplotlib'
#
#                          by Zhe Wang @iCeMS, updated on 2023-12-10
# ===================================================================

import sys, os, webbrowser, platform, openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Border, Font, Alignment, Side
from PyQt6.QtWidgets import *
from PyQt6.QtGui import *
from PyQt6.QtCore import *
from mpl_toolkits.axes_grid1 import make_axes_locatable
import readFile, geomAnalyzer, configparser, homaCalc, pathCreator, NICSInp, NICSout, poav, pynmr
import CONSTANT
from pathlib import Path
import numpy as np
import matplotlib
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg, NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from mpl_toolkits.mplot3d.art3d import Poly3DCollection

matplotlib.use('QtAgg')

'''
*************************************************************************

           FUNCTION FOR WRITE ROUTE SECTION OF GAUSSIAN INPUT

*************************************************************************
'''
def routeFunc(sp_flag, name_add):
	nmrStrList = []
	nmrStrList.append(configFile.get('calculation', 'nmr').lower())
	
	if configFile.get('calculation', 'coupling') == 'true':
		nmrStrList.append('spinspin')
	if configFile.get('calculation', 'mix') == 'true':
		nmrStrList.append('mixed')
	
	if configFile.get('calculation', 'spinstate') == 'def':
		mtdStrList = configFile.get('calculation', 'method').lower()
	elif configFile.get('calculation', 'spinstate') == 'r':
		mtdStrList = 'r' + configFile.get('calculation', 'method').lower()
	elif configFile.get('calculation', 'spinstate') == 'u':
		mtdStrList = 'u' + configFile.get('calculation', 'method').lower()
	elif configFile.get('calculation', 'spinstate') == 'ro':
		mtdStrList = 'ro' + configFile.get('calculation', 'method').lower()
	
	basisStrList = ''
	if configFile.get('calculation', 'diffuse') == 'aug-':
		basisStrList = 'aug-' + configFile.get('calculation', 'basis')
	elif configFile.get('calculation', 'diffuse') != 'none':
		if configFile.get('calculation', 'basis')[-1] == 'g':
			if configFile.get('calculation', 'diffuse') == 'p':
				basisStrList = configFile.get('calculation', 'basis')[:-1] + '+g'
			elif configFile.get('calculation', 'diffuse') == 'pp':
				basisStrList = configFile.get('calculation', 'basis')[:-1] + '++g'
		elif configFile.get('calculation', 'basis')[-1] == '\'':
			if configFile.get('calculation', 'diffuse') == 'p':
				basisStrList = configFile.get('calculation', 'basis')[:-2] + '+g'
			elif configFile.get('calculation', 'diffuse') == 'pp':
				basisStrList = configFile.get('calculation', 'basis')[:-2] + '++g'
	elif configFile.get('calculation', 'diffuse') == 'none':
		if configFile.get('calculation', 'basis') != '6-31g\'':
			basisStrList = configFile.get('calculation', 'basis')
		elif configFile.get('calculation', 'pol1') != '' or configFile.get('calculation', 'pol2') != '':
			basisStrList = configFile.get('calculation', 'basis')[:-1]
		else:
			basisStrList = configFile.get('calculation', 'basis')
	
	if configFile.get('calculation', 'pol1') != '' or configFile.get('calculation', 'pol2') != '':
		if configFile.get('calculation', 'pol1') != '':
			if configFile.get('calculation', 'basis') == '6-31g\'':
				basisStrList = basisStrList + '(' + configFile.get('calculation', 'pol1') + '\''
			else:
				basisStrList = basisStrList + '(' + configFile.get('calculation', 'pol1')
			if configFile.get('calculation', 'pol2') != '':
				if configFile.get('calculation', 'basis') == '6-31g\'':
					basisStrList = basisStrList + ',' + configFile.get('calculation', 'pol2') + '\')'
				else:
					basisStrList = basisStrList + ',' + configFile.get('calculation', 'pol2') + ')'
			else:
				basisStrList = basisStrList + ')'
		else:
			if configFile.get('calculation', 'basis') == '6-31g\'':
				basisStrList = basisStrList + '(' + configFile.get('calculation', 'pol2') + '\')'
			else:
				basisStrList = basisStrList + '(' + configFile.get('calculation', 'pol2') + ')'
	if configFile.get('calculation', 'pol1') == '' and configFile.get('calculation', 'pol2') == '':
		if configFile.get('calculation', 'diffuse') == 'aug-':
			basisStrList = 'aug-' + configFile.get('calculation', 'basis')
		elif configFile.get('calculation', 'diffuse') != 'none':
			if configFile.get('calculation', 'basis')[-1] == 'g':
				if configFile.get('calculation', 'diffuse') == 'p':
					basisStrList = configFile.get('calculation', 'basis')[:-1] + '+g'
				elif configFile.get('calculation', 'diffuse') == 'pp':
					basisStrList = configFile.get('calculation', 'basis')[:-1] + '++g'
			elif configFile.get('calculation', 'basis')[-1] == '\'':
				if configFile.get('calculation', 'diffuse') == 'p':
					basisStrList = configFile.get('calculation', 'basis')[:-2] + '+g\''
				elif configFile.get('calculation', 'diffuse') == 'pp':
					basisStrList = configFile.get('calculation', 'basis')[:-2] + '++g\''
		elif configFile.get('calculation', 'diffuse') == 'none':
			basisStrList = configFile.get('calculation', 'basis')

	chargeSpinStr = configFile.get('calculation', 'charge') + ' ' + configFile.get('calculation', 'spinmulti')
	
	memStr = ''
	if configFile.get('calculation', 'memory') == 'true':
		memStr = '%mem=' + configFile.get('calculation', 'memamount') + configFile.get('calculation', 'memunit')
	
	cpuStr = ''
	if configFile.get('calculation', 'cpu') == 'true':
		cpuStr = '%nprocshared=' + configFile.get('calculation', 'cpuno')

	chkStr = ''
	if configFile.get('calculation', 'chk') != 'false':
		if configFile.get('calculation', 'chk') == 'name':
			chkStr = '%chk=' + fileName.split('/')[-1].split('.')[0] + name_add + '.chk'
		elif configFile.get('calculation', 'chk') == 'path':
			chkStr = '%chk=' + fileName.split('.')[0] + name_add + '.chk'
		else:
			chkStr = '%chk=' + configFile.get('calculation', 'chk')
	
	otherStr = ''
	if configFile.get('calculation', 'other') == 'true':
		otherStr = configFile.get('calculation', 'otherline')

	routinMethodStr = '# nmr='
	if len(nmrStrList) == 1:
		routinMethodStr = routinMethodStr + f'{nmrStrList[0]}'
	elif len(nmrStrList) == 2:
		routinMethodStr = routinMethodStr + f'({nmrStrList[0]},{nmrStrList[1]})'
	elif len(nmrStrList) == 3:
		routinMethodStr = routinMethodStr + f'({nmrStrList[0]},{nmrStrList[1]},{nmrStrList[2]})'
	
	routinMethodStr = routinMethodStr + f' {mtdStrList}/{basisStrList}'
	
	if configFile.get('calculation', 'qc') == 'true':
		routinMethodStr = routinMethodStr + ' scf=qc'
	if configFile.get('calculation', 'sym') == 'true':
		routinMethodStr = routinMethodStr + ' nosymm'
	if configFile.get('calculation', 'addpri') == 'true':
		routinMethodStr = '#p' + routinMethodStr[1:]

	guessStr = []
	if configFile.get('calculation', 'guess') == 'core':
		guessStr.append('core')
	elif configFile.get('calculation', 'guess') == 'huckel':
		guessStr.append('huckel')
	elif configFile.get('calculation', 'guess') == 'input':
		guessStr.append('read')
	if configFile.get('calculation', 'guessmix') == 'true':
		guessStr.append('mix')
	if configFile.get('calculation', 'guessonly') == 'true':
		guessStr.append('only')
	if configFile.get('calculation', 'guesslocal') == 'true':
		guessStr.append('local')
	if configFile.get('calculation', 'save') == 'true':
		guessStr.append('save')
	if len(guessStr) != 0 and name_add != '_3D_NICS':
		if len(guessStr) == 1:
			routinMethodStr = routinMethodStr + f' guess={guessStr[0]}'
		elif len(guessStr) == 2:
			routinMethodStr = routinMethodStr + f' guess=({guessStr[0]},{guessStr[1]})'
		elif len(guessStr) == 3:
			routinMethodStr = routinMethodStr + f' guess=({guessStr[0]},{guessStr[1]},{guessStr[2]})'
		elif len(guessStr) == 4:
			routinMethodStr = routinMethodStr + f' guess=({guessStr[0]},{guessStr[1]},{guessStr[2]},{guessStr[3]})'
		elif len(guessStr) == 5:
			routinMethodStr = routinMethodStr + f' guess=({guessStr[0]},{guessStr[1]},{guessStr[2]},{guessStr[3]},{guessStr[4]})'
	
	if configFile.get('calculation', 'solmodel') != 'none':
		if configFile.get('calculation', 'solmodel') == 'default' and configFile.get('calculation', 'solvent').lower() == 'default':
			routinMethodStr = routinMethodStr + ' scrf'
		elif configFile.get('calculation', 'solmodel') == 'default' and configFile.get('calculation', 'solvent').lower() != 'default':
			routinMethodStr = routinMethodStr + f' scrf=(solvent=' + configFile.get('calculation', 'solvent').lower() + ')'
		elif configFile.get('calculation', 'solmodel') != 'default' and configFile.get('calculation', 'solvent').lower() == 'default':
			routinMethodStr = routinMethodStr + ' scrf=' + configFile.get('calculation', 'solmodel')
		elif configFile.get('calculation', 'solmodel') != 'default' and configFile.get('calculation', 'solvent').lower() != 'default':
			routinMethodStr = routinMethodStr + ' scrf=(' + configFile.get('calculation', 'solmodel') + ',' + 'solvent=' + configFile.get('calculation', 'solvent').lower() + ')'

	if configFile.get('calculation', 'addkey') != '':
		routinMethodStr = routinMethodStr + ' ' + configFile.get('calculation', 'addkey')

	if configFile.get('general', 'connectivity') == 'true' and sp_flag == 1:
		routinMethodStr = routinMethodStr + ' geom=connectivity'
	
	routeList = []
	if cpuStr != '':
		routeList.append(cpuStr)
	if memStr != '':
		routeList.append(memStr)
	if chkStr != '':
		routeList.append(chkStr)
	if otherStr != '':
		routeList.append(otherStr)
	routeList.append(routinMethodStr)
	routeList.append('')
	routeList.append('File Created by py.Aroma 4')
	routeList.append('')
	routeList.append(chargeSpinStr)

	return routeList

'''
*************************************************************************

                       CLASSES FOR MATPLOT CANVAS

*************************************************************************
'''
class Mpl3DCanvas(FigureCanvasQTAgg):
	def __init__(self):
		fig = Figure()
		self.ax = fig.add_subplot(projection = '3d')
		self.ax.set_xlabel('X (Å)')
		self.ax.set_ylabel('Y (Å)')
		self.ax.set_zlabel('Z (Å)')
		fig.tight_layout()
		self.ax.xaxis.pane.fill = False
		self.ax.yaxis.pane.fill = False
		self.ax.zaxis.pane.fill = False
		self.ax.xaxis.pane.set_edgecolor('w')
		self.ax.yaxis.pane.set_edgecolor('w')
		self.ax.zaxis.pane.set_edgecolor('w')
		self.ax.grid(False)
		super(Mpl3DCanvas, self).__init__(fig)

class MplBLACanvas(FigureCanvasQTAgg):
	def __init__(self):
		global fig
		fig = Figure()
		self.ax = fig.add_subplot(111)
		self.ax.set_xlabel('Bond No.')
		self.ax.set_ylabel('Bond Length (Å)')
		fig.tight_layout()
		super(MplBLACanvas, self).__init__(fig)

class MplScanPathCanvas(FigureCanvasQTAgg):
	def __init__(self):
		global figScan
		figScan = Figure()
		self.ax = figScan.add_subplot(111)
		self.ax.set_aspect('equal')
		self.ax.axis('off')
		figScan.tight_layout()
		super(MplScanPathCanvas, self).__init__(figScan)

class MplScanOutCanvas(FigureCanvasQTAgg):
	def __init__(self):
		global figScanOut
		figScanOut = Figure()
		self.ax = figScanOut.add_subplot(111)
		figScanOut.tight_layout()
		super(MplScanOutCanvas, self).__init__(figScanOut)

class MplN2OutCanvas(FigureCanvasQTAgg):
	def __init__(self):
		global figN2Out
		figN2Out = Figure()
		self.ax = figN2Out.add_subplot(111)
		self.ax.set_aspect('equal')
		super(MplN2OutCanvas, self).__init__(figN2Out)

class MplInicsOutCanvas(FigureCanvasQTAgg):
	def __init__(self):
		global figInicsOut
		figInicsOut = Figure()
		self.ax = figInicsOut.add_subplot(111)
		self.ax.set_xlabel('Distance (Å)')
		self.ax.set_ylabel('NICS (ppm)')
		figInicsOut.tight_layout()
		super(MplInicsOutCanvas, self).__init__(figInicsOut)

class MplNMRCanvas(FigureCanvasQTAgg):
	def __init__(self):
		global figNmr
		figNmr = Figure()
		self.ax = figNmr.add_subplot(111)
		figNmr.tight_layout()
		super(MplNMRCanvas, self).__init__(figNmr)

'''
*************************************************************************

                    GAUSSIAN CALCULATION SETUP WINDOW

*************************************************************************
'''
class CalSetupWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('Gaussian Calculation Setup')

		setupWidget = QWidget()
		setupLayout = QHBoxLayout()

		calWidget = QWidget()

		global preWidget
		preWidget = QTextEdit()
		preWidget.setMinimumSize(600, 500)
		preFont = QFont('Monaco')
		preFont.setPointSize(15)
		preWidget.setFont(preFont)
		preWidget.setPlaceholderText('Press \"Preview\"...')

		mainLayout = QVBoxLayout()

		jobTypeGroup = QGroupBox('Job Type')           # Job Type Tab
		mainLayout.addWidget(jobTypeGroup)
		jobTypeLayout = QGridLayout()

		global nmrMtdBox
		nmrMtdBox = QComboBox()
		nmrMtdBox.addItems(['GIAO Method', 'CSGT Method', 'IGAIM Method', 'CSGT, IGAIM and Single Origin'])
		if configFile.get('calculation', 'nmr') == 'giao':
			nmrMtdBox.setCurrentText('GIAO Method')
		elif configFile.get('calculation', 'nmr') == 'csgt':
			nmrMtdBox.setCurrentText('CSGT Method')
		elif configFile.get('calculation', 'nmr') == 'igaim':
			nmrMtdBox.setCurrentText('IGAIM Method')
		elif configFile.get('calculation', 'nmr') == 'all':
			nmrMtdBox.setCurrentText('CSGT, IGAIM and Single Origin')

		jobTypeLayout.addWidget(QLabel('NMR Method:'), 0, 0)
		jobTypeLayout.addWidget(nmrMtdBox, 0, 1)
		jobTypeLayout.addWidget(QLabel('						'), 1, 3)

		global couplingCheckBox, nmrMixCheckBox
		couplingCheckBox = QCheckBox('Compute spin-spin couplings (for all atoms)')
		if configFile.get('calculation', 'coupling') == 'true':
			couplingCheckBox.setCheckState(Qt.CheckState.Checked)
		jobTypeLayout.addWidget(couplingCheckBox, 1, 0, 1, 2)
		nmrMixCheckBox = QCheckBox('NMR=mixed')
		if configFile.get('calculation', 'mix') == 'true':
			nmrMixCheckBox.setCheckState(Qt.CheckState.Checked)
		jobTypeLayout.addWidget(nmrMixCheckBox, 2, 0, 1, 2)

		jobTypeGroup.setLayout(jobTypeLayout)

		methodGroup = QGroupBox('Calculation Method')			# Method Tab
		mainLayout.addWidget(methodGroup)
		methodLayout = QGridLayout()

		global mtdBox, mtd2Box
		methodLayout.addWidget(QLabel('Method:'), 0, 0)
		mtdBox = QComboBox()
		mtdBox.addItems(['HF', 'MP2', 'B3LYP', 'CAM-B3LYP', 'PBEPBE', 'WB97XD', 'M062X'])
		mtdBox.setEditable(True)
		mtdBox.setCurrentText(configFile.get('calculation', 'method').upper())
		methodLayout.addWidget(mtdBox, 0, 2)
		mtd2Box = QComboBox()
		mtd2Box.addItems(['Default Spin', 'Restricted', 'Unrestricted', 'Restricted-Open'])
		if configFile.get('calculation', 'spinstate') == 'def':
			mtd2Box.setCurrentText('Default Spin')
		elif configFile.get('calculation', 'spinstate') == 'r':
			mtd2Box.setCurrentText('Restricted')
		elif configFile.get('calculation', 'spinstate') == 'u':
			mtd2Box.setCurrentText('Unrestricted')
		elif configFile.get('calculation', 'spinstate') == 'ro':
			mtd2Box.setCurrentText('Restricted-Open')

		global bs1Box, difBox, pol1Box, pol2Box
		methodLayout.addWidget(mtd2Box, 0, 1)
		methodLayout.addWidget(QLabel('Basis Set:'), 1, 0)
		bs1Box = QComboBox()
		bs1Box.addItems(['STO-3G', '3-21G','6-31G', '6-31G\'', '6-311G', 'CC-PVDZ', 'CC-PVTZ', 'CC-PVQZ',\
			'LANL2DZ', 'LANL2MB', 'SDD'])
		bs1Box.setEditable(True)
		bs1Box.setCurrentText(configFile.get('calculation', 'basis').upper())
		methodLayout.addWidget(bs1Box, 1, 1)

		difBox = QComboBox()
		difBox.addItems(['', '+','++', 'aug-'])

		if configFile.get('calculation', 'diffuse') == 'none':
			difBox.setCurrentText('')
		elif configFile.get('calculation', 'diffuse') == 'p':
			difBox.setCurrentText('+')
		elif configFile.get('calculation', 'diffuse') == 'pp':
			difBox.setCurrentText('++')
		elif configFile.get('calculation', 'diffuse') == 'aug-':
			difBox.setCurrentText('aug-')
		methodLayout.addWidget(difBox, 1, 2)

		methodLayout.addWidget(QLabel('('), 1, 3)
		pol1Box = QComboBox()
		pol1Box.addItems(['', 'd','2d', '3d', 'df', '2df', '3df', '3d2f'])
		pol1Box.setEditable(True)
		if configFile.get('calculation', 'pol1') == '':
			pol1Box.setCurrentText('')
		else:
			pol1Box.setCurrentText(configFile.get('calculation', 'pol1').lower())

		methodLayout.addWidget(pol1Box, 1, 4)

		methodLayout.addWidget(QLabel(','), 1, 5)
		pol2Box = QComboBox()
		pol2Box.addItems(['', 'p','2p', '3p', 'pd', '2pd', '3pd', '3p2d'])
		pol2Box.setEditable(True)
		if configFile.get('calculation', 'pol2') == '':
			pol2Box.setCurrentText('')
		else:
			pol2Box.setCurrentText(configFile.get('calculation', 'pol2').lower())
		methodLayout.addWidget(pol2Box, 1, 6)

		global chargeBox, spinBox, spinStateList
		methodLayout.addWidget(QLabel(')		'), 1, 7)
		methodLayout.addWidget(QLabel('Charge:'), 2, 0)
		chargeBox = QSpinBox()
		chargeBox.setRange(-999999, 999999)
		chargeBox.setSingleStep(1)
		chargeBox.setValue(int(configFile.get('calculation', 'charge')))
		methodLayout.addWidget(chargeBox, 2, 1)
		methodLayout.addWidget(QLabel('Spin:'), 3, 0)
		spinBox = QComboBox()
		spinBox.addItems(['Singlet', 'Doublet', 'Triplet', 'Quartet', 'Qunitet', \
			'Sextet', 'Septet', 'Octet', 'Nonet', 'Dectet', '11-et', '12-et', '13-et', \
			'14-et', '15-et', '16-et', '17-et', '18-et', '19-et', '20-et'])
		spinStateList = ['Singlet', 'Doublet', 'Triplet', 'Quartet', 'Qunitet', \
			'Sextet', 'Septet', 'Octet', 'Nonet', 'Dectet', '11-et', '12-et', '13-et', \
			'14-et', '15-et', '16-et', '17-et', '18-et', '19-et', '20-et']
		spinBox.setCurrentText(spinStateList[int(configFile.get('calculation', 'spinmulti'))-1])

		methodLayout.addWidget(spinBox, 3, 1)

		methodGroup.setLayout(methodLayout)

		linkGroup = QGroupBox('Link 0')           # Link 0 Tab
		mainLayout.addWidget(linkGroup)
		linkLayout = QGridLayout()

		global memCheckBox, mem1Box, mem2Box
		memCheckBox = QCheckBox('Memory Limit:')
		linkLayout.addWidget(memCheckBox, 0, 0)
		if configFile.get('calculation', 'memory') != 'false':
			memCheckBox.setCheckState(Qt.CheckState.Checked)

		mem1Box = QSpinBox()
		mem1Box.setRange(1, 999999)
		mem1Box.setSingleStep(1)
		mem1Box.setValue(int(configFile.get('calculation', 'memamount')))
		linkLayout.addWidget(mem1Box, 0, 1)

		mem2Box = QComboBox()
		mem2Box.addItems(['words', 'KB', 'KW', 'MB', 'MW', 'GB', 'GW', 'TB', 'TW'])
		if configFile.get('calculation', 'memunit') != 'words':
			mem2Box.setCurrentText(configFile.get('calculation', 'memunit').upper())
		else:
			mem2Box.setCurrentText('words')
		linkLayout.addWidget(mem2Box, 0, 2)
		linkLayout.addWidget(QLabel(' '), 0, 3)

		global cpuCheckBox, cpuBox
		cpuCheckBox = QCheckBox('Processors:')
		if configFile.get('calculation', 'cpu') != 'false':
			cpuCheckBox.setCheckState(Qt.CheckState.Checked)
		linkLayout.addWidget(cpuCheckBox, 0, 4)

		cpuBox = QSpinBox()
		cpuBox.setRange(1, 999999)
		cpuBox.setSingleStep(1)
		cpuBox.setValue(int(configFile.get('calculation', 'cpuno')))
		linkLayout.addWidget(cpuBox, 0, 5)
		
		global chkCheckBox, chkBox, othCheckBox, othBox
		chkCheckBox = QCheckBox('Chkpoint File')
		linkLayout.addWidget(chkCheckBox, 1, 0)
		chkBox = QComboBox()
		chkBox.addItems(['Default Name', 'Full Path'])
		chkBox.setEditable(True)
		if configFile.get('calculation', 'chk') != 'false':
			chkCheckBox.setCheckState(Qt.CheckState.Checked)
			if configFile.get('calculation', 'chk') =='name':
				chkBox.setCurrentText('Default Name')
			elif configFile.get('calculation', 'chk') =='path':
				chkBox.setCurrentText('Full Path')
			else:
				chkBox.setCurrentText(configFile.get('calculation', 'chk'))

		linkLayout.addWidget(chkBox, 1, 1, 1, 5)
		othCheckBox = QCheckBox('Other:')
		if configFile.get('calculation', 'other') != 'false':
			othCheckBox.setCheckState(Qt.CheckState.Checked)
		linkLayout.addWidget(othCheckBox, 2, 0)
		othBox = QLineEdit()
		othBox.setText(configFile.get('calculation', 'otherline'))
		linkLayout.addWidget(othBox, 2, 1, 2, 5)

		linkGroup.setLayout(linkLayout)

		global qcCheckBox, symCheckBox, conCheckBox
		generalGroup = QGroupBox('General')						# General Tab
		mainLayout.addWidget(generalGroup)
		generalLayout = QGridLayout()
		qcCheckBox = QCheckBox('Use Quadratically Convergent SCF')
		if configFile.get('calculation', 'qc') != 'false':
			qcCheckBox.setCheckState(Qt.CheckState.Checked)
		generalLayout.addWidget(qcCheckBox, 0, 0)
		symCheckBox = QCheckBox('Ignore Symmetry')
		if configFile.get('calculation', 'sym') != 'false':
			symCheckBox.setCheckState(Qt.CheckState.Checked)
		generalLayout.addWidget(symCheckBox, 0, 1)
		conCheckBox = QCheckBox('Additional Print')
		if configFile.get('calculation', 'addpri') != 'false':
			conCheckBox.setCheckState(Qt.CheckState.Checked)
		generalLayout.addWidget(conCheckBox, 0, 2)

		generalGroup.setLayout(generalLayout)

		global gusMtdBox, mixChkBox, alyChkBox, locChkBox, saveChkBox
		guessGroup = QGroupBox('Guess')							# Guess Tab
		mainLayout.addWidget(guessGroup)
		guessLayout = QGridLayout()
		guessLayout.addWidget(QLabel('Guess Method:'), 0, 0)
		gusMtdBox = QComboBox()
		gusMtdBox.addItems(['Default', 'Core Hamiltonian', 'Extended Huckel', 'Read Chkpoint File'])
		if configFile.get('calculation', 'guess') == 'false':
			gusMtdBox.setCurrentText('Default')
		elif configFile.get('calculation', 'guess') == 'core':
			gusMtdBox.setCurrentText('Core Hamiltonian')
		elif configFile.get('calculation', 'guess') == 'huckel':
			gusMtdBox.setCurrentText('Extended Huckel')
		else:
			gusMtdBox.setCurrentText('Read Chkpoint File')

		guessLayout.addWidget(gusMtdBox, 0, 1, 0, 3)
		mixChkBox = QCheckBox('Mix HOMO and LUMO orbitals')
		if configFile.get('calculation', 'guessmix') != 'false':
			mixChkBox.setCheckState(Qt.CheckState.Checked)
		guessLayout.addWidget(mixChkBox, 1, 0)
		alyChkBox = QCheckBox('Only do guess (no SCF)')
		if configFile.get('calculation', 'guessonly') != 'false':
			alyChkBox.setCheckState(Qt.CheckState.Checked)
		guessLayout.addWidget(alyChkBox, 1, 1)
		locChkBox = QCheckBox('Localize orbitals')
		if configFile.get('calculation', 'guesslocal') != 'false':
			locChkBox.setCheckState(Qt.CheckState.Checked)
		guessLayout.addWidget(locChkBox, 1, 2)
		saveChkBox = QCheckBox('Save orbitals to chk file')
		if configFile.get('calculation', 'save') != 'false':
			saveChkBox.setCheckState(Qt.CheckState.Checked)
		guessLayout.addWidget(saveChkBox, 1, 3)

		guessGroup.setLayout(guessLayout)

		global solModel, solVent
		solGroup = QGroupBox('Solvation')
		mainLayout.addWidget(solGroup)
		solLayout = QGridLayout()
		solLayout.addWidget(QLabel('Model:'), 0, 0)
		solModel = QComboBox()
		solModel.addItems(['None', 'Default', 'IEFPCM', 'SMD', 'CPCM'])
		if configFile.get('calculation', 'solmodel') == 'none':
			solModel.setCurrentText('None')
		elif configFile.get('calculation', 'solmodel') == 'default':
			solModel.setCurrentText('Default')
		elif configFile.get('calculation', 'solmodel') == 'iefpcm':
			solModel.setCurrentText('IEFPCM')
		elif configFile.get('calculation', 'solmodel') == 'smd':
			solModel.setCurrentText('SMD')
		elif configFile.get('calculation', 'solmodel') == 'cpcm':
			solModel.setCurrentText('CPCM')
		solLayout.addWidget(solModel, 0, 1)

		solLayout.addWidget(QLabel('								'), 0, 2)
		solLayout.addWidget(QLabel('Solvent:'), 1, 0)
		solVent = QComboBox()
		solVent.addItems(['Default', 'Water', 'DMSO', 'Nitromethane', 'Acetonitrile', 'Methanol',\
			'Ethanol', 'Acetone', 'Dichloromethane', 'Dichloroethane', 'THF', 'Aniline', 'Chlorobenzene',\
			'Chloroform', 'Diethylether', 'Toluene', 'Benzene', 'CCl4', 'Cyclohexane', 'Heptane'])
		solVent.setEditable(True)
		solVent.setCurrentText(configFile.get('calculation', 'solvent').title())

		solLayout.addWidget(solVent, 1, 1)
		solGroup.setLayout(solLayout)

		global keyBox
		keyGroup = QGroupBox('Additional Keywords')           # Keyword Tab
		mainLayout.addWidget(keyGroup)
		keyLayout = QGridLayout()
		keyBox = QLineEdit()
		keyBox.setText(configFile.get('calculation', 'addkey'))
		keyLayout.addWidget(keyBox, 0, 0)
		keyGroup.setLayout(keyLayout)

		btnLayout = QHBoxLayout()          # Button Section
		mainLayout.addLayout(btnLayout)

		dftBtn = QPushButton('Default')
		btnLayout.addWidget(dftBtn)
		dftBtn.clicked.connect(self.calDefault)

		previewBtn = QPushButton('Preview')
		btnLayout.addWidget(previewBtn)
		previewBtn.clicked.connect(self.previewCal)

		saveBtn = QPushButton('Save')
		btnLayout.addWidget(saveBtn)
		saveBtn.clicked.connect(self.saveCal)

		clsBtn = QPushButton('Close')
		btnLayout.addWidget(clsBtn)
		clsBtn.clicked.connect(lambda: self.close())

		calWidget.setLayout(mainLayout)
		setupWidget.setLayout(setupLayout)
		setupLayout.addWidget(calWidget)
		setupLayout.addWidget(preWidget)

		self.setCentralWidget(setupWidget)

	def calDefault(self):
		nmrMtdBox.setCurrentText('GIAO Method')
		couplingCheckBox.setCheckState(Qt.CheckState.Unchecked)
		nmrMixCheckBox.setCheckState(Qt.CheckState.Unchecked)
		mtd2Box.setCurrentText('Default Spin')
		mtdBox.setCurrentText('B3LYP')
		bs1Box.setCurrentText('6-311G')
		difBox.setCurrentText('+')
		pol1Box.setCurrentText('2d')
		pol2Box.setCurrentText('p')
		chargeBox.setValue(0)
		spinBox.setCurrentText('Singlet')
		memCheckBox.setCheckState(Qt.CheckState.Checked)
		mem1Box.setValue(10)
		mem2Box.setCurrentText('GB')
		cpuCheckBox.setCheckState(Qt.CheckState.Checked)
		cpuBox.setValue(8)
		chkCheckBox.setCheckState(Qt.CheckState.Checked)
		chkBox.setCurrentText('Default Name')
		othCheckBox.setCheckState(Qt.CheckState.Unchecked)
		othBox.setText('')
		qcCheckBox.setCheckState(Qt.CheckState.Unchecked)
		symCheckBox.setCheckState(Qt.CheckState.Unchecked)
		conCheckBox.setCheckState(Qt.CheckState.Checked)
		gusMtdBox.setCurrentText('Default')
		mixChkBox.setCheckState(Qt.CheckState.Unchecked)
		alyChkBox.setCheckState(Qt.CheckState.Unchecked)
		locChkBox.setCheckState(Qt.CheckState.Unchecked)
		saveChkBox.setCheckState(Qt.CheckState.Unchecked)
		solModel.setCurrentText('None')
		solVent.setCurrentText('Default')
		keyBox.setText('')

	def saveCal(self):
		if nmrMtdBox.currentText() == 'GIAO Method':
			configFile.set('calculation', 'nmr', 'giao')
		elif nmrMtdBox.currentText() == 'CSGT Method':
			configFile.set('calculation', 'nmr', 'csgt')
		elif nmrMtdBox.currentText() == 'IGAIM Method':
			configFile.set('calculation', 'nmr', 'igaim')
		elif nmrMtdBox.currentText() == 'CSGT, IGAIM and Single Origin':
			configFile.set('calculation', 'nmr', 'all')

		if couplingCheckBox.isChecked():
			configFile.set('calculation', 'coupling', 'true')
		else:
			configFile.set('calculation', 'coupling', 'false')

		if nmrMixCheckBox.isChecked():
			configFile.set('calculation', 'mix', 'true')
		else:
			configFile.set('calculation', 'mix', 'false')

		if mtd2Box.currentText() == 'Default Spin':
			configFile.set('calculation', 'spinstate', 'def')
		elif mtd2Box.currentText() == 'Restricted':
			configFile.set('calculation', 'spinstate', 'r')
		elif mtd2Box.currentText() == 'Unrestricted':
			configFile.set('calculation', 'spinstate', 'u')
		elif mtd2Box.currentText() == 'Restricted-Open':
			configFile.set('calculation', 'spinstate', 'ro')

		configFile.set('calculation', 'method', f'{mtdBox.currentText().lower()}')
		configFile.set('calculation', 'basis', f'{bs1Box.currentText().lower()}')

		if difBox.currentText() == '':
			configFile.set('calculation', 'diffuse', 'none')
		elif difBox.currentText() == '+':
			configFile.set('calculation', 'diffuse', 'p')
		elif difBox.currentText() == '++':
			configFile.set('calculation', 'diffuse', 'pp')
		elif difBox.currentText() == 'aug-':
			configFile.set('calculation', 'diffuse', 'aug-')

		configFile.set('calculation', 'pol1', f'{pol1Box.currentText().lower()}')
		configFile.set('calculation', 'pol2', f'{pol2Box.currentText().lower()}')

		configFile.set('calculation', 'charge', f'{chargeBox.value()}')
		configFile.set('calculation', 'spinmulti', f'{spinStateList.index(spinBox.currentText()) + 1}')

		if memCheckBox.isChecked():
			configFile.set('calculation', 'memory', 'true')
		else:
			configFile.set('calculation', 'memory', 'false')

		configFile.set('calculation', 'memamount', f'{mem1Box.value()}')
		configFile.set('calculation', 'memunit', f'{mem2Box.currentText().lower()}')

		if cpuCheckBox.isChecked():
			configFile.set('calculation', 'cpu', 'true')
		else:
			configFile.set('calculation', 'cpu', 'false')

		configFile.set('calculation', 'cpuno', f'{cpuBox.value()}')

		if chkCheckBox.isChecked() == False:
			configFile.set('calculation', 'chk', 'false')
		elif chkBox.currentText() == 'Default Name':
			configFile.set('calculation', 'chk', 'name')
		elif chkBox.currentText() == 'Full Path':
			configFile.set('calculation', 'chk', 'path')
		else:
			configFile.set('calculation', 'chk', f'{chkBox.currentText()}')

		if othCheckBox.isChecked() == False:
			configFile.set('calculation', 'other', 'false')
		else:
			configFile.set('calculation', 'other', 'true')
		configFile.set('calculation', 'otherline', f'{othBox.text()}')

		if qcCheckBox.isChecked():
			configFile.set('calculation', 'qc', 'true')
		else:
			configFile.set('calculation', 'qc', 'false')

		if symCheckBox.isChecked():
			configFile.set('calculation', 'sym', 'true')
		else:
			configFile.set('calculation', 'sym', 'false')

		if conCheckBox.isChecked():
			configFile.set('calculation', 'addpri', 'true')
		else:
			configFile.set('calculation', 'addpri', 'false')

		if gusMtdBox.currentText() == 'Default':
			configFile.set('calculation', 'guess', 'false')
		elif gusMtdBox.currentText() == 'Core Hamiltonian':
			configFile.set('calculation', 'guess', 'core')
		elif gusMtdBox.currentText() == 'Extended Huckel':
			configFile.set('calculation', 'guess', 'huckel')
		elif gusMtdBox.currentText() == 'Read Chkpoint File':
			configFile.set('calculation', 'guess', 'input')
		
		if mixChkBox.isChecked():
			configFile.set('calculation', 'guessmix', 'true')
		else:
			configFile.set('calculation', 'guessmix', 'false')

		if alyChkBox.isChecked():
			configFile.set('calculation', 'guessonly', 'true')
		else:
			configFile.set('calculation', 'guessonly', 'false')

		if locChkBox.isChecked():
			configFile.set('calculation', 'guesslocal', 'true')
		else:
			configFile.set('calculation', 'guesslocal', 'false')

		if saveChkBox.isChecked():
			configFile.set('calculation', 'save', 'true')
		else:
			configFile.set('calculation', 'save', 'false')

		if solModel.currentText() == 'None':
			configFile.set('calculation', 'solmodel', 'none')
		elif solModel.currentText() == 'Default':
			configFile.set('calculation', 'solmodel', 'default')
		elif solModel.currentText() == 'IEFPCM':
			configFile.set('calculation', 'solmodel', 'iefpcm')
		elif solModel.currentText() == 'SMD':
			configFile.set('calculation', 'solmodel', 'smd')
		elif solModel.currentText() == 'CPCM':
			configFile.set('calculation', 'solmodel', 'cpcm')

		configFile.set('calculation', 'solvent', f'{solVent.currentText()}')
		configFile.set('calculation', 'addkey', f'{keyBox.text()}')

		with open(configFilePath, 'w') as configFileNew:
			configFile.write(configFileNew)
		self.close()

	def previewCal(self):
		nmrList = []
		if 'IGAIM' in nmrMtdBox.currentText():
			if 'Single' in nmrMtdBox.currentText():
				nmrList.append('all')
			else:
				nmrList.append('igaim')
		else:
			nmrList.append(nmrMtdBox.currentText()[:4].lower())

		if couplingCheckBox.isChecked():
			nmrList.append('spinspin')
		if nmrMixCheckBox.isChecked():
			nmrList.append('mixed')

		if mtd2Box.currentText() == 'Default Spin':
			mtdList = mtdBox.currentText().lower()
		elif mtd2Box.currentText() == 'Restricted':
			mtdList = 'r' + mtdBox.currentText().lower()
		elif mtd2Box.currentText() == 'Unrestricted':
			mtdList = 'u' + mtdBox.currentText().lower()
		elif mtd2Box.currentText() == 'Restricted-Open':
			mtdList = 'ro' + mtdBox.currentText().lower()

		basisList = ''
		if difBox.currentText() == 'aug-':
			basisList = 'aug-' + bs1Box.currentText().lower()
		elif difBox.currentText() != '':
			if bs1Box.currentText()[-1] == 'G':
				basisList = bs1Box.currentText()[:-1] + difBox.currentText() + 'g'
			elif bs1Box.currentText()[-1] == '\'':
				basisList = bs1Box.currentText()[:-2] + difBox.currentText() + 'g'
		elif difBox.currentText() == '':
			if bs1Box.currentText() != '6-31G\'':
				basisList = bs1Box.currentText().lower()
			elif pol1Box.currentText() != '' or pol2Box.currentText() != '':
				basisList = bs1Box.currentText().lower()[:-1]
			else:
				basisList = bs1Box.currentText().lower()
		if pol1Box.currentText() != '' or pol2Box.currentText() != '':
			if pol1Box.currentText() != '':
				if bs1Box.currentText() == '6-31G\'':
					basisList = basisList + '(' + pol1Box.currentText() + '\''
				else:
					basisList = basisList + '(' + pol1Box.currentText()
				if pol2Box.currentText() != '':
					if bs1Box.currentText() == '6-31G\'':
						basisList = basisList + ',' + pol2Box.currentText() + '\')'
					else:
						basisList = basisList + ',' + pol2Box.currentText() + ')'
				else:
					basisList = basisList + ')'
			else:
				if bs1Box.currentText() == '6-31G\'':
					basisList = basisList + '(' + pol2Box.currentText() + '\')'
				else:
					basisList = basisList + '(' + pol2Box.currentText() + ')'
		if pol1Box.currentText() == '' and pol2Box.currentText() == '':
			if difBox.currentText() == 'aug-':
				basisList = 'aug-' + bs1Box.currentText().lower()
			elif difBox.currentText() != '':
				if bs1Box.currentText()[-1] == 'G':
					basisList = bs1Box.currentText()[:-1] + difBox.currentText() + 'g'
				elif bs1Box.currentText()[-1] == '\'':
					basisList = bs1Box.currentText()[:-2] + difBox.currentText() + 'g\''
			elif difBox.currentText() == '':
				basisList = bs1Box.currentText().lower()

		chargeSpinLine = f'{chargeBox.value()} {spinStateList.index(spinBox.currentText()) + 1}'

		memLine = ''
		if memCheckBox.isChecked():
			memLine = f'%mem={mem1Box.value()}{mem2Box.currentText().lower()}\n'

		cpuLine = ''
		if cpuCheckBox.isChecked():
			cpuLine = f'%nprocshared={cpuBox.value()}\n'

		chkLine = ''
		if chkCheckBox.isChecked() and chkBox.currentText() == 'Default Name':
			chkLine = '%chk=filename.chk\n'
		elif chkCheckBox.isChecked() and chkBox.currentText() == 'Full Path':
			chkLine = '%chk=/path/to/filename.chk\n'
		elif chkCheckBox.isChecked():
			chkLine = f'%chk={chkBox.currentText()}\n'

		otherLine = ''
		if othCheckBox.isChecked():
			otherLine = f'{othBox.text()}\n'

		routinMethodLine = '# nmr='
		if len(nmrList) == 1:
			routinMethodLine = routinMethodLine + f'{nmrList[0]}'
		elif len(nmrList) == 2:
			routinMethodLine = routinMethodLine + f'({nmrList[0]},{nmrList[1]})'
		elif len(nmrList) == 3:
			routinMethodLine = routinMethodLine + f'({nmrList[0]},{nmrList[1]},{nmrList[2]})'
		
		routinMethodLine = routinMethodLine + f' {mtdList}/{basisList}'

		if qcCheckBox.isChecked():
			routinMethodLine = routinMethodLine + ' scf=qc'
		if symCheckBox.isChecked():
			routinMethodLine = routinMethodLine + ' nosymm'
		if conCheckBox.isChecked():
			routinMethodLine = '#p' + routinMethodLine[1:]

		guessList = []
		if gusMtdBox.currentText() == 'Core Hamiltonian':
			guessList.append('core')
		elif gusMtdBox.currentText() == 'Extended Huckel':
			guessList.append('huckel')
		elif gusMtdBox.currentText() == 'Read Chkpoint File':
			guessList.append('read')

		if mixChkBox.isChecked():
			guessList.append('mix')
		if alyChkBox.isChecked():
			guessList.append('only')
		if locChkBox.isChecked():
			guessList.append('local')
		if saveChkBox.isChecked():
			guessList.append('save')

		if len(guessList) != 0:
			if len(guessList) == 1:
				routinMethodLine = routinMethodLine + f' guess={guessList[0]}'
			elif len(guessList) == 2:
				routinMethodLine = routinMethodLine + f' guess=({guessList[0]},{guessList[1]})'
			elif len(guessList) == 3:
				routinMethodLine = routinMethodLine + f' guess=({guessList[0]},{guessList[1]},{guessList[2]})'
			elif len(guessList) == 4:
				routinMethodLine = routinMethodLine + f' guess=({guessList[0]},{guessList[1]},{guessList[2]},{guessList[3]})'
			elif len(guessList) == 5:
				routinMethodLine = routinMethodLine + f' guess=({guessList[0]},{guessList[1]},{guessList[2]},{guessList[3]},{guessList[4]})'
	
		if solModel.currentText() != 'None':
			if solModel.currentText() == 'Default' and solVent.currentText() == 'Default':
				routinMethodLine = routinMethodLine + ' scrf'
			elif solModel.currentText() == 'Default' and solVent.currentText() != 'Default':
				routinMethodLine = routinMethodLine + f' scrf=(solvent={solVent.currentText().lower()})'
			elif solModel.currentText() != 'Default' and solVent.currentText() == 'Default':
				routinMethodLine = routinMethodLine + f' scrf={solModel.currentText().lower()}'
			elif solModel.currentText() != 'Default' and solVent.currentText() != 'Default':
				routinMethodLine = routinMethodLine + f' scrf=({solModel.currentText().lower()},solvent={solVent.currentText().lower()})'

		if keyBox.text() != '':
			routinMethodLine = routinMethodLine + f' {keyBox.text()}'

		preWidget.setText(f'{cpuLine}{memLine}{chkLine}{otherLine}{routinMethodLine}\n\nTitle\n\n{chargeSpinLine}')

'''
*************************************************************************

                               BLA WINDOW

*************************************************************************
'''
class BLAWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('BLA: ' + fileName.split('/')[-1])

		blaMainWidget = QWidget()
		blaMainLayout = QHBoxLayout()

		molView2 = Mpl3DCanvas()
		molView2.ax.axis('off')

		if xMax == xMin:
			molView2.ax.set_xlim(xMin - 1, xMin + 1)
			molView2.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView2.ax.set_ylim(yMin - 1, yMin + 1)
			molView2.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView2.ax.set_zlim(zMin - 1, zMin + 1)
			molView2.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView2.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_j in bndAtom:
			molView2.ax.plot([xCoor[bndAtom_j[0]], xCoor[bndAtom_j[1]]], [yCoor[bndAtom_j[0]], yCoor[bndAtom_j[1]]], \
				[zCoor[bndAtom_j[0]], zCoor[bndAtom_j[1]]], '0.5')
			molView2.ax.text((xCoor[bndAtom_j[0]] + xCoor[bndAtom_j[1]]) / 2, (yCoor[bndAtom_j[0]] + yCoor[bndAtom_j[1]]) / 2,\
				(zCoor[bndAtom_j[0]] + zCoor[bndAtom_j[1]]) / 2, bndAtom.index(bndAtom_j) + 1, fontweight = 'bold',\
				bbox = dict(boxstyle = 'round', ec = (1., 0.5, 0.5), fc = (1., 0.8, 0.8)))
		molView2.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		toolBar2 = NavigationToolbar(molView2, self)

		viewLayout2 = QVBoxLayout()
		viewLayout2.addWidget(molView2)
		viewLayout2.addWidget(toolBar2)
		viewLayout2.setContentsMargins(0, 0, 0, 0)

		viewer2 = QWidget()
		viewer2.setLayout(viewLayout2)

		blaPlot = QWidget()
		blaLayout = QVBoxLayout()
		blaPlot.setLayout(blaLayout)

		global blaView
		blaView = MplBLACanvas()

		blaToolBar = NavigationToolbar(blaView, self)

		blaLayout.addWidget(blaView)
		blaLayout.addWidget(blaToolBar)

		global bndNoText
		bndNoText = QLineEdit()
		bndNoText.setPlaceholderText('Enter bond No. here, separate by space. (e.g.: 1 2 3 4 5 6)')
		blaLayout.addWidget(bndNoText)

		btnWidget = QWidget()
		btnLayout = QHBoxLayout()
		btnWidget.setLayout(btnLayout)

		plotBtn = QPushButton('Plot')
		plotBtn.clicked.connect(self.plotBLAFunc)
		btnLayout.addWidget(plotBtn)

		clrBtn = QPushButton('Clear')
		clrBtn.clicked.connect(self.clrBLAFunc)
		btnLayout.addWidget(clrBtn)

		savePngBlaBtn = QPushButton('Save .png')
		savePngBlaBtn.clicked.connect(self.savePngBlaFunc)
		btnLayout.addWidget(savePngBlaBtn)

		saveTxtBlaBtn = QPushButton('Save .txt')
		saveTxtBlaBtn.clicked.connect(self.saveTxtBLAFunc)
		btnLayout.addWidget(saveTxtBlaBtn)

		saveXlsxBlaBtn = QPushButton('Save .xlsx')
		saveXlsxBlaBtn.clicked.connect(self.saveXlsxBLAFunc)
		btnLayout.addWidget(saveXlsxBlaBtn)

		blaLayout.addWidget(btnWidget)

		blaLayout.setContentsMargins(0, 0, 0, 0)

		blaMainLayout.addWidget(viewer2)
		blaMainLayout.addWidget(blaPlot)
		blaMainWidget.setLayout(blaMainLayout)

		self.setCentralWidget(blaMainWidget)

	def plotBLAFunc(self):
		bndNoStr = str(bndNoText.text())
		global blaErrorFlag
		blaErrorFlag = 1
		if bndNoStr.strip() != '':
			global bndNoList
			bndNoList = bndNoStr.split()
			global bndLengthList
			bndLengthList = []
			for bndNo_i in bndNoList:
				if int(bndNo_i) <= len(bndAtom):
					x1 = geomList[bndAtom[int(bndNo_i) - 1][0]][1]
					x2 = geomList[bndAtom[int(bndNo_i) - 1][1]][1]
					y1 = geomList[bndAtom[int(bndNo_i) - 1][0]][2]
					y2 = geomList[bndAtom[int(bndNo_i) - 1][1]][2]
					z1 = geomList[bndAtom[int(bndNo_i) - 1][0]][3]
					z2 = geomList[bndAtom[int(bndNo_i) - 1][1]][3]
					bndLengthList.append(np.sqrt((x1-x2)**2+(y1-y2)**2+(z1-z2)**2))
			if len(bndLengthList) > 1 and len(bndLengthList) == len(bndNoList):
				bndEven = sum(bndLengthList[0::2])/len(bndLengthList[0::2])
				bndOdd = sum(bndLengthList[1::2])/len(bndLengthList[1::2])
				global blaValue
				blaValue = abs(bndEven - bndOdd)

				blaView.ax.cla()
				blaView.ax.plot(np.array(bndNoList), np.array(bndLengthList), c = '0.')
				blaView.ax.scatter(np.array(bndNoList), np.array(bndLengthList), c = '0.')
				blaView.ax.text(bndNoList[0], bndLengthList[0], f'  BLA = {blaValue:.3f}')
				blaView.ax.set_xlabel('Bond No.')
				blaView.ax.set_ylabel('Bond Length (Å)')
				fig.tight_layout()
				blaView.draw()
			else:
				blaErrorFlag = 0
		else:
			blaErrorFlag = 0
			QMessageBox.critical(None, 'Error', 'Please input more than two bond indices.')

	def clrBLAFunc(self):
		blaView.ax.cla()
		blaView.ax.set_xlabel('Bond No.')
		blaView.ax.set_ylabel('Bond Length (Å)')
		blaView.draw()
		bndNoText.clear()

	def savePngBlaFunc(self):
		self.plotBLAFunc()
		if blaErrorFlag == 1:
			fig.savefig(f'{os.path.splitext(fileName)[0]}_BLA.png', dpi = 300)

	def saveTxtBLAFunc(self):
		self.plotBLAFunc()
		if blaErrorFlag == 1:
			blaTxt = open(f'{os.path.splitext(fileName)[0]}_BLA.txt', 'w')
			blaTxt.write('#\n#   File Created by py.Aroma 4\n#   Author: Zhe Wang (Ph.D.)\n')
			blaTxt.write('#   https://wongzit.github.com/program/pyaroma/\n#\n')
			blaTxt.write(f'\nBLA = {blaValue:.6f}')
			blaTxt.write('\n\n No.     Bond Index      Bond Length (Å)      Bong Type\n')
			for i in range(len(bndNoList)):
				blaTxt.write(f'{str(i + 1).rjust(3)}         {str(bndNoList[i]).rjust(3)}             {bndLengthList[i]:.6f}\
	            {geomList[bndAtom[int(bndNoList[i]) - 1][0]][0]}{bndAtom[int(bndNoList[i]) - 1][0] + 1}-{geomList[bndAtom[int(bndNoList[i]) - 1][1]][0]}{bndAtom[int(bndNoList[i]) - 1][1] + 1}\n')

			blaTxt.write('\n\n')
			blaTxt.close()

	def saveXlsxBLAFunc(self):
		self.plotBLAFunc()
		if blaErrorFlag == 1:
			blaWB = openpyxl.Workbook()
			blaWS = blaWB.active
			blaWS['A1'] = 'Bond No.'
			blaWS['B1'] = 'Bond Index'
			blaWS['C1'] = 'Bond Length (Å)'
			blaWS['D1'] = 'Bond Type'
			blaWS['E1'] = f'BLA = {blaValue:.6f}'
			for j in range(len(bndNoList)):
				blaWS.append([j + 1, bndNoList[j], bndLengthList[j],\
					f'{geomList[bndAtom[int(bndNoList[j]) - 1][0]][0]}{bndAtom[int(bndNoList[j]) - 1][0] + 1}-{geomList[bndAtom[int(bndNoList[j]) - 1][1]][0]}{bndAtom[int(bndNoList[j]) - 1][1] + 1}'])
			blaChart = ScatterChart()
			xValue = Reference(blaWS, min_col = 1, min_row = 2, max_row = len(bndNoList) + 1)
			yValue = Reference(blaWS, min_col = 3, min_row = 2, max_row = len(bndNoList) + 1)
			blaChart.series.append(Series(yValue, xValue))
			blaChart.legend = None
			blaChart.x_axis.title = 'Bond No.'
			blaChart.y_axis.title = 'Bond Length (Å)'
			blaWS.add_chart(blaChart, 'E2')
			blaWB.save(f'{os.path.splitext(fileName)[0]}_BLA.xlsx')

'''
*************************************************************************

                                HOMA WINDOW

*************************************************************************
'''
class HOMAWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('HOMA: ' + fileName.split('/')[-1])

		homaMainWidget = QWidget()
		homaMainLayout = QHBoxLayout()

		global molView3
		molView3 = Mpl3DCanvas()
		molView3.ax.axis('off')

		# Load config.ini
		config_file = configparser.ConfigParser()
		config_file_path = os.path.join(os.path.dirname(__file__), 'config.ini')
		config_file.read(config_file_path)

		# Load homa from config.ini
		global homaPara
		homaPara = {
		    'CC':[float(config_file.get('homa', 'cc').split(',')[0]), \
		    float(config_file.get('homa', 'cc').split(',')[1])],
		    'CN':[float(config_file.get('homa', 'cn').split(',')[0]), \
		    float(config_file.get('homa', 'cn').split(',')[1])],
		    'NC':[float(config_file.get('homa', 'cn').split(',')[0]), \
		    float(config_file.get('homa', 'cn').split(',')[1])],
		    'CO':[float(config_file.get('homa', 'co').split(',')[0]), \
		    float(config_file.get('homa', 'co').split(',')[1])],
		    'OC':[float(config_file.get('homa', 'co').split(',')[0]), \
		    float(config_file.get('homa', 'co').split(',')[1])],
		    'CP':[float(config_file.get('homa', 'cp').split(',')[0]), \
		    float(config_file.get('homa', 'cp').split(',')[1])],
		    'PC':[float(config_file.get('homa', 'cp').split(',')[0]), \
		    float(config_file.get('homa', 'cp').split(',')[1])],
		    'CS':[float(config_file.get('homa', 'cs').split(',')[0]), \
		    float(config_file.get('homa', 'cs').split(',')[1])],
		    'SC':[float(config_file.get('homa', 'cs').split(',')[0]), \
		    float(config_file.get('homa', 'cs').split(',')[1])],
		    'NN':[float(config_file.get('homa', 'nn').split(',')[0]), \
		    float(config_file.get('homa', 'nn').split(',')[1])],
		    'NO':[float(config_file.get('homa', 'no').split(',')[0]), \
		    float(config_file.get('homa', 'no').split(',')[1])],
		    'ON':[float(config_file.get('homa', 'no').split(',')[0]), \
		    float(config_file.get('homa', 'no').split(',')[1])],
		    'BN':[float(config_file.get('homa', 'bn').split(',')[0]), \
		    float(config_file.get('homa', 'bn').split(',')[1])],
		    'NB':[float(config_file.get('homa', 'bn').split(',')[0]), \
		    float(config_file.get('homa', 'bn').split(',')[1])],
		    'CSE':[float(config_file.get('homa', 'cse').split(',')[0]), \
		    float(config_file.get('homa', 'cse').split(',')[1])],
		    'SEC':[float(config_file.get('homa', 'cse').split(',')[0]), \
		    float(config_file.get('homa', 'cse').split(',')[1])]
		}

		if xMax == xMin:
			molView3.ax.set_xlim(xMin - 1, xMin + 1)
			molView3.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView3.ax.set_ylim(yMin - 1, yMin + 1)
			molView3.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView3.ax.set_zlim(zMin - 1, zMin + 1)
			molView3.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView3.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_k in bndAtom:
			molView3.ax.plot([xCoor[bndAtom_k[0]], xCoor[bndAtom_k[1]]], [yCoor[bndAtom_k[0]], yCoor[bndAtom_k[1]]], \
				[zCoor[bndAtom_k[0]], zCoor[bndAtom_k[1]]], '0.5') 
		molView3.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		global monoCycles, centerHOMAList
		monoCycles = geomAnalyzer.find_monocycle(bndAtom)
		centerHOMAList = []  # Save center coordiate and HOMA value of each monocycle
		for monoCycle in monoCycles:
			homaValue = homaCalc.mono_homa(monoCycle, geomList, homaPara)
			if homaValue == -9999999999:
				pass
			else:
				# find center coordinate
				x_center = 0
				y_center = 0
				z_center = 0
				for cycleAtom_i in monoCycle:
					x_center += geomList[cycleAtom_i][1]/len(monoCycle)
					y_center += geomList[cycleAtom_i][2]/len(monoCycle)
					z_center += geomList[cycleAtom_i][3]/len(monoCycle)
				centerHOMAList.append([x_center, y_center, z_center, homaValue])

		toolBar3 = NavigationToolbar(molView3, self)

		viewLayout3 = QVBoxLayout()
		viewLayout3.addWidget(molView3)
		viewLayout3.addWidget(toolBar3)
		viewLayout3.setContentsMargins(0, 0, 0, 0)

		viewer3 = QWidget()
		viewer3.setLayout(viewLayout3)

		homaResult = QWidget()
		homaResultLayout = QVBoxLayout()
		homaResult.setLayout(homaResultLayout)

		global homaSum
		homaSum = QTextEdit()
		homaSum.append('Monocycle:\nHOMA Value\n-------------------------------\n')
		for cycle_i in monoCycles:
			ringStr = ''
			for cycle_len in range(len(cycle_i)):
				ringStr = ringStr + geomList[cycle_i[cycle_len]][0] + str(cycle_i[cycle_len] + 1) + '-'
			homaSum.append(f'{ringStr[:-1]}:\n{round(centerHOMAList[monoCycles.index(cycle_i)][3], 5)}\n')

		homaResultLayout.addWidget(homaSum)

		global homaNoText
		homaNoText = QLineEdit()
		homaNoText.setPlaceholderText('Enter atom No. here, e.g.: 1 2 3 4 5 6')
		homaResultLayout.addWidget(homaNoText)

		homaCompBtn = QPushButton('Calculate HOMA for specified ring.')
		homaCompBtn.clicked.connect(self.calUserHOMA)
		homaResultLayout.addWidget(homaCompBtn)

		homaBtnWidget = QWidget()
		homaBtnLayout = QHBoxLayout()
		homaBtnWidget.setLayout(homaBtnLayout)

		self.showHOMABtn = QCheckBox('Show HOMA Value')
		self.showHOMABtn.setCheckState(Qt.CheckState.Unchecked)
		self.showHOMABtn.stateChanged.connect(self.showHOMA)
		homaBtnLayout.addWidget(self.showHOMABtn)

		self.showLabelBtn = QCheckBox('Show Atom No.')
		self.showLabelBtn.setCheckState(Qt.CheckState.Unchecked)
		self.showLabelBtn.stateChanged.connect(self.showHOMA)
		homaBtnLayout.addWidget(self.showLabelBtn)

		homaResultLayout.addWidget(homaBtnWidget)
		homaResultLayout.setContentsMargins(0, 0, 0, 0)

		homaMainLayout.addWidget(viewer3)
		homaMainLayout.addWidget(homaResult)
		homaMainWidget.setLayout(homaMainLayout)

		self.setCentralWidget(homaMainWidget)

	def showHOMA(self):
		molView3.ax.cla()
		molView3.ax.grid(False)
		molView3.ax.axis('off')

		if xMax == xMin:
			molView3.ax.set_xlim(xMin - 1, xMin + 1)
			molView3.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView3.ax.set_ylim(yMin - 1, yMin + 1)
			molView3.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView3.ax.set_zlim(zMin - 1, zMin + 1)
			molView3.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView3.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_k in bndAtom:
			molView3.ax.plot([xCoor[bndAtom_k[0]], xCoor[bndAtom_k[1]]], [yCoor[bndAtom_k[0]], yCoor[bndAtom_k[1]]], \
				[zCoor[bndAtom_k[0]], zCoor[bndAtom_k[1]]], '0.5') 
		molView3.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		if self.showHOMABtn.isChecked():
			if len(centerHOMAList) != 0:
				for homaEle in centerHOMAList:
					molView3.ax.text(homaEle[0], homaEle[1], homaEle[2], round(homaEle[3], 3),\
					bbox = dict(boxstyle = 'round', ec = (0., 0.5, 0.5), fc = (0., 0.8, 0.8)))

		if self.showLabelBtn.isChecked():
			for atom_i in range(len(geomList)):
				molView3.ax.text(geomList[atom_i][1], geomList[atom_i][2], geomList[atom_i][3], atom_i + 1, color = 'darkred')
		
		molView3.draw()
		
	def calUserHOMA(self):
		atmNoStr = str(homaNoText.text())
		if atmNoStr.strip() != '':
			atmNoList = []
			for atom_j in atmNoStr.split():
				atmNoList.append(int(atom_j) - 1)
			if len(atmNoList) >= 3:
				usrHoma = homaCalc.mono_homa(atmNoList, geomList, homaPara)
				if usrHoma == -9999999999:
					QMessageBox.critical(None, 'Unsupported Atom Type', 'Including unsupported atom type.')
				else:
					ringStr2 = ''
					x_cen = 0
					y_cen = 0
					z_cen = 0
					for cycleAtom_j in atmNoList:
						ringStr2 = ringStr2 + geomList[cycleAtom_j][0] + str(cycleAtom_j + 1) + '-'
						x_cen += geomList[cycleAtom_j][1]/len(atmNoList)
						y_cen += geomList[cycleAtom_j][2]/len(atmNoList)
						z_cen += geomList[cycleAtom_j][3]/len(atmNoList)
					centerHOMAList.append([x_cen, y_cen, z_cen, usrHoma])
					homaSum.append(f'{ringStr2[:-1]}:\n{round(usrHoma, 5)}')
			else:
				QMessageBox.critical(None, 'Bad Number of Atoms', 'Please input at least three atoms.')

			self.showHOMA()
'''
*************************************************************************

                               POAV WINDOW

*************************************************************************
'''
class PovaWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('POAV: ' + fileName.split('/')[-1])

		povaMainWidget = QWidget()
		povaMainLayout = QHBoxLayout()

		global showLChk
		showLChk = QCheckBox('Show Atom No.')
		showLChk.setCheckState(Qt.CheckState.Checked)
		showLChk.stateChanged.connect(self.showLabelPoav)

		global molView18
		molView18 = Mpl3DCanvas()
		molView18.ax.axis('off')

		if xMax == xMin:
			molView18.ax.set_xlim(xMin - 1, xMin + 1)
			molView18.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView18.ax.set_ylim(yMin - 1, yMin + 1)
			molView18.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView18.ax.set_zlim(zMin - 1, zMin + 1)
			molView18.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView18.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_u in bndAtom:
			molView18.ax.plot([xCoor[bndAtom_u[0]], xCoor[bndAtom_u[1]]], [yCoor[bndAtom_u[0]], yCoor[bndAtom_u[1]]], \
				[zCoor[bndAtom_u[0]], zCoor[bndAtom_u[1]]], '0.5') 
		molView18.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		for atom_j in range(len(geomList)):
			molView18.ax.text(geomList[atom_j][1], geomList[atom_j][2], geomList[atom_j][3], atom_j + 1, color = 'darkred')

		toolBar19 = NavigationToolbar(molView18, self)

		viewLayout18 = QVBoxLayout()
		viewLayout18.addWidget(molView18)
		viewLayout18.addWidget(toolBar19)
		viewLayout18.setContentsMargins(0, 0, 0, 0)

		viewer18 = QWidget()
		viewer18.setLayout(viewLayout18)

		povaWidget = QWidget()
		povaLayout = QVBoxLayout()
		povaWidget.setLayout(povaLayout)

		povaSetGroup = QGroupBox()
		povaSetLayout = QHBoxLayout()
		povaSetGroup.setLayout(povaSetLayout)

		povaSetLayout.addWidget(showLChk)

		global inputLine
		inputLine = QLineEdit()
		inputLine.setPlaceholderText('Input atom No. here, only one atom!')
		povaSetLayout.addWidget(inputLine)

		povaBtn = QPushButton('Compute')
		povaBtn.clicked.connect(self.computPOAV)
		povaSetLayout.addWidget(povaBtn)

		povaLayout.addWidget(povaSetGroup)

		povaResult = QWidget()
		povaresultLayout = QGridLayout()
		povaResult.setLayout(povaresultLayout)

		povaresultLayout.addWidget(QLabel('POAV1'), 0, 1)
		povaresultLayout.addWidget(QLabel('POAV2'), 0, 3)
		povaresultLayout.addWidget(QLabel('θ(σπ) / ˚'), 1, 0)
		povaresultLayout.addWidget(QLabel('θ(p) / ˚'), 2, 0)
		povaresultLayout.addWidget(QLabel('n'), 3, 0)
		povaresultLayout.addWidget(QLabel('m'), 4, 0)
		povaresultLayout.addWidget(QLabel('θ(1π) / ˚'), 1, 2)
		povaresultLayout.addWidget(QLabel('θ(2π) / ˚'), 2, 2)
		povaresultLayout.addWidget(QLabel('θ(3π) / ˚'), 3, 2)
		povaresultLayout.addWidget(QLabel('n(1)'), 4, 2)
		povaresultLayout.addWidget(QLabel('n(2)'), 5, 2)
		povaresultLayout.addWidget(QLabel('n(3)'), 6, 2)
		povaresultLayout.addWidget(QLabel('m'), 7, 2)

		global thetaSP, thetaP, povaN, povaM, theta1P, theta2P, theta3P, pova2N1, pova2N2, pova2N3, pova2M
		thetaSP = QLineEdit()
		povaresultLayout.addWidget(thetaSP, 1, 1)
		thetaP = QLineEdit()
		povaresultLayout.addWidget(thetaP, 2, 1)
		povaN = QLineEdit()
		povaresultLayout.addWidget(povaN, 3, 1)
		povaM = QLineEdit()
		povaresultLayout.addWidget(povaM, 4, 1)
		theta1P = QLineEdit()
		povaresultLayout.addWidget(theta1P, 1, 3)
		theta2P = QLineEdit()
		povaresultLayout.addWidget(theta2P, 2, 3)
		theta3P = QLineEdit()
		povaresultLayout.addWidget(theta3P, 3, 3)
		pova2N1 = QLineEdit()
		povaresultLayout.addWidget(pova2N1, 4, 3)
		pova2N2 = QLineEdit()
		povaresultLayout.addWidget(pova2N2, 5, 3)
		pova2N3 = QLineEdit()
		povaresultLayout.addWidget(pova2N3, 6, 3)
		pova2M = QLineEdit()
		povaresultLayout.addWidget(pova2M, 7, 3)

		povaLayout.addWidget(povaResult)

		global infoTextPova
		infoTextPova = QTextEdit()
		infoTextPova.setDisabled(True)
		povaLayout.addWidget(infoTextPova)

		povaMainLayout.addWidget(viewer18)
		povaMainLayout.addWidget(povaWidget)
		povaMainWidget.setLayout(povaMainLayout)

		self.setCentralWidget(povaMainWidget)

	def showLabelPoav(self):
		molView18.ax.cla()
		molView18.ax.axis('off')

		if xMax == xMin:
			molView18.ax.set_xlim(xMin - 1, xMin + 1)
			molView18.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView18.ax.set_ylim(yMin - 1, yMin + 1)
			molView18.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView18.ax.set_zlim(zMin - 1, zMin + 1)
			molView18.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView18.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_u in bndAtom:
			molView18.ax.plot([xCoor[bndAtom_u[0]], xCoor[bndAtom_u[1]]], [yCoor[bndAtom_u[0]], yCoor[bndAtom_u[1]]], \
				[zCoor[bndAtom_u[0]], zCoor[bndAtom_u[1]]], '0.5') 
		molView18.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		if showLChk.isChecked():
			for atom_j in range(len(geomList)):
				molView18.ax.text(geomList[atom_j][1], geomList[atom_j][2], geomList[atom_j][3], atom_j + 1, color = 'darkred')

		molView18.draw()

	def computPOAV(self):
		if inputLine.text().strip() != '' and inputLine.text().strip().isdigit():
			connectedAtoms = poav.find_connect_atom(int(inputLine.text()) - 1, bndAtom)
			if len(connectedAtoms) == 0:
				#infoTextPova.setText('Bad atom, please input atom with three bonded atoms.')
				QMessageBox.critical(None, 'Bad Atom', 'Please input the atom which has and only has three bonded atoms.')
			else:
				theta12 = poav.calc_angle(geomList[int(inputLine.text()) - 1][1], geomList[int(inputLine.text()) - 1][2], geomList[int(inputLine.text()) - 1][3],\
					geomList[connectedAtoms[0]][1], geomList[connectedAtoms[0]][2], geomList[connectedAtoms[0]][3],\
					geomList[connectedAtoms[1]][1], geomList[connectedAtoms[1]][2], geomList[connectedAtoms[1]][3])
				theta13 = poav.calc_angle(geomList[int(inputLine.text()) - 1][1], geomList[int(inputLine.text()) - 1][2], geomList[int(inputLine.text()) - 1][3],\
					geomList[connectedAtoms[0]][1], geomList[connectedAtoms[0]][2], geomList[connectedAtoms[0]][3],\
					geomList[connectedAtoms[2]][1], geomList[connectedAtoms[2]][2], geomList[connectedAtoms[2]][3])
				theta23 = poav.calc_angle(geomList[int(inputLine.text()) - 1][1], geomList[int(inputLine.text()) - 1][2], geomList[int(inputLine.text()) - 1][3],\
					geomList[connectedAtoms[1]][1], geomList[connectedAtoms[1]][2], geomList[connectedAtoms[1]][3],\
					geomList[connectedAtoms[2]][1], geomList[connectedAtoms[2]][2], geomList[connectedAtoms[2]][3])
				theta_sigma_pi, theta_p, spn_n, smp_m, theta_1pi, theta_2pi, theta_3pi, n1, n2, n3, smp = poav.calc_poav(theta12, theta23, theta13)
				thetaSP.setText(format(theta_sigma_pi, '.6f'))
				thetaP.setText(format(theta_p, '.6f'))
				povaN.setText(format(spn_n, '.6f'))
				povaM.setText(format(smp_m, '.6f'))
				theta1P.setText(format(theta_1pi, '.6f'))
				theta2P.setText(format(theta_2pi, '.6f'))
				theta3P.setText(format(theta_3pi, '.6f'))
				pova2N1.setText(format(n1, '.6f'))
				pova2N2.setText(format(n2, '.6f'))
				pova2N3.setText(format(n3, '.6f'))
				pova2M.setText(format(smp, '.6f'))

			molView18.ax.cla()
			molView18.ax.axis('off')
			showLChk.setCheckState(Qt.CheckState.Unchecked)

			if xMax == xMin:
				molView18.ax.set_xlim(xMin - 1, xMin + 1)
				molView18.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
			elif yMax == yMin:
				molView18.ax.set_ylim(yMin - 1, yMin + 1)
				molView18.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
			elif zMax == zMin:
				molView18.ax.set_zlim(zMin - 1, zMin + 1)
				molView18.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
			else:
				molView18.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
			
			for bndAtom_u in bndAtom:
				molView18.ax.plot([xCoor[bndAtom_u[0]], xCoor[bndAtom_u[1]]], [yCoor[bndAtom_u[0]], yCoor[bndAtom_u[1]]], \
					[zCoor[bndAtom_u[0]], zCoor[bndAtom_u[1]]], '0.5') 
			molView18.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)
			molView18.ax.scatter(np.array([geomList[int(inputLine.text()) - 1][1]]), np.array([geomList[int(inputLine.text()) - 1][2]]), np.array([geomList[int(inputLine.text()) - 1][3]]), color = 'tomato', edgecolors = 'tomato', s = 100, depthshade = 0)

			if showLChk.isChecked():
				for atom_j in range(len(geomList)):
					molView18.ax.text(geomList[atom_j][1], geomList[atom_j][2], geomList[atom_j][3], atom_j + 1, color = 'darkred')

			molView18.draw()

		else:
			QMessageBox.critical(None, 'Bad Atom', 'Please input the atom number.')

'''
*************************************************************************

                              SP NICS WINDOW

*************************************************************************
'''
class SpWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('Single Point NICS: ' + fileName.split('/')[-1])

		spMainWidget = QWidget()
		spMainLayout = QHBoxLayout()

		global molView5
		molView5 = Mpl3DCanvas()
		molView5.ax.axis('off')

		if xMax == xMin:
			molView5.ax.set_xlim(xMin - 1, xMin + 1)
			molView5.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView5.ax.set_ylim(yMin - 1, yMin + 1)
			molView5.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView5.ax.set_zlim(zMin - 1, zMin + 1)
			molView5.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView5.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_m in bndAtom:
			molView5.ax.plot([xCoor[bndAtom_m[0]], xCoor[bndAtom_m[1]]], [yCoor[bndAtom_m[0]], yCoor[bndAtom_m[1]]], \
				[zCoor[bndAtom_m[0]], zCoor[bndAtom_m[1]]], '0.5') 
		molView5.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		toolBar5 = NavigationToolbar(molView5, self)

		viewLayout5 = QVBoxLayout()
		viewLayout5.addWidget(molView5)
		viewLayout5.addWidget(toolBar5)
		viewLayout5.setContentsMargins(0, 0, 0, 0)

		viewer5 = QWidget()
		viewer5.setLayout(viewLayout5)

		spWidget = QWidget()
		spLayout = QVBoxLayout()
		spWidget.setLayout(spLayout)

		usrRingGroup = QWidget()
		usrRingLayout = QHBoxLayout()
		usrRingGroup.setLayout(usrRingLayout)

		global spTextBox
		spTextBox = QTextEdit()
		spTextBox.setText('  List of Ghost Atoms (X, Y, Z):\n------------------------------------')
		spLayout.addWidget(spTextBox)

		global spAtmLine
		spAtmLine = QLineEdit()
		spAtmLine.setPlaceholderText('Enter atom No. here, e.g.: 1 2 3 4 5 6')
		usrRingLayout.addWidget(spAtmLine)

		heiGroup = QWidget()
		heiLayout = QGridLayout()
		heiGroup.setLayout(heiLayout)

		self.showSpLabelBtn = QCheckBox('Show Atom No.')
		self.showSpLabelBtn.setCheckState(Qt.CheckState.Unchecked)
		self.showSpLabelBtn.stateChanged.connect(self.reDrawSpGeom)
		heiLayout.addWidget(self.showSpLabelBtn, 0, 0)

		heiLayout.addWidget(QLabel('	Height: '), 0, 1)

		global heiSpinBox
		heiSpinBox = QLineEdit()
		heiSpinBox.setText('1.0')
		heiLayout.addWidget(heiSpinBox, 0, 2)

		bqAddBtn = QPushButton('Add Bq atom')
		bqAddBtn.clicked.connect(self.addSpFunc)
		usrRingLayout.addWidget(bqAddBtn)
		spLayout.addWidget(usrRingGroup)

		spNICSBtnGroup = QGroupBox('Ghost Atoms Editing')
		spNICSBtnLayout = QVBoxLayout()
		spNICSBtnGroup.setLayout(spNICSBtnLayout)

		undoClrGroup = QWidget()
		undoClrLayout = QHBoxLayout()
		undoClrGroup.setLayout(undoClrLayout)

		spUndoBtn = QPushButton('Undo')
		spUndoBtn.clicked.connect(self.spUndoFunc)
		undoClrLayout.addWidget(spUndoBtn)

		spClrBtn = QPushButton('Clear')
		spClrBtn.clicked.connect(self.spClrFunc)
		undoClrLayout.addWidget(spClrBtn)

		allBqAddBtn = QPushButton('Add Bq atoms for all monocycles')
		allBqAddBtn.clicked.connect(self.addAllBq)
		spNICSBtnLayout.addWidget(allBqAddBtn)

		gcBqAddBtn = QPushButton('Add Bq atoms at geometric center')
		gcBqAddBtn.clicked.connect(self.addGCBq)
		spNICSBtnLayout.addWidget(gcBqAddBtn)

		gchBqAddBtn = QPushButton('Add Bq atoms at geometric center (heavy atoms only)')
		gchBqAddBtn.clicked.connect(self.addGCHBq)
		spNICSBtnLayout.addWidget(gchBqAddBtn)

		mcBqAddBtn = QPushButton('Add Bq atoms at mass center')
		mcBqAddBtn.clicked.connect(self.addMCBq)
		spNICSBtnLayout.addWidget(mcBqAddBtn)

		mchBqAddBtn = QPushButton('Add Bq atoms at mass center (heavy atoms only)')
		mchBqAddBtn.clicked.connect(self.addMCHBq)
		spNICSBtnLayout.addWidget(mchBqAddBtn)

		spLayout.addWidget(heiGroup)
		spLayout.addWidget(undoClrGroup)
		spLayout.addWidget(spNICSBtnGroup)

		spBtnGroup = QWidget()
		spBtnLayout = QHBoxLayout()
		spBtnGroup.setLayout(spBtnLayout)

		self.spCalSetup_window = None
		spCalBtn = QPushButton('Calculation Setup')
		spCalBtn.clicked.connect(self.spCalSetup)
		spBtnLayout.addWidget(spCalBtn)

		spInpBtn = QPushButton('Save Input File')
		spInpBtn.clicked.connect(self.spInpSave)
		spBtnLayout.addWidget(spInpBtn)

		spLayout.addWidget(spBtnGroup)

		spMainLayout.addWidget(viewer5)
		spMainLayout.addWidget(spWidget)
		spMainWidget.setLayout(spMainLayout)

		self.setCentralWidget(spMainWidget)

		global spBqList
		spBqList = []

	def spCalSetup(self):
		if self.spCalSetup_window is None:
			self.spCalSetup_window = CalSetupWindow()
		self.spCalSetup_window.show()

	def spInpSave(self):
		spRouteLine = routeFunc(0, '_NICS')
		nicsSpInp = open(f'{os.path.splitext(fileName)[0]}_NICS.gjf', 'w')
		if (fileType.lower() == 'gjf' or fileType.lower() == 'com') and configFile.get('general', 'input') == 'false':
			with open(fileName, 'r') as usrInp:
				usrInpLines = usrInp.readlines()
			for usrInpLine in usrInpLines:
				if usrInpLine != '\n' and len(usrInpLine.split()) >= 4 and usrInpLine.count('.') == 3:
					break
				else:
					nicsSpInp.write(usrInpLine)
		else:
			for rou in spRouteLine:
				nicsSpInp.write(rou + '\n')
		for geomRou in geomList:
			nicsSpInp.write(f'{geomRou[0]}      {geomRou[1]:.6f}      {geomRou[2]:.6f}      {geomRou[3]:.6f}\n')
		for bqSpAtm in spBqList:
			nicsSpInp.write(f'Bq      {bqSpAtm[1]:.6f}      {bqSpAtm[2]:.6f}      {bqSpAtm[3]:.6f}\n')
		nicsSpInp.write('\n')

		nicsSpInp.write('\n')
		nicsSpInp.close()

	def reDrawSpGeom(self):
		molView5.ax.cla()
		molView5.ax.grid(False)
		molView5.ax.axis('off')

		currentList = []
		if len(spBqList) != 0:
			currentList = spBqList + geomList
		else:
			currentList = geomList
		
		xMin, xMax, yMin, yMax, zMin, zMax = geomAnalyzer.find_max_min(currentList)

		if xMax == xMin:
			molView5.ax.set_xlim(xMin - 1, xMin + 1)
			molView5.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView5.ax.set_ylim(yMin - 1, yMin + 1)
			molView5.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView5.ax.set_zlim(zMin - 1, zMin + 1)
			molView5.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView5.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_p in bndAtom:
			molView5.ax.plot([xCoor[bndAtom_p[0]], xCoor[bndAtom_p[1]]], [yCoor[bndAtom_p[0]], yCoor[bndAtom_p[1]]], \
				[zCoor[bndAtom_p[0]], zCoor[bndAtom_p[1]]], '0.5') 
		molView5.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		if len(spBqList) != 0:
			for bqAtom in spBqList:
				molView5.ax.scatter(np.array([bqAtom[1]]), np.array([bqAtom[2]]), np.array([bqAtom[3]]), color = [0.898, 0.2, 1.0], edgecolors = [0.898, 0.2, 1.0], s = 40, depthshade = 0)

		if self.showSpLabelBtn.isChecked():
			for atom_k in range(len(geomList)):
				molView5.ax.text(geomList[atom_k][1], geomList[atom_k][2], geomList[atom_k][3], atom_k + 1, color = 'darkred')

		molView5.draw()

	def addSpFunc(self):
		usrSpAtm = str(spAtmLine.text())
		if usrSpAtm.strip() != '':
			usrSpAtm2 = []
			for spAtm in usrSpAtm.split():
				usrSpAtm2.append(int(spAtm) - 1)

			if float(heiSpinBox.text()) == 0.0:
				bq0x_u, bq0y_u, bq0z_u = NICSInp.calCoor(usrSpAtm2, heiSpinBox.text(), geomList)
				spBqList.append(['Bq', bq0x_u, bq0y_u, bq0z_u])
			else:
				bq1x_u, bq1y_u, bq1z_u, bq2x_u, bq2y_u, bq2z_u = NICSInp.calCoor(usrSpAtm2, heiSpinBox.text(), geomList)
				spBqList.append(['Bq', bq1x_u, bq1y_u, bq1z_u])
				spBqList.append(['Bq', bq2x_u, bq2y_u, bq2z_u])
			spAtmLine.clear()
			self.reDrawSpGeom()

			spTextBox.setText('  List of Ghost Atoms (X, Y, Z):\n------------------------------------')
			for bq_i in spBqList:
				spTextBox.append(f'{round(bq_i[1], 5)}, {round(bq_i[2], 5)}, {round(bq_i[3], 5)}')

	def addGCBq(self):
		global gcX, gcY, gcZ
		gcX, gcY, gcZ = NICSInp.geom_center_all(geomList)
		spBqList.append(['Bq', gcX, gcY, gcZ])
		spAtmLine.clear()
		self.reDrawSpGeom()

		spTextBox.setText('  List of Ghost Atoms (X, Y, Z):\n------------------------------------')
		for bq_ii in spBqList:
			spTextBox.append(f'{round(bq_ii[1], 5)}, {round(bq_ii[2], 5)}, {round(bq_ii[3], 5)}')

	def addGCHBq(self):
		global gchX, gchY, gchZ
		gchX, gchY, gchZ = NICSInp.geom_center_heavy(geomList)
		spBqList.append(['Bq', gchX, gchY, gchZ])
		spAtmLine.clear()
		self.reDrawSpGeom()

		spTextBox.setText('  List of Ghost Atoms (X, Y, Z):\n------------------------------------')
		for bq_iii in spBqList:
			spTextBox.append(f'{round(bq_iii[1], 5)}, {round(bq_iii[2], 5)}, {round(bq_iii[3], 5)}')

	def addMCBq(self):
		global mcX, mcY, mcZ
		mcX, mcY, mcZ = NICSInp.mass_center_all(geomList)
		spBqList.append(['Bq', mcX, mcY, mcZ])
		spAtmLine.clear()
		self.reDrawSpGeom()

		spTextBox.setText('  List of Ghost Atoms (X, Y, Z):\n------------------------------------')
		for bq_jj in spBqList:
			spTextBox.append(f'{round(bq_jj[1], 5)}, {round(bq_jj[2], 5)}, {round(bq_jj[3], 5)}')

	def addMCHBq(self):
		global mchX, mchY, mchZ
		mchX, mchY, mchZ = NICSInp.mass_center_heavy(geomList)
		spBqList.append(['Bq', mchX, mchY, mchZ])
		spAtmLine.clear()
		self.reDrawSpGeom()

		spTextBox.setText('  List of Ghost Atoms (X, Y, Z):\n------------------------------------')
		for bq_jjj in spBqList:
			spTextBox.append(f'{round(bq_jjj[1], 5)}, {round(bq_jjj[2], 5)}, {round(bq_jjj[3], 5)}')

	def spUndoFunc(self):
		if len(spBqList) >= 1:
			spBqList.pop()
		self.reDrawSpGeom()
		spTextBox.setText('  List of Ghost Atoms (X, Y, Z):\n------------------------------------')
		for bq_j in spBqList:
			spTextBox.append(f'{round(bq_j[1], 5)}, {round(bq_j[2], 5)}, {round(bq_j[3], 5)}')

	def spClrFunc(self):
		spBqList.clear()
		self.reDrawSpGeom()
		spTextBox.setText('  List of Ghost Atoms (X, Y, Z):\n------------------------------------')

	def addAllBq(self):
		global spMonoCycles
		spMonoCycles = geomAnalyzer.find_monocycle(bndAtom)
		for cpCycle in spMonoCycles:
			if float(heiSpinBox.text()) == 0.0:
				bq0x, bq0y, bq0z = NICSInp.calCoor(cpCycle, heiSpinBox.text(), geomList)
				spBqList.append(['Bq', bq0x, bq0y, bq0z])
			else:
				bq1x, bq1y, bq1z, bq2x, bq2y, bq2z = NICSInp.calCoor(cpCycle, heiSpinBox.text(), geomList)
				spBqList.append(['Bq', bq1x, bq1y, bq1z])
				spBqList.append(['Bq', bq2x, bq2y, bq2z])

		self.reDrawSpGeom()
		spTextBox.setText('  List of Ghost Atoms (X, Y, Z):\n------------------------------------')
		for bq_k in spBqList:
			spTextBox.append(f'{round(bq_k[1], 5)}, {round(bq_k[2], 5)}, {round(bq_k[3], 5)}')

'''
*************************************************************************

                         1D NICS SCAN WINDOW

*************************************************************************
'''
class ScanWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('1D NICS Scan: ' + fileName.split('/')[-1])

		scanMainWidget = QWidget()
		scanMainLayout = QHBoxLayout()

		global molView4
		molView4 = Mpl3DCanvas()

		if xMax == xMin:
			molView4.ax.set_xlim(xMin - 1, xMin + 1)
			molView4.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView4.ax.set_ylim(yMin - 1, yMin + 1)
			molView4.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView4.ax.set_zlim(zMin - 1, zMin + 1)
			molView4.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView4.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_l in bndAtom:
			molView4.ax.plot([xCoor[bndAtom_l[0]], xCoor[bndAtom_l[1]]], [yCoor[bndAtom_l[0]], yCoor[bndAtom_l[1]]], \
				[zCoor[bndAtom_l[0]], zCoor[bndAtom_l[1]]], '0.5') 
		molView4.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		toolBar4 = NavigationToolbar(molView4, self)

		viewLayout4 = QVBoxLayout()
		viewLayout4.addWidget(molView4)
		viewLayout4.addWidget(toolBar4)
		viewLayout4.setContentsMargins(0, 0, 0, 0)

		viewer4 = QWidget()
		viewer4.setLayout(viewLayout4)

		global pathWin
		pathWin = MplScanPathCanvas()

		pathToolBar = NavigationToolbar(pathWin, self)

		pathViewLayout = QVBoxLayout()

		pathViewLayout.addWidget(pathWin)
		pathViewLayout.addWidget(pathToolBar)

		pathView = QWidget()
		pathView.setLayout(pathViewLayout)

		scanWidget = QWidget()
		scanLayout = QVBoxLayout()
		scanWidget.setLayout(scanLayout)

		global knotText
		knotText = QTextEdit()
		scanLayout.addWidget(knotText)
		knotText.append('  Knot List:\n--------------------')

		global atmScan, knotList
		atmScan = QLineEdit()
		atmScan.setPlaceholderText('Enter atom No. here, e.g.: 1 2 3 4 5 6')
		scanLayout.addWidget(atmScan)
		knotList = []

		scanBtnWidget = QWidget()
		scanBtnLayout = QHBoxLayout()
		scanBtnWidget.setLayout(scanBtnLayout)

		self.showLabelBtn2 = QCheckBox('Show Atom No.')
		self.showLabelBtn2.setCheckState(Qt.CheckState.Unchecked)
		self.showLabelBtn2.stateChanged.connect(self.showLabel)
		scanBtnLayout.addWidget(self.showLabelBtn2)

		addKnotBtn = QPushButton('Add Knot')
		addKnotBtn.clicked.connect(self.addKnot)
		scanBtnLayout.addWidget(addKnotBtn)

		undoBtn = QPushButton('Undo')
		undoBtn.clicked.connect(self.undoKnot)
		scanBtnLayout.addWidget(undoBtn)

		clrBtnScan = QPushButton('Clear All')
		clrBtnScan.clicked.connect(self.clrAllKnot)
		scanBtnLayout.addWidget(clrBtnScan)

		scanLayout.addWidget(scanBtnWidget)

		scanParaGroup = QGroupBox('NICS Scan Parameters')
		scanParaLayout = QGridLayout()
		scanParaGroup.setLayout(scanParaLayout)

		scanParaLayout.addWidget(QLabel('Plane: '), 0, 0)

		radioGroup = QButtonGroup()
		xyRadioBtn = QRadioButton('XY', self)
		xyRadioBtn.toggled.connect(self.planeOnClicked)
		scanParaLayout.addWidget(xyRadioBtn, 0, 1)
		yzRadioBtn = QRadioButton('YZ', self)
		yzRadioBtn.toggled.connect(self.planeOnClicked)
		scanParaLayout.addWidget(yzRadioBtn, 0, 2)
		xzRadioBtn = QRadioButton('XZ', self)
		xzRadioBtn.toggled.connect(self.planeOnClicked)
		scanParaLayout.addWidget(xzRadioBtn, 0, 3)

		radioGroup.addButton(xyRadioBtn, 1)
		radioGroup.addButton(yzRadioBtn, 2)
		radioGroup.addButton(xzRadioBtn, 3)

		xyRadioBtn.setChecked(True)

		scanParaLayout.addWidget(QLabel('Interval: '), 1, 0)

		global gridBox
		gridBox = QDoubleSpinBox()
		gridBox.setRange(0.1, 999999.0)
		gridBox.setSingleStep(0.1)
		gridBox.setValue(0.1)
		scanParaLayout.addWidget(gridBox, 1, 1)

		scanParaLayout.addWidget(QLabel('Height: '), 1, 2)

		global heightBox
		heightBox = QDoubleSpinBox()
		heightBox.setRange(-999999.0, 999999.0)
		heightBox.setSingleStep(0.1)
		heightBox.setValue(1.0)
		scanParaLayout.addWidget(heightBox, 1, 3)

		scanLayout.addWidget(scanParaGroup)

		saveBtnGroup = QWidget()
		saveBtnLayout = QHBoxLayout()
		saveBtnGroup.setLayout(saveBtnLayout)

		savePngScanBtn = QPushButton('Save .png')
		savePngScanBtn.clicked.connect(self.savePngScanFunc)
		saveBtnLayout.addWidget(savePngScanBtn)

		self.calSetup_window = None
		calSetupBtn = QPushButton('Calculation Setup')
		calSetupBtn.clicked.connect(self.calSetup)
		saveBtnLayout.addWidget(calSetupBtn)

		saveScanInpBtn = QPushButton('Save Input File')
		saveScanInpBtn.clicked.connect(self.saveScanInp)
		saveBtnLayout.addWidget(saveScanInpBtn)

		scanLayout.addWidget(saveBtnGroup)

		scanMainLayout.addWidget(viewer4)
		scanMainLayout.addWidget(pathView)
		scanMainLayout.addWidget(scanWidget)
		scanMainWidget.setLayout(scanMainLayout)

		self.setCentralWidget(scanMainWidget)
		self.setMinimumSize(1500, 600)

	def reDrawScanGeom(self):
		no_H_bond = []
		for i in bndAtom:
			if geomList[i[0]][0].upper() == 'H' or geomList[i[1]][0].upper() == 'H':
				pass
			else:
				no_H_bond.append(i)
		if scanPlane == 'XY':
			pathWin.ax.cla()
			pathWin.ax.axis('off')
			for j in no_H_bond:
				color_1 = CONSTANT.atom_colors[CONSTANT.period_table.index(geomList[j[0]][0].upper())]
				color_2 = CONSTANT.atom_colors[CONSTANT.period_table.index(geomList[j[1]][0].upper())]
				pathWin.ax.plot([geomList[j[0]][1], geomList[j[1]][1]], [geomList[j[0]][2], geomList[j[1]][2]], '0.5', zorder=0)
				for geom in geomList:
					if geom[0] != 'H':
						pathWin.ax.scatter(geom[1], geom[2], color = color_1, s = 40, edgecolors = '0.0')
			if len(knotList) != 0:
				for knot_j in knotList:
					pathWin.ax.scatter(knot_j[0], knot_j[1], color = 'red', s = 20, edgecolors = 'red')
			pathWin.draw()
		elif scanPlane == 'YZ':
			pathWin.ax.cla()
			pathWin.ax.axis('off')
			for j in no_H_bond:
				color_1 = CONSTANT.atom_colors[CONSTANT.period_table.index(geomList[j[0]][0].upper())]
				color_2 = CONSTANT.atom_colors[CONSTANT.period_table.index(geomList[j[1]][0].upper())]
				pathWin.ax.plot([geomList[j[0]][2], geomList[j[1]][2]], [geomList[j[0]][3], geomList[j[1]][3]], '0.5', zorder=0)
				for geom in geomList:
					if geom[0] != 'H':
						pathWin.ax.scatter(geom[2], geom[3], color = color_1, s = 40, edgecolors = '0.0')
			if len(knotList) != 0:
				for knot_j in knotList:
					pathWin.ax.scatter(knot_j[0], knot_j[1], color = 'red', s = 20, edgecolors = 'red')
			pathWin.draw()
		elif scanPlane == 'XZ':
			pathWin.ax.cla()
			pathWin.ax.axis('off')
			for j in no_H_bond:
				color_1 = CONSTANT.atom_colors[CONSTANT.period_table.index(geomList[j[0]][0].upper())]
				color_2 = CONSTANT.atom_colors[CONSTANT.period_table.index(geomList[j[1]][0].upper())]
				pathWin.ax.plot([geomList[j[0]][1], geomList[j[1]][1]], [geomList[j[0]][3], geomList[j[1]][3]], '0.5', zorder=0)
				for geom in geomList:
					if geom[0] != 'H':
						pathWin.ax.scatter(geom[1], geom[3], color = color_1, s = 40, edgecolors = '0.0')
			if len(knotList) != 0:
				for knot_j in knotList:
					pathWin.ax.scatter(knot_j[0], knot_j[1], color = 'red', s = 20, edgecolors = 'red')
			pathWin.draw()

	def planeOnClicked(self):
		global scanPlane
		scanPlane = ''
		if self.sender().isChecked():
			scanPlane = self.sender().text()
		self.reDrawScanGeom()

	def showLabel(self):
		molView4.ax.cla()
		molView4.ax.grid(False)
		molView4.ax.set_xlabel('X (Å)')
		molView4.ax.set_ylabel('Y (Å)')
		molView4.ax.set_zlabel('Z (Å)')

		if xMax == xMin:
			molView4.ax.set_xlim(xMin - 1, xMin + 1)
			molView4.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView4.ax.set_ylim(yMin - 1, yMin + 1)
			molView4.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView4.ax.set_zlim(zMin - 1, zMin + 1)
			molView4.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView4.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_l in bndAtom:
			molView4.ax.plot([xCoor[bndAtom_l[0]], xCoor[bndAtom_l[1]]], [yCoor[bndAtom_l[0]], yCoor[bndAtom_l[1]]], \
				[zCoor[bndAtom_l[0]], zCoor[bndAtom_l[1]]], '0.5') 
		molView4.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		if self.showLabelBtn2.isChecked():
			for atom_j in range(len(geomList)):
				molView4.ax.text(geomList[atom_j][1], geomList[atom_j][2], geomList[atom_j][3], atom_j + 1, color = 'darkred')
		molView4.draw()

	def addKnot(self):
		atmNoStr2 = str(atmScan.text())
		if atmNoStr2.strip() != '':
			ringStr3 = ''
			xCen = 0.0
			yCen = 0.0
			for atom_k in atmNoStr2.split():
				ringStr3 = ringStr3 + geomList[int(atom_k) - 1][0] + atom_k + '-'
				if scanPlane == 'XY':
					xCen += geomList[int(atom_k) - 1][1]/len(atmNoStr2.split())
					yCen += geomList[int(atom_k) - 1][2]/len(atmNoStr2.split())
				elif scanPlane == 'YZ':
					xCen += geomList[int(atom_k) - 1][2]/len(atmNoStr2.split())
					yCen += geomList[int(atom_k) - 1][3]/len(atmNoStr2.split())
				elif scanPlane == 'XZ':
					xCen += geomList[int(atom_k) - 1][1]/len(atmNoStr2.split())
					yCen += geomList[int(atom_k) - 1][3]/len(atmNoStr2.split())
			knotText.append(f'{ringStr3[:-1]}')
			knotList.append([xCen, yCen])
			atmScan.clear()
		self.showPath()

	def showPath(self):
		if len(knotList) > 1:
			self.reDrawScanGeom()
			for knot_k in range(len(knotList) - 1):
				pathWin.ax.plot([knotList[knot_k][0], knotList[knot_k+1][0]], [knotList[knot_k][1], knotList[knot_k+1][1]], 'r', lw = 1.0)
			pathWin.ax.arrow(knotList[-2][0], knotList[-2][1], knotList[-1][0]-knotList[-2][0], knotList[-1][1]-knotList[-2][1], lw = 1.0, head_width = 0.1, length_includes_head = True, fc = 'r' , ec = 'r')
			pathWin.draw()
		elif len(knotList) == 1:
			self.reDrawScanGeom()
		else:
			pass

	def undoKnot(self):
		knotText.append('(Last knot has been deleted!)')
		knotList.pop()
		pathWin.ax.cla()
		pathWin.ax.set_aspect('equal')
		pathWin.ax.axis('off')
		pathWin.draw()
		self.showPath()

	def clrAllKnot(self):
		knotText.clear()
		knotList.clear()
		pathWin.ax.cla()
		pathWin.ax.set_aspect('equal')
		pathWin.ax.axis('off')
		self.reDrawScanGeom()
		self.showPath()

	def savePngScanFunc(self):
		figScan.savefig(f'{os.path.splitext(fileName)[0]}_NICS_Scan.png', dpi = 300)

	def calSetup(self):
		if self.calSetup_window is None:
			self.calSetup_window = CalSetupWindow()
		self.calSetup_window.show()

	def saveScanInp(self):
		routeLine = routeFunc(1, '_NICS_Scan')
		nicsScanInp = open(f'{os.path.splitext(fileName)[0]}_NICS_Scan.gjf', 'w')
		if (fileType.lower() == 'gjf' or fileType.lower() == 'com') and configFile.get('general', 'input') == 'false':
			with open(fileName, 'r') as usrInp:
				usrInpLines = usrInp.readlines()
			for usrInpLine in usrInpLines:
				if usrInpLine != '\n' and len(usrInpLine.split()) >= 4 and usrInpLine.count('.') == 3:
					break
				else:
					if configFile.get('general', 'connectivity') == 'true' and usrInpLine[0] == '#' and 'geom=connectivity' not in usrInpLine:
						nicsScanInp.write(usrInpLine.strip() + ' geom=connectivity' + '\n')
					else:
						nicsScanInp.write(usrInpLine)

		else:
			for rou in routeLine:
				nicsScanInp.write(rou + '\n')
		for geomRou in geomList:
			nicsScanInp.write(f'{geomRou[0]}      {geomRou[1]:.6f}      {geomRou[2]:.6f}      {geomRou[3]:.6f}\n')
		bqScanList = pathCreator.creat_path(knotList, gridBox.value(), heightBox.value(), scanPlane)
		for bqScanRou in bqScanList:
			nicsScanInp.write(bqScanRou)

		nicsScanInp.write('\n')
		
		if configFile.get('general', 'connectivity') == 'true':
			for no_i in range(len(geomList) + len(bqScanList)):
				nicsScanInp.write(f'{no_i + 1}\n')

		nicsScanInp.write('\n')
		nicsScanInp.close()

'''
*************************************************************************

                            2D NICS WINDOW

*************************************************************************
'''
class Nics2dWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('2D NICS: ' + fileName.split('/')[-1])

		nics2dMainWidget = QWidget()
		nics2dMainLayout = QHBoxLayout()

		global molView6
		molView6 = Mpl3DCanvas()

		if xMax == xMin:
			molView6.ax.set_xlim(xMin - 1, xMin + 1)
			molView6.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView6.ax.set_ylim(yMin - 1, yMin + 1)
			molView6.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView6.ax.set_zlim(zMin - 1, zMin + 1)
			molView6.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView6.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_n in bndAtom:
			molView6.ax.plot([xCoor[bndAtom_n[0]], xCoor[bndAtom_n[1]]], [yCoor[bndAtom_n[0]], yCoor[bndAtom_n[1]]], \
				[zCoor[bndAtom_n[0]], zCoor[bndAtom_n[1]]], '0.5') 
		molView6.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		polyX0 = [xMin - 1.5, xMax + 1.5, xMax + 1.5, xMin - 1.5]
		polyY0 = [yMin - 1.5, yMin - 1.5, yMax + 1.5, yMax + 1.5]
		polyZ0 = [0.0, 0.0, 0.0, 0.0]
		poly0 = list(zip(polyX0, polyY0, polyZ0))
		molView6.ax.add_collection3d(Poly3DCollection([poly0], facecolor = '#000080', alpha = 0.3))

		toolBar6 = NavigationToolbar(molView6, self)

		viewLayout6 = QVBoxLayout()
		viewLayout6.addWidget(molView6)
		viewLayout6.addWidget(toolBar6)
		viewLayout6.setContentsMargins(0, 0, 0, 0)

		viewer6 = QWidget()
		viewer6.setLayout(viewLayout6)

		n2Widget = QWidget()
		n2Layout = QVBoxLayout()
		n2Widget.setLayout(n2Layout)

		global n2TextBox
		n2TextBox = QTextEdit()
		n2TextBox.setText('  2D NICS Information:\n-------------------------------------')
		n2Layout.addWidget(n2TextBox)
		n2TextBox.append('\nBq atoms would be added on XY plane, with height of 0.0.')
		n2TextBox.append(f'X range: from {round(xMin - 1.5, 2)} to {round(xMax + 1.5, 2)}; Y range: from {round(yMin - 1.5, 2)} to {round(yMax + 1.5, 2)}.')

		mmTable = QGroupBox('2D NICS Parameters')
		mmLayout = QGridLayout()
		mmTable.setLayout(mmLayout)

		mmLayout.addWidget(QLabel('Minimum'), 1, 0)
		mmLayout.addWidget(QLabel('Maximum'), 2, 0)
		mmLayout.addWidget(QLabel('X'), 0, 1)
		mmLayout.addWidget(QLabel('Y'), 0, 2)
		mmLayout.addWidget(QLabel('Z'), 0, 3)

		global n2XMin, n2XMax, n2YMin, n2YMax, n2ZMin, n2ZMax
		n2XMin = QDoubleSpinBox()
		n2XMin.setRange(-999999.0, 999999.0)
		n2XMin.setSingleStep(0.1)
		n2XMin.setValue(round(xMin - 1.5, 2))
		mmLayout.addWidget(n2XMin, 1, 1)
		n2XMax = QDoubleSpinBox()
		n2XMax.setRange(-999999.0, 999999.0)
		n2XMax.setSingleStep(0.1)
		n2XMax.setValue(round(xMax + 1.5, 2))
		mmLayout.addWidget(n2XMax, 2, 1)
		n2YMin = QDoubleSpinBox()
		n2YMin.setRange(-999999.0, 999999.0)
		n2YMin.setSingleStep(0.1)
		n2YMin.setValue(round(yMin - 1.5, 2))
		mmLayout.addWidget(n2YMin, 1, 2)
		n2YMax = QDoubleSpinBox()
		n2YMax.setRange(-999999.0, 999999.0)
		n2YMax.setSingleStep(0.1)
		n2YMax.setValue(round(yMax + 1.5, 2))
		mmLayout.addWidget(n2YMax, 2, 2)
		n2ZMin = QDoubleSpinBox()
		n2ZMin.setRange(-999999.0, 999999.0)
		n2ZMin.setSingleStep(0.1)
		n2ZMin.setValue(round(zMin - 1.5, 2))
		mmLayout.addWidget(n2ZMin, 1, 3)
		n2ZMax = QDoubleSpinBox()
		n2ZMax.setRange(-999999.0, 999999.0)
		n2ZMax.setSingleStep(0.1)
		n2ZMax.setValue(round(zMax + 1.5, 2))
		mmLayout.addWidget(n2ZMax, 2, 3)
		n2Layout.addWidget(mmTable)

		mmLayout.addWidget(QLabel(' '), 4, 0)

		mmLayout.addWidget(QLabel('Plane:'), 5, 0)

		radion2Group = QButtonGroup()
		xyRadion2Btn = QRadioButton('XY', self)
		xyRadion2Btn.toggled.connect(self.planeClicked)
		mmLayout.addWidget(xyRadion2Btn, 5, 1)
		yzRadion2Btn = QRadioButton('YZ', self)
		yzRadion2Btn.toggled.connect(self.planeClicked)
		mmLayout.addWidget(yzRadion2Btn, 5, 2)
		xzRadion2Btn = QRadioButton('XZ', self)
		xzRadion2Btn.toggled.connect(self.planeClicked)
		mmLayout.addWidget(xzRadion2Btn, 5, 3)

		radion2Group.addButton(xyRadion2Btn, 1)
		radion2Group.addButton(yzRadion2Btn, 2)
		radion2Group.addButton(xzRadion2Btn, 3)

		xyRadion2Btn.setChecked(True)

		mmLayout.addWidget(QLabel('Height:'), 6, 0)

		global heiN2Spin
		heiN2Spin = QDoubleSpinBox()
		heiN2Spin.setRange(-999999.0, 999999.0)
		heiN2Spin.setSingleStep(0.1)
		heiN2Spin.setValue(0.0)
		mmLayout.addWidget(heiN2Spin, 6, 1)

		mmLayout.addWidget(QLabel('	Interval:'), 6, 2)

		global gridN2Spin
		gridN2Spin = QDoubleSpinBox()
		gridN2Spin.setRange(0.1, 999999.0)
		gridN2Spin.setSingleStep(0.1)
		gridN2Spin.setValue(0.2)
		mmLayout.addWidget(gridN2Spin, 6, 3)

		N2BtnGroup = QWidget()
		N2BtnLayout = QHBoxLayout()
		N2BtnGroup.setLayout(N2BtnLayout)

		n2UpdBtn = QPushButton('Preview')
		n2UpdBtn.clicked.connect(self.reDraw2DGeom)
		N2BtnLayout.addWidget(n2UpdBtn)

		self.calN2Setup_window = None
		n2CalSetBtn = QPushButton('Calculation Setup')
		n2CalSetBtn.clicked.connect(self.calN2Setup)
		N2BtnLayout.addWidget(n2CalSetBtn)

		n2InpBtn = QPushButton('Save Input File')
		n2InpBtn.clicked.connect(self.saveN2Inp)
		N2BtnLayout.addWidget(n2InpBtn)

		n2Layout.addWidget(N2BtnGroup)

		nics2dMainLayout.addWidget(viewer6)
		nics2dMainLayout.addWidget(n2Widget)
		nics2dMainWidget.setLayout(nics2dMainLayout)

		self.setCentralWidget(nics2dMainWidget)

	def reDraw2DGeom(self):
		molView6.ax.cla()
		molView6.ax.grid(False)
		molView6.ax.set_xlabel('X (Å)')
		molView6.ax.set_ylabel('Y (Å)')
		molView6.ax.set_zlabel('Z (Å)')
		n2TextBox.append(f'\nBq atoms would be added on {n2DPlane} plane, with height of {round(heiN2Spin.value(), 2)}.')

		if xMax == xMin:
			molView6.ax.set_xlim(xMin - 1, xMin + 1)
			molView6.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView6.ax.set_ylim(yMin - 1, yMin + 1)
			molView6.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView6.ax.set_zlim(zMin - 1, zMin + 1)
			molView6.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView6.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_n in bndAtom:
			molView6.ax.plot([xCoor[bndAtom_n[0]], xCoor[bndAtom_n[1]]], [yCoor[bndAtom_n[0]], yCoor[bndAtom_n[1]]], \
				[zCoor[bndAtom_n[0]], zCoor[bndAtom_n[1]]], '0.5') 
		molView6.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		if n2DPlane == 'XY':
			polyX = [n2XMin.value(), n2XMax.value(), n2XMax.value(), n2XMin.value()]
			polyY = [n2YMin.value(), n2YMin.value(), n2YMax.value(), n2YMax.value()]
			polyZ = [heiN2Spin.value(), heiN2Spin.value(), heiN2Spin.value(), heiN2Spin.value()]
			poly = list(zip(polyX, polyY, polyZ))
			molView6.ax.add_collection3d(Poly3DCollection([poly], facecolor = '#000080', alpha = 0.3))
			n2TextBox.append(f'X range: from {round(n2XMin.value(), 2)} to {round(n2XMax.value(), 2)}; Y range: from {round(n2YMin.value(), 2)} to {round(n2YMax.value(), 2)}.')
		elif n2DPlane == 'XZ':
			polyX = [n2XMin.value(), n2XMax.value(), n2XMax.value(), n2XMin.value()]
			polyY = [heiN2Spin.value(), heiN2Spin.value(), heiN2Spin.value(), heiN2Spin.value()]
			polyZ = [n2ZMin.value(), n2ZMin.value(), n2ZMax.value(), n2ZMax.value()]
			poly = list(zip(polyX, polyY, polyZ))
			molView6.ax.add_collection3d(Poly3DCollection([poly], facecolor = '#000080', alpha = 0.3))
			n2TextBox.append(f'X range: from {round(n2XMin.value(), 2)} to {round(n2XMax.value(), 2)}; Z range: from {round(n2ZMin.value(), 2)} to {round(n2ZMax.value(), 2)}.')
		elif n2DPlane == 'YZ':
			polyX = [heiN2Spin.value(), heiN2Spin.value(), heiN2Spin.value(), heiN2Spin.value()]
			polyY = [n2YMin.value(), n2YMax.value(), n2YMax.value(), n2YMin.value()]
			polyZ = [n2ZMin.value(), n2ZMin.value(), n2ZMax.value(), n2ZMax.value()]
			poly = list(zip(polyX, polyY, polyZ))
			molView6.ax.add_collection3d(Poly3DCollection([poly], facecolor = '#000080', alpha = 0.3))
			n2TextBox.append(f'Y range: from {round(n2YMin.value(), 2)} to {round(n2YMax.value(), 2)}; Z range: from {round(n2ZMin.value(), 2)} to {round(n2ZMax.value(), 2)}.')
		molView6.draw()

	def planeClicked(self):
		global n2DPlane
		n2DPlane = ''
		if self.sender().isChecked():
			n2DPlane = self.sender().text()
		if n2DPlane == 'XY':
			n2XMin.setEnabled(True)
			n2XMax.setEnabled(True)
			n2YMin.setEnabled(True)
			n2YMax.setEnabled(True)
			n2ZMin.setEnabled(False)
			n2ZMax.setEnabled(False)
		elif n2DPlane == 'XZ':
			n2XMin.setEnabled(True)
			n2XMax.setEnabled(True)
			n2YMin.setEnabled(False)
			n2YMax.setEnabled(False)
			n2ZMin.setEnabled(True)
			n2ZMax.setEnabled(True)
		elif n2DPlane == 'YZ':
			n2XMin.setEnabled(False)
			n2XMax.setEnabled(False)
			n2YMin.setEnabled(True)
			n2YMax.setEnabled(True)
			n2ZMin.setEnabled(True)
			n2ZMax.setEnabled(True)
	
	def calN2Setup(self):
		if self.calN2Setup_window is None:
			self.calN2Setup_window = CalSetupWindow()
		self.calN2Setup_window.show()

	def saveN2Inp(self):
		routeN2Line = routeFunc(1, '_2D_NICS')
		nics2DInp = open(f'{os.path.splitext(fileName)[0]}_2D_NICS.gjf', 'w')
		if (fileType.lower() == 'gjf' or fileType.lower() == 'com') and configFile.get('general', 'input') == 'false':
			with open(fileName, 'r') as usr2DInp:
				usr2DInpLines = usr2DInp.readlines()
			for usr2DInpLine in usr2DInpLines:
				if usr2DInpLine != '\n' and len(usr2DInpLine.split()) >= 4 and usr2DInpLine.count('.') == 3:
					break
				else:
					if configFile.get('general', 'connectivity') == 'true' and usr2DInpLine[0] == '#' and 'geom=connectivity' not in usr2DInpLine:
						nics2DInp.write(usr2DInpLine.strip() + ' geom=connectivity' + '\n')
					else:
						nics2DInp.write(usr2DInpLine)

		else:
			for rou in routeN2Line:
				nics2DInp.write(rou + '\n')
		for geomRou in geomList:
			nics2DInp.write(f'{geomRou[0]}      {geomRou[1]:.6f}      {geomRou[2]:.6f}      {geomRou[3]:.6f}\n')

		if n2DPlane == 'XY':
			bq2DList = NICSInp.nics2Dxy(float(n2XMin.value()), float(n2XMax.value()), float(n2YMin.value()), float(n2YMax.value()), float(heiN2Spin.value()), float(gridN2Spin.value()))
		elif n2DPlane == 'XZ':
			bq2DList = NICSInp.nics2Dxz(float(n2XMin.value()), float(n2XMax.value()), float(n2ZMin.value()), float(n2ZMax.value()), float(heiN2Spin.value()), float(gridN2Spin.value()))
		elif n2DPlane == 'YZ':
			bq2DList = NICSInp.nics2Dyz(float(n2YMin.value()), float(n2YMax.value()), float(n2ZMin.value()), float(n2ZMax.value()), float(heiN2Spin.value()), float(gridN2Spin.value()))

		for bq2DRou in bq2DList:
			nics2DInp.write(bq2DRou)
		nics2DInp.write('\n')
		
		if configFile.get('general', 'connectivity') == 'true':
			for no_j in range(len(geomList) + len(bq2DList)):
				nics2DInp.write(f'{no_j + 1}\n')

		nics2DInp.write('\n')
		nics2DInp.close()

'''
*************************************************************************

                            3D NICS WINDOW

*************************************************************************
'''
class Nics3dWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('3D NICS: ' + fileName.split('/')[-1])

		nics3dMainWidget = QWidget()
		nics3dMainLayout = QHBoxLayout()

		global molView7
		molView7 = Mpl3DCanvas()

		if xMax == xMin:
			molView7.ax.set_xlim(xMin - 1, xMin + 1)
			molView7.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView7.ax.set_ylim(yMin - 1, yMin + 1)
			molView7.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView7.ax.set_zlim(zMin - 1, zMin + 1)
			molView7.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView7.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_o in bndAtom:
			molView7.ax.plot([xCoor[bndAtom_o[0]], xCoor[bndAtom_o[1]]], [yCoor[bndAtom_o[0]], yCoor[bndAtom_o[1]]], \
				[zCoor[bndAtom_o[0]], zCoor[bndAtom_o[1]]], '0.5') 
		molView7.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		# Poly surface 1
		polyX1 = [xMax + 1.5, xMax + 1.5, xMax + 1.5, xMax + 1.5]
		polyY1 = [yMin - 1.5, yMax + 1.5, yMax + 1.5, yMin - 1.5]
		polyZ1 = [zMin - 1.5, zMin - 1.5, zMax + 1.5, zMax + 1.5]
		poly1 = list(zip(polyX1, polyY1, polyZ1))
		molView7.ax.add_collection3d(Poly3DCollection([poly1], facecolor = '#000080', alpha = 0.3))
		# Poly surface 2
		polyX2 = [xMax + 1.5, xMin - 1.5, xMin - 1.5, xMax + 1.5]
		polyY2 = [yMax + 1.5, yMax + 1.5, yMax + 1.5, yMax + 1.5]
		polyZ2 = [zMin - 1.5, zMin - 1.5, zMax + 1.5, zMax + 1.5]
		poly2 = list(zip(polyX2, polyY2, polyZ2))
		molView7.ax.add_collection3d(Poly3DCollection([poly2], facecolor = '#55559e', alpha = 0.3))
		# Poly surface 3
		polyX3 = [xMin - 1.5, xMin - 1.5, xMin - 1.5, xMin - 1.5]
		polyY3 = [yMin - 1.5, yMax + 1.5, yMax + 1.5, yMin - 1.5]
		polyZ3 = [zMin - 1.5, zMin - 1.5, zMax + 1.5, zMax + 1.5]
		poly3 = list(zip(polyX3, polyY3, polyZ3))
		molView7.ax.add_collection3d(Poly3DCollection([poly3], facecolor = '#000080', alpha = 0.3))
		# Poly surface 4
		polyX4 = [xMax + 1.5, xMin - 1.5, xMin - 1.5, xMax + 1.5]
		polyY4 = [yMin - 1.5, yMin - 1.5, yMin - 1.5, yMin - 1.5]
		polyZ4 = [zMin - 1.5, zMin - 1.5, zMax + 1.5, zMax + 1.5]
		poly4 = list(zip(polyX4, polyY4, polyZ4))
		molView7.ax.add_collection3d(Poly3DCollection([poly4], facecolor = '#55559e', alpha = 0.3))
		# Poly surface 5
		polyX5 = [xMax + 1.5, xMax + 1.5, xMin - 1.5, xMin - 1.5]
		polyY5 = [yMin - 1.5, yMax + 1.5, yMax + 1.5, yMin - 1.5]
		polyZ5 = [zMin - 1.5, zMin - 1.5, zMin - 1.5, zMin - 1.5]
		poly5 = list(zip(polyX5, polyY5, polyZ5))
		molView7.ax.add_collection3d(Poly3DCollection([poly5], facecolor = '#3939e3', alpha = 0.3))
		# Poly surface 6
		polyX6 = [xMax + 1.5, xMax + 1.5, xMin - 1.5, xMin - 1.5]
		polyY6 = [yMin - 1.5, yMax + 1.5, yMax + 1.5, yMin - 1.5]
		polyZ6 = [zMax + 1.5, zMax + 1.5, zMax + 1.5, zMax + 1.5]
		poly6 = list(zip(polyX6, polyY6, polyZ6))
		molView7.ax.add_collection3d(Poly3DCollection([poly6], facecolor = '#3939e3', alpha = 0.3))

		toolBar7 = NavigationToolbar(molView7, self)

		viewLayout7 = QVBoxLayout()
		viewLayout7.addWidget(molView7)
		viewLayout7.addWidget(toolBar7)
		viewLayout7.setContentsMargins(0, 0, 0, 0)

		viewer7 = QWidget()
		viewer7.setLayout(viewLayout7)

		n3DWidget = QWidget()
		n3DLayout = QVBoxLayout()
		n3DWidget.setLayout(n3DLayout)

		global n3InfoBox
		n3InfoBox = QTextEdit()
		n3InfoBox.append('3D NICS Information:\n-------------------------------------')
		n3DLayout.addWidget(n3InfoBox)
		n3InfoBox.append(f'\nX range: from {round(xMin - 1.5, 2)} to {round(xMax + 1.5, 2)}; Y range: from \
{round(yMin - 1.5, 2)} to {round(yMax + 1.5, 2)}; Z range: from \
{round(zMin - 1.5, 2)} to {round(zMax + 1.5, 2)}.')

		mm3Table = QGroupBox('3D NICS Parameters')
		mm3Layout = QGridLayout()
		mm3Table.setLayout(mm3Layout)

		mm3Layout.addWidget(QLabel('Minimum'), 1, 0)
		mm3Layout.addWidget(QLabel('Maximum'), 2, 0)
		mm3Layout.addWidget(QLabel('X'), 0, 1)
		mm3Layout.addWidget(QLabel('Y'), 0, 2)
		mm3Layout.addWidget(QLabel('Z'), 0, 3)

		global n3XMin, n3XMax, n3YMin, n3YMax, n3ZMin, n3ZMax
		n3XMin = QDoubleSpinBox()
		n3XMin.setRange(-999999.0, 999999.0)
		n3XMin.setSingleStep(0.1)
		n3XMin.setValue(round(xMin - 1.5, 2))
		mm3Layout.addWidget(n3XMin, 1, 1)
		n3XMax = QDoubleSpinBox()
		n3XMax.setRange(-999999.0, 999999.0)
		n3XMax.setSingleStep(0.1)
		n3XMax.setValue(round(xMax + 1.5, 2))
		mm3Layout.addWidget(n3XMax, 2, 1)
		n3YMin = QDoubleSpinBox()
		n3YMin.setRange(-999999.0, 999999.0)
		n3YMin.setSingleStep(0.1)
		n3YMin.setValue(round(yMin - 1.5, 2))
		mm3Layout.addWidget(n3YMin, 1, 2)
		n3YMax = QDoubleSpinBox()
		n3YMax.setRange(-999999.0, 999999.0)
		n3YMax.setSingleStep(0.1)
		n3YMax.setValue(round(yMax + 1.5, 2))
		mm3Layout.addWidget(n3YMax, 2, 2)
		n3ZMin = QDoubleSpinBox()
		n3ZMin.setRange(-999999.0, 999999.0)
		n3ZMin.setSingleStep(0.1)
		n3ZMin.setValue(round(zMin - 1.5, 2))
		mm3Layout.addWidget(n3ZMin, 1, 3)
		n3ZMax = QDoubleSpinBox()
		n3ZMax.setRange(-999999.0, 999999.0)
		n3ZMax.setSingleStep(0.1)
		n3ZMax.setValue(round(zMax + 1.5, 2))
		mm3Layout.addWidget(n3ZMax, 2, 3)
		n3DLayout.addWidget(mm3Table)

		mm3Layout.addWidget(QLabel(''), 3, 0)

		mm3Layout.addWidget(QLabel('Interval:'), 4, 0)

		global gridN3Spin
		gridN3Spin = QDoubleSpinBox()
		gridN3Spin.setRange(0.1, 999999.0)
		gridN3Spin.setSingleStep(0.1)
		gridN3Spin.setValue(0.2)
		mm3Layout.addWidget(gridN3Spin, 4, 1)

		N3BtnGroup = QWidget()
		N3BtnLayout = QHBoxLayout()
		N3BtnGroup.setLayout(N3BtnLayout)

		n3UpdateBtn = QPushButton('Preview')
		n3UpdateBtn.clicked.connect(self.reDraw3DGeom)
		N3BtnLayout.addWidget(n3UpdateBtn)

		self.calN3Setup_window = None
		n3CalSetBtn = QPushButton('Calculation Setup')
		n3CalSetBtn.clicked.connect(self.calN3Setup)
		N3BtnLayout.addWidget(n3CalSetBtn)

		n3InpBtn = QPushButton('Save Input File')
		n3InpBtn.clicked.connect(self.saveN3Inp)
		N3BtnLayout.addWidget(n3InpBtn)

		n3DLayout.addWidget(N3BtnGroup)

		nics3dMainLayout.addWidget(viewer7)
		nics3dMainLayout.addWidget(n3DWidget)
		nics3dMainWidget.setLayout(nics3dMainLayout)

		self.setCentralWidget(nics3dMainWidget)

	def reDraw3DGeom(self):
		molView7.ax.cla()
		molView7.ax.grid(False)
		molView7.ax.set_xlabel('X (Å)')
		molView7.ax.set_ylabel('Y (Å)')
		molView7.ax.set_zlabel('Z (Å)')

		if xMax == xMin:
			molView7.ax.set_xlim(xMin - 1, xMin + 1)
			molView7.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView7.ax.set_ylim(yMin - 1, yMin + 1)
			molView7.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView7.ax.set_zlim(zMin - 1, zMin + 1)
			molView7.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView7.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_o in bndAtom:
			molView7.ax.plot([xCoor[bndAtom_o[0]], xCoor[bndAtom_o[1]]], [yCoor[bndAtom_o[0]], yCoor[bndAtom_o[1]]], \
				[zCoor[bndAtom_o[0]], zCoor[bndAtom_o[1]]], '0.5') 
		molView7.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		# Poly surface 1
		polyX1 = [n3XMax.value(), n3XMax.value(), n3XMax.value(), n3XMax.value()]
		polyY1 = [n3YMin.value(), n3YMax.value(), n3YMax.value(), n3YMin.value()]
		polyZ1 = [n3ZMin.value(), n3ZMin.value(), n3ZMax.value(), n3ZMax.value()]
		poly1 = list(zip(polyX1, polyY1, polyZ1))
		molView7.ax.add_collection3d(Poly3DCollection([poly1], facecolor = '#000080', alpha = 0.3))
		# Poly surface 2
		polyX2 = [n3XMax.value(), n3XMin.value(), n3XMin.value(), n3XMax.value()]
		polyY2 = [n3YMax.value(), n3YMax.value(), n3YMax.value(), n3YMax.value()]
		polyZ2 = [n3ZMin.value(), n3ZMin.value(), n3ZMax.value(), n3ZMax.value()]
		poly2 = list(zip(polyX2, polyY2, polyZ2))
		molView7.ax.add_collection3d(Poly3DCollection([poly2], facecolor = '#55559e', alpha = 0.3))
		# Poly surface 3
		polyX3 = [n3XMin.value(), n3XMin.value(), n3XMin.value(), n3XMin.value()]
		polyY3 = [n3YMin.value(), n3YMax.value(), n3YMax.value(), n3YMin.value()]
		polyZ3 = [n3ZMin.value(), n3ZMin.value(), n3ZMax.value(), n3ZMax.value()]
		poly3 = list(zip(polyX3, polyY3, polyZ3))
		molView7.ax.add_collection3d(Poly3DCollection([poly3], facecolor = '#000080', alpha = 0.3))
		# Poly surface 4
		polyX4 = [n3XMax.value(), n3XMin.value(), n3XMin.value(), n3XMax.value()]
		polyY4 = [n3YMin.value(), n3YMin.value(), n3YMin.value(), n3YMin.value()]
		polyZ4 = [n3ZMin.value(), n3ZMin.value(), n3ZMax.value(), n3ZMax.value()]
		poly4 = list(zip(polyX4, polyY4, polyZ4))
		molView7.ax.add_collection3d(Poly3DCollection([poly4], facecolor = '#55559e', alpha = 0.3))
		# Poly surface 5
		polyX5 = [n3XMax.value(), n3XMax.value(), n3XMin.value(), n3XMin.value()]
		polyY5 = [n3YMin.value(), n3YMax.value(), n3YMax.value(), n3YMin.value()]
		polyZ5 = [n3ZMin.value(), n3ZMin.value(), n3ZMin.value(), n3ZMin.value()]
		poly5 = list(zip(polyX5, polyY5, polyZ5))
		molView7.ax.add_collection3d(Poly3DCollection([poly5], facecolor = '#3939e3', alpha = 0.3))
		# Poly surface 6
		polyX6 = [n3XMax.value(), n3XMax.value(), n3XMin.value(), n3XMin.value()]
		polyY6 = [n3YMin.value(), n3YMax.value(), n3YMax.value(), n3YMin.value()]
		polyZ6 = [n3ZMax.value(), n3ZMax.value(), n3ZMax.value(), n3ZMax.value()]
		poly6 = list(zip(polyX6, polyY6, polyZ6))
		molView7.ax.add_collection3d(Poly3DCollection([poly6], facecolor = '#3939e3', alpha = 0.3))

		molView7.draw()

		n3InfoBox.append(f'\nX range: from {round(n3XMin.value(), 2)} to {round(n3XMax.value(), 2)}; \
Y range: from {round(n3YMin.value(), 2)} to {round(n3YMax.value(), 2)}; Z range: from \
{round(n3ZMin.value(), 2)} to {round(n3ZMax.value(), 2)}.')

	def calN3Setup(self):
		if self.calN3Setup_window is None:
			self.calN3Setup_window = CalSetupWindow()
		self.calN3Setup_window.show()

	def saveN3Inp(self):
		routeN3Line = routeFunc(1, '_3D_NICS')
		bq3DList = NICSInp.nics3D(float(n3XMin.value()), float(n3XMax.value()), float(n3YMin.value()), float(n3YMax.value()), float(n3ZMin.value()), float(n3ZMax.value()), float(gridN3Spin.value()))

		fileNumbers = 1
		if len(bq3DList) <= 7000 - len(geomList):
			fileNumbers = 1
		elif len(bq3DList) % 7000 == 0:
			fileNumbers = int(len(bq3DList) / 7000)
		else:
			fileNumbers = int(len(bq3DList) / 7000 + 1)

		nics3DInp = open(f'{os.path.splitext(fileName)[0]}_3D_NICS_0001.gjf', 'w')
		if (fileType.lower() == 'gjf' or fileType.lower() == 'com') and configFile.get('general', 'input') == 'false':
			with open(fileName, 'r') as usr3DInp:
				usr3DInpLines = usr3DInp.readlines()
			for usr3DInpLine in usr3DInpLines:
				if usr3DInpLine != '\n' and len(usr3DInpLine.split()) >= 4 and usr3DInpLine.count('.') == 3:
					break
				else:
					if configFile.get('general', 'connectivity') == 'true' and usr3DInpLine[0] == '#' and 'geom=connectivity' not in usr3DInpLine:
						nics3DInp.write(usr3DInpLine.strip() + ' geom=connectivity' + '\n')
					else:
						nics3DInp.write(usr3DInpLine)
		else:
			for rou in routeN3Line:
				nics3DInp.write(rou + '\n')
		for geomRou in geomList:
			nics3DInp.write(f'{geomRou[0]}      {geomRou[1]:.6f}      {geomRou[2]:.6f}      {geomRou[3]:.6f}\n')

		if fileNumbers == 1:
			for bq3DRou in bq3DList:
				nics3DInp.write(bq3DRou)
			nics3DInp.write('\n')
			for atmNo3D in range(1, len(bq3DList) + len(geomList) + 1):
				nics3DInp.write(f'{atmNo3D}\n')
		else:
			bqCounter1 = 0
			for counter_i in range(7000):
				nics3DInp.write(bq3DList[counter_i])
				bqCounter1 += 1
			nics3DInp.write('\n')
			for bqNumber1 in range(1, 7000 + len(geomList) + 1):
				nics3DInp.write(f'{bqNumber1}\n')

		nics3DInp.write('\n\n')
		nics3DInp.close()

		if fileNumbers > 1:
			for fileNumber in range(2, fileNumbers + 1):
				fileNameNumber = '%04d' % fileNumber
				nics3DInp = open(f'{os.path.splitext(fileName)[0]}_3D_NICS_{fileNameNumber}.gjf', 'w')

				if (fileType.lower() == 'gjf' or fileType.lower() == 'com') and configFile.get('general', 'input') == 'false':
					with open(fileName, 'r') as usr3DInp:
						usr3DInpLines = usr3DInp.readlines()
					for usr3DInpLine in usr3DInpLines:
						if usr3DInpLine != '\n' and len(usr3DInpLine.split()) >= 4 and usr3DInpLine.count('.') == 3:
							break
						else:
							if configFile.get('general', 'connectivity') == 'true' and usr3DInpLine[0] == '#' and 'geom=connectivity' not in usr3DInpLine:
								nics3DInp.write(usr3DInpLine.strip() + ' geom=connectivity guess=read' + '\n')
							elif configFile.get('general', 'connectivity') == 'true' and usr3DInpLine[0] == '#' and 'geom=connectivity' in usr3DInpLine:
								nics3DInp.write(usr3DInpLine.strip() + ' guess=read' + '\n')
							elif configFile.get('general', 'connectivity') == 'false' and usr3DInpLine[0] == '#':
								nics3DInp.write(usr3DInpLine.strip() + ' guess=read' + '\n')
							else:
								nics3DInp.write(usr3DInpLine)

				else:
					for rou in routeN3Line:
						if rou != '':
							if rou[0] == '#':
								nics3DInp.write(rou + ' guess=read\n')
							else:
								nics3DInp.write(rou + '\n')
						else:
							nics3DInp.write(rou + '\n')
				for geomRou in geomList:
					nics3DInp.write(f'{geomRou[0]}      {geomRou[1]:.6f}      {geomRou[2]:.6f}      {geomRou[3]:.6f}\n')

				bqCounter2 = 0
				while bqCounter2 < 7000 and bqCounter1 < len(bq3DList):
					nics3DInp.write(bq3DList[bqCounter1])
					bqCounter1 += 1
					bqCounter2 += 1

				for bqNumber2 in range(1, len(geomList) + bqCounter2 + 1):
					nics3DInp.write('\n' + str(bqNumber2))

				nics3DInp.write('\n\n')
				nics3DInp.close()

'''
*************************************************************************

                               INICS WINDOW

*************************************************************************
'''
class InicsWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('INICS: ' + fileName.split('/')[-1])

		inicsMainWidget = QWidget()
		inicsMainLayout = QHBoxLayout()

		global molView19
		molView19 = Mpl3DCanvas()
		molView19.ax.axis('off')

		if xMax == xMin:
			molView19.ax.set_xlim(xMin - 1, xMin + 1)
			molView19.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView19.ax.set_ylim(yMin - 1, yMin + 1)
			molView19.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView19.ax.set_zlim(zMin - 1, zMin + 1)
			molView19.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView19.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_m in bndAtom:
			molView19.ax.plot([xCoor[bndAtom_m[0]], xCoor[bndAtom_m[1]]], [yCoor[bndAtom_m[0]], yCoor[bndAtom_m[1]]], \
				[zCoor[bndAtom_m[0]], zCoor[bndAtom_m[1]]], '0.5') 
		molView19.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		toolBar19 = NavigationToolbar(molView19, self)

		viewLayout19 = QVBoxLayout()
		viewLayout19.addWidget(molView19)
		viewLayout19.addWidget(toolBar19)
		viewLayout19.setContentsMargins(0, 0, 0, 0)

		viewer19 = QWidget()
		viewer19.setLayout(viewLayout19)

		inicsWidget = QWidget()
		inicsLayout = QVBoxLayout()
		inicsWidget.setLayout(inicsLayout)

		global inicsTextBox
		inicsTextBox = QTextEdit()
		inicsTextBox.setText('INICS informations:\n------------------------------')
		inicsLayout.addWidget(inicsTextBox)

		inicsParaGroup = QWidget()
		inicsParaLayout = QGridLayout()
		inicsParaGroup.setLayout(inicsParaLayout)

		global inicsAtmLine
		inicsAtmLine = QLineEdit()
		inicsAtmLine.setPlaceholderText('Enter atom No. here, e.g.: 1 2 3 4 5 6')
		inicsParaLayout.addWidget(inicsAtmLine, 1, 0, 1, 4)

		self.showInicsLabelBtn = QCheckBox('Show Atom No.')
		self.showInicsLabelBtn.setCheckState(Qt.CheckState.Unchecked)
		self.showInicsLabelBtn.stateChanged.connect(self.reDrawInicsGeom)
		inicsParaLayout.addWidget(self.showInicsLabelBtn, 3, 3)

		inicsParaLayout.addWidget(QLabel('   Range: '), 2, 0)

		global rangeBox
		rangeBox = QLineEdit()
		rangeBox.setText('10.0')
		inicsParaLayout.addWidget(rangeBox, 2, 1)

		inicsParaLayout.addWidget(QLabel('   Split: '), 2, 2)

		global splitBox
		splitBox = QLineEdit()
		splitBox.setText('0.2')
		inicsParaLayout.addWidget(splitBox, 2, 3)

		bqAddBtn = QPushButton('Add Bq atom')
		bqAddBtn.clicked.connect(self.addInicsFunc)
		inicsParaLayout.addWidget(bqAddBtn, 3, 1)

		inicsClrBtn = QPushButton('Clear')
		inicsClrBtn.clicked.connect(self.inicsClrFunc)
		inicsParaLayout.addWidget(inicsClrBtn, 3, 2)

		allBqAddBtn = QPushButton('Add Bq atoms for all monocycles')
		allBqAddBtn.clicked.connect(self.addAllBq)
		inicsParaLayout.addWidget(allBqAddBtn, 4, 0, 4, 4)

		inicsLayout.addWidget(inicsParaGroup)

		inicsBtnGroup = QWidget()
		inicsBtnLayout = QHBoxLayout()
		inicsBtnGroup.setLayout(inicsBtnLayout)

		self.inicsCalSetup_window = None
		inicsCalBtn = QPushButton('Calculation Setup')
		inicsCalBtn.clicked.connect(self.inicsCalSetup)
		inicsBtnLayout.addWidget(inicsCalBtn)

		inicsInpBtn = QPushButton('Save Input File')
		inicsInpBtn.clicked.connect(self.inicsInpSave)
		inicsBtnLayout.addWidget(inicsInpBtn)

		inicsLayout.addWidget(inicsBtnGroup)

		inicsMainLayout.addWidget(viewer19)
		inicsMainLayout.addWidget(inicsWidget)
		inicsMainWidget.setLayout(inicsMainLayout)

		self.setCentralWidget(inicsMainWidget)

		global inicsBqList
		inicsBqList = []

	def inicsCalSetup(self):
		if self.inicsCalSetup_window is None:
			self.inicsCalSetup_window = CalSetupWindow()
		self.inicsCalSetup_window.show()

	def inicsInpSave(self):
		inicsRouteLine = routeFunc(1, '_INICS')
		inicsInp = open(f'{os.path.splitext(fileName)[0]}_INICS.gjf', 'w')
		if (fileType.lower() == 'gjf' or fileType.lower() == 'com') and configFile.get('general', 'input') == 'false':
			with open(fileName, 'r') as usrInp:
				usrInpLines = usrInp.readlines()
			for usrInpLine in usrInpLines:
				if usrInpLine != '\n' and len(usrInpLine.split()) >= 4 and usrInpLine.count('.') == 3:
					break
				else:
					inicsInp.write(usrInpLine)
		else:
			for rou in inicsRouteLine:
				if 'File Created' in rou:
					ND = int(float(rangeBox.text())/float(splitBox.text()))
					inicsInp.write(f'INICS {float(rangeBox.text()):.2f} {float(splitBox.text()):.2f} {ND} // py.Aroma 4\n')
				else:
					inicsInp.write(rou + '\n')
		for geomRou in geomList:
			inicsInp.write(f'{geomRou[0]}      {geomRou[1]:.6f}      {geomRou[2]:.6f}      {geomRou[3]:.6f}\n')
		for bqInicsAtm in inicsBqList:
			inicsInp.write(f'Bq      {bqInicsAtm[1]:.6f}      {bqInicsAtm[2]:.6f}      {bqInicsAtm[3]:.6f}\n')
		inicsInp.write('\n')

		for atm_no in range(1, (len(geomList)+len(inicsBqList)+1)):
			inicsInp.write(f'{atm_no}\n')

		inicsInp.write('\n')
		inicsInp.close()

	def reDrawInicsGeom(self):
		molView19.ax.cla()
		molView19.ax.grid(False)
		molView19.ax.axis('off')

		currentList = []
		if len(inicsBqList) != 0:
			currentList = inicsBqList + geomList
		else:
			currentList = geomList
		
		xMin, xMax, yMin, yMax, zMin, zMax = geomAnalyzer.find_max_min(currentList)

		if xMax == xMin:
			molView19.ax.set_xlim(xMin - 1, xMin + 1)
			molView19.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView19.ax.set_ylim(yMin - 1, yMin + 1)
			molView19.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView19.ax.set_zlim(zMin - 1, zMin + 1)
			molView19.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView19.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_p in bndAtom:
			molView19.ax.plot([xCoor[bndAtom_p[0]], xCoor[bndAtom_p[1]]], [yCoor[bndAtom_p[0]], yCoor[bndAtom_p[1]]], \
				[zCoor[bndAtom_p[0]], zCoor[bndAtom_p[1]]], '0.5') 
		molView19.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		if len(inicsBqList) != 0:
			for bqAtom in inicsBqList:
				molView19.ax.scatter(np.array([bqAtom[1]]), np.array([bqAtom[2]]), np.array([bqAtom[3]]), color = [0.898, 0.2, 1.0], edgecolors = [0.898, 0.2, 1.0], s = 40, depthshade = 0)

		if self.showInicsLabelBtn.isChecked():
			for atom_k in range(len(geomList)):
				molView19.ax.text(geomList[atom_k][1], geomList[atom_k][2], geomList[atom_k][3], atom_k + 1, color = 'darkred')

		molView19.draw()

	def addInicsFunc(self):
		usrInicsAtm = str(inicsAtmLine.text())
		usrHight = float(rangeBox.text())
		usrSplit = float(splitBox.text())
		if usrHight < 0:
			usrHight = -usrHight
		if usrSplit < 0:
			usrSplit = -usrSplit
		
		heightList = [0]
		for inics_n in range(1, 999):
			if inics_n*usrSplit <= usrHight:
				heightList.append(float(inics_n*usrSplit))
			else:
				break

		if usrInicsAtm.strip() != '':
			usrInicsAtm2 = []
			for inicsAtm in usrInicsAtm.split():
				usrInicsAtm2.append(int(inicsAtm) - 1)

		for heightEle in heightList:
			if heightEle == 0.0:
				bq0x_u, bq0y_u, bq0z_u = NICSInp.calCoor(usrInicsAtm2, 0.0, geomList)
				inicsBqList.append(['Bq', bq0x_u, bq0y_u, bq0z_u])
			else:
				bq1x_u, bq1y_u, bq1z_u, bq2x_u, bq2y_u, bq2z_u = NICSInp.calCoor(usrInicsAtm2, heightEle, geomList)
				inicsBqList.append(['Bq', bq1x_u, bq1y_u, bq1z_u])
				inicsBqList.insert(0, ['Bq', bq2x_u, bq2y_u, bq2z_u])

		inicsTextBox.setText('INICS informations:\n------------------------------')
		inicsTextBox.append(f'Add Bq atoms for ring: [{str(inicsAtmLine.text())}] in range from {-usrHight} to {usrHight} with {usrSplit} interval.')

		inicsAtmLine.clear()
		self.reDrawInicsGeom()

	def inicsClrFunc(self):
		inicsBqList.clear()
		self.reDrawInicsGeom()
		inicsTextBox.setText('INICS informations:\n------------------------------')

	def addAllBq(self):
		global inicsMonoCycles
		inicsMonoCycles = geomAnalyzer.find_monocycle(bndAtom)
		
		usrHight = float(rangeBox.text())
		usrSplit = float(splitBox.text())
		if usrHight < 0:
			usrHight = -usrHight
		if usrSplit < 0:
			usrSplit = -usrSplit
		
		heightList2 = [0]
		for inics_m in range(1, 999):
			if inics_m*usrSplit <= usrHight:
				heightList2.append(float(inics_m*usrSplit))
			else:
				break

		for cpCycle in inicsMonoCycles:
			monobqList = []
			for heightEle2 in heightList2:
				if heightEle2 == 0.0:
					bq0x, bq0y, bq0z = NICSInp.calCoor(cpCycle, 0.0, geomList)
					monobqList.append(['Bq', bq0x, bq0y, bq0z])
				else:
					bq1x, bq1y, bq1z, bq2x, bq2y, bq2z = NICSInp.calCoor(cpCycle, heightEle2, geomList)
					monobqList.append(['Bq', bq1x, bq1y, bq1z])
					monobqList.insert(0, ['Bq', bq2x, bq2y, bq2z])
			for monobqListEle in monobqList:
				inicsBqList.append(monobqListEle)

		inicsTextBox.setText('INICS informations:\n------------------------------')
		if len(inicsMonoCycles) == 1:
			inicsTextBox.append(f'Add Bq atoms for 1 monocycle in range from {-usrHight} to {usrHight} with {usrSplit} interval.')
		else:
			inicsTextBox.append(f'Add Bq atoms for {len(inicsMonoCycles)} monocycles in range from {-usrHight} to {usrHight} with {usrSplit} interval.')
		self.reDrawInicsGeom()

'''
*************************************************************************

               TYPE-A WINDOW: for structure info, .gjf files

*************************************************************************
'''
class TypeAWindow(QMainWindow):
	def __init__(self):
		super().__init__()

		self.setWindowTitle(fileName.split('/')[-1])

		window_toolbar = QToolBar('Type A Toolbar')
		self.addToolBar(window_toolbar)
		window_toolbar.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)

		self.bla_window = None
		blaAction = QAction(QIcon('./assets/bla.png'), 'BLA', self)
		blaAction.setStatusTip('Plot bond length alternation graph for specified bonds.')
		blaAction.triggered.connect(self.blaFunction)

		self.homa_window = None
		homaAction = QAction(QIcon('./assets/homa.png'), 'HOMA', self)
		homaAction.setStatusTip('Compute HOMA value for monocycles.')
		homaAction.triggered.connect(self.homaFunction)

		self.pova_window = None
		povaAction = QAction(QIcon('./assets/poav.png'), 'POAV', self)
		povaAction.setStatusTip('Compute POAV1 and POAV2 for specified atom.')
		povaAction.triggered.connect(self.povaFunction)

		self.sp_window = None
		spAction = QAction(QIcon('./assets/sp.png'), 'NICS', self)
		spAction.setStatusTip('Create input file for single point NICS calculations.')
		spAction.triggered.connect(self.spFunction)

		self.scan_window = None
		nicsScanAction = QAction(QIcon('./assets/scan.png'), 'NICS Scan', self)
		nicsScanAction.setStatusTip('Create input file for 1D NICS scan calculations.')
		nicsScanAction.triggered.connect(self.scanFunction)

		self.nics2d_window = None
		nics2DAction = QAction(QIcon('./assets/2d.png'), '2D NICS', self)
		nics2DAction.setStatusTip('Create input file for 2D NICS (ICSS) calculations.')
		nics2DAction.triggered.connect(self.nics2DFunction)

		self.nics3d_window = None
		nics3DAction = QAction(QIcon('./assets/3d.png'), '3D NICS', self)
		nics3DAction.setStatusTip('Create input file for 3D NICS (ICSS) calculations.')
		nics3DAction.triggered.connect(self.nics3DFunction)

		self.inics_window = None
		inicsAction = QAction(QIcon('./assets/inics.png'), 'INICS', self)
		inicsAction.setStatusTip('Create input file for INICS calculations.')
		inicsAction.triggered.connect(self.inicsFunction)

		window_toolbar.addAction(blaAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(homaAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(povaAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(spAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(nicsScanAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(nics2DAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(nics3DAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(inicsAction)

		window_toolbar.setMovable(False)

		self.setStatusBar(QStatusBar(self))

		molView = Mpl3DCanvas()
		molView.ax.axis('off')

		if xMax == xMin:
			molView.ax.set_xlim(xMin - 1, xMin + 1)
			molView.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView.ax.set_ylim(yMin - 1, yMin + 1)
			molView.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView.ax.set_zlim(zMin - 1, zMin + 1)
			molView.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_i in bndAtom:
			molView.ax.plot([xCoor[bndAtom_i[0]], xCoor[bndAtom_i[1]]], [yCoor[bndAtom_i[0]], yCoor[bndAtom_i[1]]], \
				[zCoor[bndAtom_i[0]], zCoor[bndAtom_i[1]]], '0.5')
		molView.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		viewLayout = QVBoxLayout()
		viewLayout.addWidget(molView)

		viewer = QWidget()
		viewer.setLayout(viewLayout)

		self.setCentralWidget(viewer)

	def blaFunction(self):
		if self.bla_window is None:
			self.bla_window = BLAWindow()
		self.bla_window.show()

	def homaFunction(self):
		if self.homa_window is None:
			self.homa_window = HOMAWindow()
		self.homa_window.show()

	def povaFunction(self):
		if self.pova_window is None:
			self.pova_window = PovaWindow()
		self.pova_window.show()

	def spFunction(self):
		if self.sp_window is None:
			self.sp_window = SpWindow()
		self.sp_window.show()

	def scanFunction(self):
		if self.scan_window is None:
			self.scan_window = ScanWindow()
		self.scan_window.show()

	def nics2DFunction(self):
		if self.nics2d_window is None:
			self.nics2d_window = Nics2dWindow()
		self.nics2d_window.show()

	def nics3DFunction(self):
		if self.nics3d_window is None:
			self.nics3d_window = Nics3dWindow()
		self.nics3d_window.show()

	def inicsFunction(self):
		if self.inics_window is None:
			self.inics_window = InicsWindow()
		self.inics_window.show()
		
'''
*************************************************************************

                               NMR WINDOW

*************************************************************************
'''
class NmrSpecWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('NMR Spectrum: ' + fileName.split('/')[-1])

		nmrSpecMainWidget = QWidget()
		nmrSpecMainLayout = QVBoxLayout()

		global minRange, maxRange
		if len(plotPeak) == 1:
			minRange = plotPeak[0] - 3.0
			maxRange = plotPeak[0] + 3.0

		else:
			minRange = min(plotPeak) - (max(plotPeak) - min(plotPeak))*0.16
			maxRange = max(plotPeak) + (max(plotPeak) - min(plotPeak))*0.16
		
		nmrCsMat, nmrIntMat = pynmr.nmr_intensity(plotPeak, minRange, maxRange, 0.001, 0.01)

		global nmrView
		nmrView = MplNMRCanvas()

		nmrView.ax.plot(np.array(nmrCsMat), np.array(nmrIntMat), c = 'steelblue')
		nmrView.ax.set_xlabel('Chemical Shift (ppm)')
		nmrView.ax.tick_params(left=False, labelleft=False)
		nmrView.ax.set(ylabel=None)
		nmrView.ax.set_xlim(maxRange, minRange)
		figNmr.tight_layout()

		nmrToolbar = NavigationToolbar(nmrView, self)

		nmrViewGroup = QWidget()
		nmrViewLayout = QVBoxLayout()
		nmrViewGroup.setLayout(nmrViewLayout)
		nmrViewLayout.addWidget(nmrView)
		nmrViewLayout.addWidget(nmrToolbar)
		nmrViewLayout.setContentsMargins(0, 0, 0, 0)

		specSetMainWid = QWidget()
		specSetMainLayout = QHBoxLayout()
		specSetMainWid.setLayout(specSetMainLayout)

		specSettingGroup = QGroupBox('Spectrum Setting')
		specSettingLayout = QGridLayout()
		specSettingGroup.setLayout(specSettingLayout)

		specSettingLayout.addWidget(QLabel('Range: '), 0, 0)

		global minRangeBox, maxRangeBox
		minRangeBox = QLineEdit()
		minRangeBox.setText(f'{minRange:.4f}')
		minRangeBox.setFixedWidth(75)
		specSettingLayout.addWidget(minRangeBox, 0, 1)

		specSettingLayout.addWidget(QLabel('(ppm)   ~ '), 0, 2)

		maxRangeBox = QLineEdit()
		maxRangeBox.setText(f'{maxRange:.4f}')
		maxRangeBox.setFixedWidth(75)
		specSettingLayout.addWidget(maxRangeBox, 0, 3)

		specSettingLayout.addWidget(QLabel('(ppm)'), 0, 4)

		splitLabel = QLabel('Split: ')
		splitLabel.setAlignment(Qt.AlignmentFlag.AlignRight)
		specSettingLayout.addWidget(splitLabel, 1, 0)
		fwhmLabel = QLabel('FWHM: ')
		fwhmLabel.setAlignment(Qt.AlignmentFlag.AlignRight)
		specSettingLayout.addWidget(fwhmLabel, 1, 2)

		global specSplitBox
		specSplitBox = QDoubleSpinBox()
		specSplitBox.setRange(0.001, 100.0)
		specSplitBox.setDecimals(3)
		specSplitBox.setSingleStep(0.001)
		specSplitBox.setValue(0.001)
		specSettingLayout.addWidget(specSplitBox, 1, 1)

		global specFwhmBox
		specFwhmBox = QDoubleSpinBox()
		specFwhmBox.setRange(0.01, 1000.0)
		specFwhmBox.setSingleStep(0.01)
		specFwhmBox.setValue(0.01)
		specSettingLayout.addWidget(specFwhmBox, 1, 3)

		specPlotGroup = QWidget()
		specPlotLayout = QGridLayout()
		specPlotGroup.setLayout(specPlotLayout)

		redrawBtn = QPushButton('Redraw')
		redrawBtn.clicked.connect(self.redrawNMR)
		specPlotLayout.addWidget(redrawBtn, 0, 0)

		setDefBtn = QPushButton('Default')
		setDefBtn.clicked.connect(self.setDefNMR)
		specPlotLayout.addWidget(setDefBtn, 0, 1)

		pngNMRBtn = QPushButton('Save .png')
		pngNMRBtn.clicked.connect(self.savePngNMR)
		specPlotLayout.addWidget(pngNMRBtn, 1, 0)

		xlsxNMRBtn = QPushButton('Save .xlsx')
		xlsxNMRBtn.clicked.connect(self.saveXlsxNMR)
		specPlotLayout.addWidget(xlsxNMRBtn, 1, 1)

		specSetMainLayout.addWidget(specSettingGroup)
		specSetMainLayout.addWidget(specPlotGroup)
		nmrSpecMainLayout.addWidget(nmrViewGroup)
		nmrSpecMainLayout.addWidget(specSetMainWid)

		nmrSpecMainWidget.setLayout(nmrSpecMainLayout)

		self.setCentralWidget(nmrSpecMainWidget)

	def redrawNMR(self):
		nmrView.ax.cla()
		global nmrCsMat_2, nmrIntMat_2
		nmrCsMat_2, nmrIntMat_2 = pynmr.nmr_intensity(plotPeak, float(minRangeBox.text()), \
			float(maxRangeBox.text()), float(specSplitBox.text()), float(specFwhmBox.text()))
		nmrView.ax.plot(np.array(nmrCsMat_2), np.array(nmrIntMat_2), c = 'steelblue')
		nmrView.ax.set_xlabel('Chemical Shift (ppm)')
		nmrView.ax.tick_params(left=False, labelleft=False)
		nmrView.ax.set(ylabel=None)
		nmrView.ax.set_xlim(float(maxRangeBox.text()), float(minRangeBox.text()))
		figNmr.tight_layout()
		nmrView.draw()

	def setDefNMR(self):
		minRangeBox.setText(f'{minRange:.4f}')
		maxRangeBox.setText(f'{maxRange:.4f}')
		specSplitBox.setValue(0.001)
		specFwhmBox.setValue(0.01)

	def savePngNMR(self):
		figNmr.savefig(f'{os.path.splitext(fileName)[0]}_{eleCombo.currentText().title()}_NMR.png', dpi = 300)

	def saveXlsxNMR(self):
		self.redrawNMR()
		nmrWB = openpyxl.Workbook()
		nmrWS = nmrWB.active
		nmrWS['A1'] = 'Chemical Shift (ppm)'
		nmrWS['B1'] = 'Intensity'
		nmrDatas = []
		for nmr_i in range(len(nmrCsMat_2)):
			nmrData = []
			nmrData.append(nmrCsMat_2[nmr_i])
			nmrData.append(nmrIntMat_2[nmr_i])
			nmrDatas.append(nmrData)
		for nmrData_i in nmrDatas:
			nmrWS.append(nmrData_i)
		nmrWB.save(f'{os.path.splitext(fileName)[0]}_{eleCombo.currentText().title()}_NMR.xlsx')

class NmrWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('NMR: ' + fileName.split('/')[-1])

		nmrMainWidget = QWidget()
		nmrMainLayout = QHBoxLayout()

		molView25 = Mpl3DCanvas()
		molView25.ax.axis('off')

		if xMax == xMin:
			molView25.ax.set_xlim(xMin - 1, xMin + 1)
			molView25.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView25.ax.set_ylim(yMin - 1, yMin + 1)
			molView25.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView25.ax.set_zlim(zMin - 1, zMin + 1)
			molView25.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView25.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_jj in bndAtom:
			molView25.ax.plot([xCoor[bndAtom_jj[0]], xCoor[bndAtom_jj[1]]], [yCoor[bndAtom_jj[0]], yCoor[bndAtom_jj[1]]], \
				[zCoor[bndAtom_jj[0]], zCoor[bndAtom_jj[1]]], '0.5')
		for atom_ii in range(len(geomList)):
			molView25.ax.text(geomList[atom_ii][1], geomList[atom_ii][2], geomList[atom_ii][3], atom_ii + 1, color = 'darkred')
		
		molView25.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		toolBar25 = NavigationToolbar(molView25, self)

		viewLayout25 = QVBoxLayout()
		viewLayout25.addWidget(molView25)
		viewLayout25.addWidget(toolBar25)
		viewLayout25.setContentsMargins(0, 0, 0, 0)

		viewer25 = QWidget()
		viewer25.setLayout(viewLayout25)

		nmrResult = QWidget()
		nmrReLayout = QVBoxLayout()
		nmrResult.setLayout(nmrReLayout)

		global nmrShield
		nmrShield = QTextEdit()
		nmrReLayout.addWidget(nmrShield)
		nmrShield.setText('    Atom       Shielding (ppm)\n------------------------------')

		global nmrMat
		nmrMat = readFile.nmr_isotropy(fileName)

		eleList = []
		for shiel_j in nmrMat:
			eleList.append(shiel_j[1])
		eleSetList = list(set(eleList))

		if 'h' in eleSetList:
			for shiel_i in nmrMat:
				if shiel_i[1].lower() == 'h':
					nmrShield.append(f"  {format((shiel_i[1].title()+shiel_i[0]).rjust(6))}         {format(str(round(shiel_i[2], 4)).rjust(9))}")
		else:
			for shiel_k in nmrMat:
				if shiel_k[1].lower() == eleSetList[0].lower():
					nmrShield.append(f"  {format((shiel_k[1].title()+shiel_k[0]).rjust(6))}         {format(str(round(shiel_k[2], 4)).rjust(9))}")

		nmrSetting = QWidget()
		nmrSetLayout = QVBoxLayout()
		nmrSetting.setLayout(nmrSetLayout)

		nmrSetLayout.addWidget(nmrResult)

		modeBtnGroup = QGroupBox('Select Mode')
		modeBtnLayout = QGridLayout()
		modeBtnGroup.setLayout(modeBtnLayout)

		modeGroup = QButtonGroup()
		nothingBtn = QRadioButton('Show shielding tensors', self)
		nothingBtn.toggled.connect(self.modeClicked)
		modeBtnLayout.addWidget(nothingBtn, 1, 0, 1, 1)

		refBtn = QRadioButton('Set reference (ppm): ', self)
		refBtn.toggled.connect(self.modeClicked)
		modeBtnLayout.addWidget(refBtn, 2, 0)

		global tmsRef
		tmsRef = QLineEdit()
		tmsRef.setText('31.8821')
		tmsRef.setFixedWidth(75)
		modeBtnLayout.addWidget(tmsRef, 2, 1)

		scaleBtn = QRadioButton('Apply scaling factor: ', self)
		scaleBtn.toggled.connect(self.modeClicked)
		modeBtnLayout.addWidget(scaleBtn, 3, 0)

		CHESHIREBtn = QPushButton('CHESHIRE CCAT')
		CHESHIREBtn.clicked.connect(self.cheshirenmr)
		modeBtnLayout.addWidget(CHESHIREBtn, 3, 1, 3, 2)

		global slopeText, interText
		slopeText = QLineEdit()
		slopeText.setText('-0.9957')
		slopeText.setFixedWidth(75)
		slopeLabel = QLabel('Slope: ')
		slopeLabel.setAlignment(Qt.AlignmentFlag.AlignRight)
		modeBtnLayout.addWidget(slopeLabel, 6, 0)
		modeBtnLayout.addWidget(slopeText, 6, 1)
		interText = QLineEdit()
		interText.setText('32.2884')
		interText.setFixedWidth(75)
		inteLabel = QLabel('Intercept: ')
		inteLabel.setAlignment(Qt.AlignmentFlag.AlignRight)
		modeBtnLayout.addWidget(inteLabel, 6, 2)
		modeBtnLayout.addWidget(interText, 6, 3)

		modeGroup.addButton(nothingBtn, 1)
		modeGroup.addButton(refBtn, 2)
		modeGroup.addButton(scaleBtn, 3)

		nothingBtn.setChecked(True)

		nmrSetLayout.addWidget(modeBtnGroup)

		nmrPlotting = QWidget()
		nmrPlotLayout = QHBoxLayout()
		nmrPlotting.setLayout(nmrPlotLayout)

		eleLabel = QLabel('Elements: ')
		eleLabel.setAlignment(Qt.AlignmentFlag.AlignRight)
		nmrPlotLayout.addWidget(eleLabel)

		global eleCombo
		eleCombo = QComboBox()
		eleCombo.addItems([eleII.title() for eleII in eleSetList])
		eleCombo.setEditable(False)
		if 'h' in eleSetList:
			eleCombo.setCurrentText('H')
		else:
			eleCombo.setCurrentText(eleSetList[0].title())
		nmrPlotLayout.addWidget(eleCombo)

		clacBtn = QPushButton('Calculate')
		clacBtn.clicked.connect(self.calcShieldFunc)
		nmrPlotLayout.addWidget(clacBtn)

		spectraBtn = QPushButton('Show Spectrum')
		spectraBtn.clicked.connect(self.showSpectraFunc)
		nmrPlotLayout.addWidget(spectraBtn)

		nmrSetLayout.addWidget(nmrPlotting)

		nmrMainLayout.addWidget(viewer25)
		nmrMainLayout.addWidget(nmrSetting)
		nmrMainWidget.setLayout(nmrMainLayout)

		self.setCentralWidget(nmrMainWidget)

		self.nmr_spec_window = None

	def cheshirenmr(self):
		webbrowser.open('http://cheshirenmr.info/ScalingFactors.htm')

	def calcShieldFunc(self):
		global plotPeak
		plotPeak = []

		global refMat, scaleMat
		refMat = []
		scaleMat = []

		if currentMode == 'Show shielding tensors':
			nmrShield.setText('    Atom       Shielding (ppm)\n------------------------------')
			for shiel_ii in nmrMat:
				if shiel_ii[1].lower() == eleCombo.currentText().lower():
					nmrShield.append(f"  {format((shiel_ii[1].title()+shiel_ii[0]).rjust(6))}         {format(str(round(shiel_ii[2], 4)).rjust(9))}")
			plotPeak = pynmr.plot_peak(eleCombo.currentText(), nmrMat)

		elif currentMode == 'Set reference (ppm): ':
			refMat = pynmr.ref_peak(nmrMat, float(tmsRef.text()))
			nmrShield.setText('    Atom     Chemical Shift (ppm)\n---------------------------------')
			for shiel_jj in refMat:
				if shiel_jj[1].lower() == eleCombo.currentText().lower():
					nmrShield.append(f"  {format((shiel_jj[1].title()+shiel_jj[0]).rjust(6))}         {format(str(round(shiel_jj[2], 4)).rjust(9))}")
			plotPeak = pynmr.plot_peak(eleCombo.currentText(), refMat)

		elif currentMode == 'Apply scaling factor: ':
			scaleMat = pynmr.scale_peak(nmrMat, float(slopeText.text()), float(interText.text()))
			nmrShield.setText('    Atom     Chemical Shift (ppm)\n---------------------------------')
			for shiel_ll in scaleMat:
				if shiel_ll[1].lower() == eleCombo.currentText().lower():
					nmrShield.append(f"  {format((shiel_ll[1].title()+shiel_ll[0]).rjust(6))}         {format(str(round(shiel_ll[2], 4)).rjust(9))}")
			plotPeak = pynmr.plot_peak(eleCombo.currentText(), scaleMat)

	def showSpectraFunc(self):
		self.calcShieldFunc()
		self.nmr_spec_window = NmrSpecWindow()
		self.nmr_spec_window.show()

	def modeClicked(self):
		global currentMode
		currentMode = ''
		if self.sender().isChecked():
			currentMode = self.sender().text()
		if currentMode == 'Show shielding tensors':
			tmsRef.setEnabled(False)
			slopeText.setEnabled(False)
			interText.setEnabled(False)
		elif currentMode == 'Set reference (ppm): ':
			tmsRef.setEnabled(True)
			slopeText.setEnabled(False)
			interText.setEnabled(False)
		elif currentMode == 'Apply scaling factor: ':
			tmsRef.setEnabled(False)
			slopeText.setEnabled(True)
			interText.setEnabled(True)

'''
*************************************************************************

                                CSI WINDOW

*************************************************************************
'''
class CsiWindow(QMainWindow):
	def __init__(self):
		super().__init__()

		self.setWindowTitle('CSIgen: ' + fileName.split('/')[-1])

		csiMainWidget = QWidget()
		csiMainLayout = QHBoxLayout()
		csiMainWidget.setLayout(csiMainLayout)

		global molView15
		molView15 = Mpl3DCanvas()
		molView15.ax.axis('off')

		if xMax == xMin:
			molView15.ax.set_xlim(xMin - 1, xMin + 1)
			molView15.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView15.ax.set_ylim(yMin - 1, yMin + 1)
			molView15.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView15.ax.set_zlim(zMin - 1, zMin + 1)
			molView15.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView15.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_o in bndAtom:
			molView15.ax.plot([xCoor[bndAtom_o[0]], xCoor[bndAtom_o[1]]], [yCoor[bndAtom_o[0]], yCoor[bndAtom_o[1]]], \
				[zCoor[bndAtom_o[0]], zCoor[bndAtom_o[1]]], '0.5') 
		molView15.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		toolBar16 = NavigationToolbar(molView15, self)

		viewWidget = QWidget()
		viewLayout15 = QVBoxLayout()
		viewWidget.setLayout(viewLayout15)
		viewLayout15.addWidget(molView15)
		viewLayout15.addWidget(toolBar16)
		viewLayout15.setContentsMargins(0, 0, 0, 0)
		csiMainLayout.addWidget(viewWidget)

		optionMainWideget = QWidget()
		optionMainLayout = QVBoxLayout()
		optionMainWideget.setLayout(optionMainLayout)

		with open(fileName, 'r') as outPutFile:
			outLine = outPutFile.readlines()

		global outRouteLine
		outRouteLine = ''
		for out_i in range(len(outLine)):
			if '------' in outLine[out_i] and '#' in outLine[out_i + 1]:
				outRouteLine = outLine[out_i + 1].strip().lower()
				if '------' not in outLine[out_i + 2]:
					outRouteLine = outLine[out_i + 1].strip().lower() + ' ' + outLine[out_i + 2].strip().lower()
				break

		optionWidget = QGroupBox('Option')
		optionLayout = QVBoxLayout()
		optionWidget.setLayout(optionLayout)

		global rotChkBox, chargeChkBox, spinChkBox, pgChkBox
		rotChkBox = QCheckBox('Routine Section')
		rotChkBox.setCheckState(Qt.CheckState.Checked)

		chargeChkBox = QCheckBox('Charge')
		chargeChkBox.setCheckState(Qt.CheckState.Checked)
		spinChkBox = QCheckBox('Spin Multiplicity')
		spinChkBox.setCheckState(Qt.CheckState.Checked)
		pgChkBox = QCheckBox('Point Group')
		pgChkBox.setCheckState(Qt.CheckState.Checked)

		global eleEnChkBox, nifChkBox, zpeChkBox, thrChkBox, hChkBox, feChkBox
		eleEnChkBox = QCheckBox('Electronic Energy')
		eleEnChkBox.setCheckState(Qt.CheckState.Checked)

		nifChkBox = QCheckBox('Number of Imaginary Frequencies')
		zpeChkBox = QCheckBox('Zero-Point Energy')
		thrChkBox = QCheckBox('Thermal Energy')
		hChkBox = QCheckBox('Enthalpy')
		feChkBox = QCheckBox('Free Energy')

		if 'freq' in outRouteLine:
			nifChkBox.setCheckState(Qt.CheckState.Checked)
			zpeChkBox.setCheckState(Qt.CheckState.Checked)
			thrChkBox.setCheckState(Qt.CheckState.Checked)
			hChkBox.setCheckState(Qt.CheckState.Checked)
			feChkBox.setCheckState(Qt.CheckState.Checked)
		else:
			nifChkBox.setCheckable(False)
			zpeChkBox.setCheckable(False)
			thrChkBox.setCheckable(False)
			hChkBox.setCheckable(False)
			feChkBox.setCheckable(False)

		optionLayout.addWidget(rotChkBox)
		optionLayout.addWidget(chargeChkBox)
		optionLayout.addWidget(spinChkBox)
		optionLayout.addWidget(pgChkBox)
		optionLayout.addWidget(eleEnChkBox)
		optionLayout.addWidget(nifChkBox)
		optionLayout.addWidget(zpeChkBox)
		optionLayout.addWidget(thrChkBox)
		optionLayout.addWidget(hChkBox)
		optionLayout.addWidget(feChkBox)

		global xlscColChkBox
		xlscColChkBox = QCheckBox('Save coordinates in one column.')
		xlscColChkBox.setCheckState(Qt.CheckState.Unchecked)

		optionMainLayout.addWidget(optionWidget)

		optionWidget2 = QWidget()
		optionLayout2 = QVBoxLayout()
		optionWidget2.setLayout(optionLayout2)

		chkAllCsiBtn = QPushButton('Select All')
		chkAllCsiBtn.clicked.connect(self.chkAllCsi)
		optionLayout.addWidget(chkAllCsiBtn)

		unchkAllCsiBtn = QPushButton('Unselect All')
		unchkAllCsiBtn.clicked.connect(self.unchkAllCsi)
		optionLayout.addWidget(unchkAllCsiBtn)

		chkInvCsiBtn = QPushButton('Select Inverse')
		chkInvCsiBtn.clicked.connect(self.chkInvCsi)
		optionLayout.addWidget(chkInvCsiBtn)

		optionLayout.addWidget(QLabel('  '))
		optionLayout.addWidget(QLabel('For .xlsx file: '))
		optionLayout.addWidget(xlscColChkBox)

		saveTxtCsiBtn = QPushButton('Save .txt')
		saveTxtCsiBtn.clicked.connect(self.csiSaveTxt)
		optionLayout2.addWidget(saveTxtCsiBtn)

		saveXlsxCsiBtn = QPushButton('Save .xlsx')
		saveXlsxCsiBtn.clicked.connect(self.csiSaveXlsx)
		optionLayout2.addWidget(saveXlsxCsiBtn)

		infoText = QTextEdit()
		infoText.setDisabled(True)
		optionLayout2.addWidget(infoText)

		optionMainLayout.addWidget(optionWidget2)

		csiMainLayout.addWidget(optionMainWideget)

		self.setCentralWidget(csiMainWidget)

		global freqZPE, freqThr, freqH, freqFE
		freqZPE, freqThr, freqH, freqFE = 999, 999, 999, 999

		global imFreq, freqValue
		imFreq = 0
		freqValue = []

		global outCharge, outMultiplicity, outPointGroup
		outCharge = 999
		outMultiplicity = 999
		outPointGroup = '999'

		for out_i in range(len(outLine)):
			if 'Charge =' in outLine[out_i] and 'Multiplicity =' in outLine[out_i]:
				outCharge = int(outLine[out_i].split()[2])
				outMultiplicity = int(outLine[out_i].split()[5])
			if 'Full point group' in outLine[out_i]:
				outPointGroup = outLine[out_i].split()[3]
			if 'freq' in outRouteLine:
				if 'Sum of electronic and zero-point Energies=' in outLine[out_i]:
					freqZPE = format(float(outLine[out_i].split()[6]), '.6f')
					freqThr = format(float(outLine[out_i + 1].split()[6]), '.6f')
					freqH = format(float(outLine[out_i + 2].split()[6]), '.6f')
					freqFE = format(float(outLine[out_i + 3].split()[7]), '.6f')
				if 'Frequencies -- ' in outLine[out_i]:
					freqValue.append(format(float(outLine[out_i].split()[2]), '.2f'))
					freqValue.append(format(float(outLine[out_i].split()[3]), '.2f'))
					freqValue.append(format(float(outLine[out_i].split()[4]), '.2f'))

		if len(freqValue) != 0:
			for freq_i in freqValue:
				if float(freq_i) < 0:
					imFreq += 1

		global eleEn
		eleEn = readFile.read_energy(outRouteLine, fileName)

	def chkAllCsi(self):
		rotChkBox.setCheckState(Qt.CheckState.Checked)
		chargeChkBox.setCheckState(Qt.CheckState.Checked)
		spinChkBox.setCheckState(Qt.CheckState.Checked)
		pgChkBox.setCheckState(Qt.CheckState.Checked)
		eleEnChkBox.setCheckState(Qt.CheckState.Checked)
		if 'freq' in outRouteLine:
			nifChkBox.setCheckState(Qt.CheckState.Checked)
			zpeChkBox.setCheckState(Qt.CheckState.Checked)
			thrChkBox.setCheckState(Qt.CheckState.Checked)
			hChkBox.setCheckState(Qt.CheckState.Checked)
			feChkBox.setCheckState(Qt.CheckState.Checked)
		else:
			nifChkBox.setCheckable(False)
			zpeChkBox.setCheckable(False)
			thrChkBox.setCheckable(False)
			hChkBox.setCheckable(False)
			feChkBox.setCheckable(False)

	def unchkAllCsi(self):
		rotChkBox.setCheckState(Qt.CheckState.Unchecked)
		chargeChkBox.setCheckState(Qt.CheckState.Unchecked)
		spinChkBox.setCheckState(Qt.CheckState.Unchecked)
		pgChkBox.setCheckState(Qt.CheckState.Unchecked)
		eleEnChkBox.setCheckState(Qt.CheckState.Unchecked)
		if 'freq' in outRouteLine:
			nifChkBox.setCheckState(Qt.CheckState.Unchecked)
			zpeChkBox.setCheckState(Qt.CheckState.Unchecked)
			thrChkBox.setCheckState(Qt.CheckState.Unchecked)
			hChkBox.setCheckState(Qt.CheckState.Unchecked)
			feChkBox.setCheckState(Qt.CheckState.Unchecked)
		else:
			nifChkBox.setCheckable(False)
			zpeChkBox.setCheckable(False)
			thrChkBox.setCheckable(False)
			hChkBox.setCheckable(False)
			feChkBox.setCheckable(False)

	def chkInvCsi(self):
		if rotChkBox.isChecked():
			rotChkBox.setCheckState(Qt.CheckState.Unchecked)
		else:
			rotChkBox.setCheckState(Qt.CheckState.Checked)
		if chargeChkBox.isChecked():
			chargeChkBox.setCheckState(Qt.CheckState.Unchecked)
		else:
			chargeChkBox.setCheckState(Qt.CheckState.Checked)
		if spinChkBox.isChecked():
			spinChkBox.setCheckState(Qt.CheckState.Unchecked)
		else:
			spinChkBox.setCheckState(Qt.CheckState.Checked)
		if pgChkBox.isChecked():
			pgChkBox.setCheckState(Qt.CheckState.Unchecked)
		else:
			pgChkBox.setCheckState(Qt.CheckState.Checked)
		if eleEnChkBox.isChecked():
			eleEnChkBox.setCheckState(Qt.CheckState.Unchecked)
		else:
			eleEnChkBox.setCheckState(Qt.CheckState.Checked)
		if 'freq' in outRouteLine:
			if nifChkBox.isChecked():
				nifChkBox.setCheckState(Qt.CheckState.Unchecked)
			else:
				nifChkBox.setCheckState(Qt.CheckState.Checked)
			if zpeChkBox.isChecked():
				zpeChkBox.setCheckState(Qt.CheckState.Unchecked)
			else:
				zpeChkBox.setCheckState(Qt.CheckState.Checked)
			if thrChkBox.isChecked():
				thrChkBox.setCheckState(Qt.CheckState.Unchecked)
			else:
				thrChkBox.setCheckState(Qt.CheckState.Checked)
			if hChkBox.isChecked():
				hChkBox.setCheckState(Qt.CheckState.Unchecked)
			else:
				hChkBox.setCheckState(Qt.CheckState.Checked)
			if feChkBox.isChecked():
				feChkBox.setCheckState(Qt.CheckState.Unchecked)
			else:
				feChkBox.setCheckState(Qt.CheckState.Checked)
		else:
			nifChkBox.setCheckable(False)
			zpeChkBox.setCheckable(False)
			thrChkBox.setCheckable(False)
			hChkBox.setCheckable(False)
			feChkBox.setCheckable(False)

	def csiSaveTxt(self):
		csiTxt = open(f'{os.path.splitext(fileName)[0]}_CSI.txt', 'w')
		csiTxt.write(fileName.split('/')[-1].split('.')[0])
		csiTxt.write('\n===================================================')

		if rotChkBox.isChecked():
			csiTxt.write(f'\n{outRouteLine}')
		if chargeChkBox.isChecked():
			csiTxt.write(f'\nCharge = {outCharge}')
		if spinChkBox.isChecked():
			csiTxt.write(f'\nMultiplicity = {outMultiplicity}')
		if pgChkBox.isChecked():
			csiTxt.write(f'\nPoint Group = {outPointGroup}')
		if nifChkBox.isChecked():
			csiTxt.write(f'\nNumber of imaginary frequencies = {imFreq}')
			if imFreq != 0:
				csiTxt.write(f', vi = {freqValue[0]}')
		if eleEnChkBox.isChecked():
			csiTxt.write(f'\nElectronic Energy = {eleEn} Hartree')
		if zpeChkBox.isChecked():
			csiTxt.write(f'\nSum of electronic and zero-point Energies = {freqZPE} Hartree')
		if thrChkBox.isChecked():
			csiTxt.write(f'\nSum of electronic and thermal Energies = {freqThr} Hartree')
		if hChkBox.isChecked():
			csiTxt.write(f'\nSum of electronic and thermal Enthalpies = {freqH} Hartree')
		if feChkBox.isChecked():
			csiTxt.write(f'\nSum of electronic and thermal Free Energies = {freqFE} Hartree')

		csiTxt.write('\n---------------------------------------------------')
		csiTxt.write('\n                      Coordinates (Å)')
		csiTxt.write('\n Atoms        X              Y              Z')
		csiTxt.write('\n---------------------------------------------------\n')
		for ele_i in geomList:
			if len(ele_i[0]) == 1:
				csiTxt.write(f'   {ele_i[0]}')
			elif len(ele_i[0]) == 2:
				csiTxt.write(f'  {ele_i[0]}')
			elif len(ele_i[0]) == 3:
				csiTxt.write(f' {ele_i[0]}')
			spX = ''
			for sp_no in range(15 - len(format(ele_i[1], '.6f'))):
				spX += ' '
			csiTxt.write(spX + format(ele_i[1], '.6f'))
			spY = ''
			for sp_no_y in range(15 - len(format(ele_i[2], '.6f'))):
				spY += ' '
			csiTxt.write(spY + format(ele_i[2], '.6f'))
			spZ = ''
			for sp_no_z in range(15 - len(format(ele_i[3], '.6f'))):
				spZ += ' '
			csiTxt.write(spZ + format(ele_i[3], '.6f') + '\n')
		csiTxt.write('---------------------------------------------------')

		csiTxt.write('\n\n')
		csiTxt.close()

	def csiSaveXlsx(self):
		if xlscColChkBox.isChecked():
			outWB = openpyxl.Workbook()
			outWS = outWB.active
			mediumB = Side(border_style = 'medium', color = '000000')
			thinB = Side(border_style = 'thin', color = '000000')
			dashB = Side(border_style = 'dashed', color = '000000')

			outWS.column_dimensions['A'].width = 8.25
			outWS.column_dimensions['B'].width = 12.25
			outWS.column_dimensions['C'].width = 12.25
			outWS.column_dimensions['D'].width = 12.25

			cellRange = outWS['A2:D9']
			for cellNum in cellRange:
				for cellNum1 in cellNum:
					cellNum1.font = Font(name = 'Times New Roman', size = 10)

			for w in range(1, 8000):
				outWS.row_dimensions[w].height = 16.0

			cellRange3 = outWS['A1:D1']
			for cellNum3 in cellRange3:
				for cellNum4 in cellNum3:
					cellNum4.border = Border(bottom = mediumB)

			cellRange5 = outWS['A9:D9']
			for cellNum5 in cellRange5:
				for cellNum6 in cellNum5:
					cellNum6.border = Border(bottom = thinB)

			outWS.merge_cells('A1:D1')
			outWS['A1'] = fileName.split('/')[-1].split('.')[0]
			outWS['A1'].font = Font(name = 'Times New Roman', size = 10.5, bold = True)
			outWS['A1'].alignment = Alignment(horizontal = 'center', vertical = 'center')

			outWS.merge_cells('A2:D2')
			if rotChkBox.isChecked():
				outWS['A2'].font = Font(name = 'Courier', size = 10)
			outWS.merge_cells('A3:D3')
			outWS.merge_cells('A4:D4')
			outWS.merge_cells('A5:D5')
			outWS.merge_cells('A6:D6')
			outWS.merge_cells('A7:D7')
			outWS.merge_cells('A8:D8')
			outWS.merge_cells('A9:D9')

			optionSec = []
			if rotChkBox.isChecked():
				optionSec.append(outRouteLine)
			if chargeChkBox.isChecked() and spinChkBox.isChecked() and pgChkBox.isChecked():
				optionSec.append(f'Charge = {outCharge}, Multiplicity = {outMultiplicity}, Point group = {outPointGroup}')
			elif chargeChkBox.isChecked() == False and spinChkBox.isChecked() and pgChkBox.isChecked():
				optionSec.append(f'Multiplicity = {outMultiplicity}, Point group = {outPointGroup}')
			elif chargeChkBox.isChecked() and spinChkBox.isChecked() == False and pgChkBox.isChecked():
				optionSec.append(f'Charge = {outCharge}, Point group = {outPointGroup}')
			elif chargeChkBox.isChecked() and spinChkBox.isChecked() and pgChkBox.isChecked() == False:
				optionSec.append(f'Charge = {outCharge}, Multiplicity = {outMultiplicity}')
			elif chargeChkBox.isChecked() and spinChkBox.isChecked() == False and pgChkBox.isChecked() == False:
				optionSec.append(f'Charge = {outCharge}')
			elif chargeChkBox.isChecked() == False and spinChkBox.isChecked() and pgChkBox.isChecked() == False:
				optionSec.append(f'Multiplicity = {outMultiplicity}')
			elif chargeChkBox.isChecked() == False and spinChkBox.isChecked() == False and pgChkBox.isChecked():
				optionSec.append(f'Point group = {outPointGroup}')

			if nifChkBox.isChecked():
				if imFreq != 0:
					optionSec.append(f'Number of imaginary frequencies = {imFreq}, vi = {freqValue[0]}')
				else:
					optionSec.append(f'Number of imaginary frequencies = {imFreq}')
			if eleEnChkBox.isChecked():
				optionSec.append(f'Electronic Energy = {eleEn} Hartree')
			if zpeChkBox.isChecked():
				optionSec.append(f'Sum of electronic and zero-point Energies = {freqZPE} Hartree')
			if thrChkBox.isChecked():
				optionSec.append(f'Sum of electronic and thermal Energies = {freqThr} Hartree')
			if hChkBox.isChecked():
				optionSec.append(f'Sum of electronic and thermal Enthalpies = {freqH} Hartree')
			if feChkBox.isChecked():
				optionSec.append(f'Sum of electronic and thermal Free Energies = {freqFE} Hartree')

			if len(optionSec) != 0:
				for wb_i in range(len(optionSec)):
					outWS[f'A{2+wb_i}'] = optionSec[wb_i]

			outWS.merge_cells('A10:A11')
			outWS.merge_cells('B10:D10')

			outWS['A10'] = 'Atoms'
			outWS['A10'].font = Font(name = 'Times New Roman', size = 10, bold = True)
			outWS['A10'].alignment = Alignment(horizontal = 'center', vertical = 'bottom')
			outWS['A11'].border = Border(bottom = thinB)

			outWS['B10'] = 'Cartesian Coordinates'
			outWS['B10'].font = Font(name = 'Times New Roman', size = 10, bold = True)
			outWS['B10'].alignment = Alignment(horizontal = 'center', vertical = 'center')

			outWS['B11'] = 'X'
			outWS['B11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
			outWS['B11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			outWS['B11'].border = Border(top = dashB, bottom = thinB)
			outWS['C11'] = 'Y'
			outWS['C11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
			outWS['C11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			outWS['C11'].border = Border(top = dashB, bottom = thinB)
			outWS['D11'] = 'Z'
			outWS['D11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
			outWS['D11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			outWS['D11'].border = Border(top = dashB, bottom = thinB)

			excelCoors = []
			for geom_r in geomList:
				excelCoor = []
				excelCoor.append(geom_r[0])
				excelCoor.append(format(geom_r[1], '.6f'))
				excelCoor.append(format(geom_r[2], '.6f'))
				excelCoor.append(format(geom_r[3], '.6f'))
				excelCoors.append(excelCoor)

			for line in excelCoors:
				outWS.append(line)

			cellRange1 = outWS['A7:D8007']
			for cellNum in cellRange1:
				for cellNum1 in cellNum:
					cellNum1.font = Font(name = 'Times New Roman', size = 10)
					cellNum1.alignment = Alignment(horizontal = 'center', vertical = 'center')
			
			lastLine = len(geomList) + 11

			cellRange7 = outWS[f'A{lastLine}:D{lastLine}']
			for cellNum7 in cellRange7:
				for cellNum8 in cellNum7:
					cellNum8.border = Border(bottom = mediumB)

			outWB.save(f'{os.path.splitext(fileName)[0]}_CSI.xlsx')

		else:
			outWB = openpyxl.Workbook()
			outWS = outWB.active
			mediumB = Side(border_style = 'medium', color = '000000')
			thinB = Side(border_style = 'thin', color = '000000')
			dashB = Side(border_style = 'dashed', color = '000000')

			outWS.column_dimensions['A'].width = 7.25
			outWS.column_dimensions['B'].width = 11.25
			outWS.column_dimensions['C'].width = 11.25
			outWS.column_dimensions['D'].width = 11.25
			outWS.column_dimensions['E'].width = 7.25
			outWS.column_dimensions['F'].width = 11.25
			outWS.column_dimensions['G'].width = 11.25
			outWS.column_dimensions['H'].width = 11.25
			for w in range(1, 5000):
				outWS.row_dimensions[w].height = 16.0

			cellRange3 = outWS['A1:H1']
			for cellNum3 in cellRange3:
				for cellNum4 in cellNum3:
					cellNum4.border = Border(bottom = mediumB)

			cellRange5 = outWS['A9:H9']
			for cellNum5 in cellRange5:
				for cellNum6 in cellNum5:
					cellNum6.border = Border(bottom = thinB)

			cellRange = outWS['A2:H9']
			for cellNum in cellRange:
				for cellNum1 in cellNum:
					cellNum1.font = Font(name = 'Times New Roman', size = 10)

			outWS.merge_cells('A1:H1')
			outWS['A1'] = fileName.split('/')[-1].split('.')[0]
			outWS['A1'].font = Font(name = 'Times New Roman', size = 10.5, bold = True)
			outWS['A1'].alignment = Alignment(horizontal = 'center', vertical = 'center')

			outWS.merge_cells('A2:C9')
			outWS['A2'].font = Font(name = 'Times New Roman', size = 10.5, color = '929292')
			outWS['A2'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			outWS['A2'] = 'Insert molecular geometry here.'

			outWS.merge_cells('D2:H2')
			if rotChkBox.isChecked():
				outWS['D2'].font = Font(name = 'Courier', size = 10)
			outWS.merge_cells('D3:H3')
			outWS.merge_cells('D4:H4')
			outWS.merge_cells('D5:H5')
			outWS.merge_cells('D6:H6')
			outWS.merge_cells('D7:H7')
			outWS.merge_cells('D8:H8')
			outWS.merge_cells('D9:H9')

			optionSec = []
			if rotChkBox.isChecked():
				optionSec.append(outRouteLine)
			if chargeChkBox.isChecked() and spinChkBox.isChecked() and pgChkBox.isChecked():
				optionSec.append(f'Charge = {outCharge}, Multiplicity = {outMultiplicity}, Point group = {outPointGroup}')
			elif chargeChkBox.isChecked() == False and spinChkBox.isChecked() and pgChkBox.isChecked():
				optionSec.append(f'Multiplicity = {outMultiplicity}, Point group = {outPointGroup}')
			elif chargeChkBox.isChecked() and spinChkBox.isChecked() == False and pgChkBox.isChecked():
				optionSec.append(f'Charge = {outCharge}, Point group = {outPointGroup}')
			elif chargeChkBox.isChecked() and spinChkBox.isChecked() and pgChkBox.isChecked() == False:
				optionSec.append(f'Charge = {outCharge}, Multiplicity = {outMultiplicity}')
			elif chargeChkBox.isChecked() and spinChkBox.isChecked() == False and pgChkBox.isChecked() == False:
				optionSec.append(f'Charge = {outCharge}')
			elif chargeChkBox.isChecked() == False and spinChkBox.isChecked() and pgChkBox.isChecked() == False:
				optionSec.append(f'Multiplicity = {outMultiplicity}')
			elif chargeChkBox.isChecked() == False and spinChkBox.isChecked() == False and pgChkBox.isChecked():
				optionSec.append(f'Point group = {outPointGroup}')

			if nifChkBox.isChecked():
				if imFreq != 0:
					optionSec.append(f'Number of imaginary frequencies = {imFreq}, vi = {freqValue[0]}')
				else:
					optionSec.append(f'Number of imaginary frequencies = {imFreq}')
			if eleEnChkBox.isChecked():
				optionSec.append(f'Electronic Energy = {eleEn} Hartree')
			if zpeChkBox.isChecked():
				optionSec.append(f'Sum of electronic and zero-point Energies = {freqZPE} Hartree')
			if thrChkBox.isChecked():
				optionSec.append(f'Sum of electronic and thermal Energies = {freqThr} Hartree')
			if hChkBox.isChecked():
				optionSec.append(f'Sum of electronic and thermal Enthalpies = {freqH} Hartree')
			if feChkBox.isChecked():
				optionSec.append(f'Sum of electronic and thermal Free Energies = {freqFE} Hartree')

			if len(optionSec) != 0:
				for wb_i in range(len(optionSec)):
					outWS[f'D{2+wb_i}'] = optionSec[wb_i]

			outWS.merge_cells('A10:A11')
			outWS.merge_cells('B10:D10')
			outWS.merge_cells('E10:E11')
			outWS.merge_cells('F10:H10')

			outWS['A10'] = 'Atoms'
			outWS['A10'].font = Font(name = 'Times New Roman', size = 10, bold = True)
			outWS['A10'].alignment = Alignment(horizontal = 'center', vertical = 'bottom')
			outWS['A11'].border = Border(bottom = thinB)

			outWS['B10'] = 'Cartesian Coordinates'
			outWS['B10'].font = Font(name = 'Times New Roman', size = 10, bold = True)
			outWS['B10'].alignment = Alignment(horizontal = 'center', vertical = 'center')

			outWS['B11'] = 'X'
			outWS['B11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
			outWS['B11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			outWS['B11'].border = Border(top = dashB, bottom = thinB)
			outWS['C11'] = 'Y'
			outWS['C11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
			outWS['C11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			outWS['C11'].border = Border(top = dashB, bottom = thinB)
			outWS['D11'] = 'Z'
			outWS['D11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
			outWS['D11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			outWS['D11'].border = Border(top = dashB, bottom = thinB)

			outWS['E10'] = 'Atoms'
			outWS['E10'].font = Font(name = 'Times New Roman', size = 10, bold = True)
			outWS['E10'].alignment = Alignment(horizontal = 'center', vertical = 'bottom')
			outWS['E11'].border = Border(bottom = thinB)

			outWS['F10'] = 'Cartesian Coordinates'
			outWS['F10'].font = Font(name = 'Times New Roman', size = 10, bold = True)
			outWS['F10'].alignment = Alignment(horizontal = 'center', vertical = 'center')

			outWS['F11'] = 'X'
			outWS['F11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
			outWS['F11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			outWS['F11'].border = Border(top = dashB, bottom = thinB)
			outWS['G11'] = 'Y'
			outWS['G11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
			outWS['G11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			outWS['G11'].border = Border(top = dashB, bottom = thinB)
			outWS['H11'] = 'Z'
			outWS['H11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
			outWS['H11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			outWS['H11'].border = Border(top = dashB, bottom = thinB)

			excelCoors = []
			if len(geomList) % 2 == 0:
				excelCoorLineNO = int(len(geomList) / 2)
				for r in range(0, len(geomList), 2):
					excelCoor = []
					excelCoor.append(geomList[r][0])
					excelCoor.append(format(geomList[r][1], '.6f'))
					excelCoor.append(format(geomList[r][2], '.6f'))
					excelCoor.append(format(geomList[r][3], '.6f'))
					excelCoor.append(geomList[r + 1][0])
					excelCoor.append(format(geomList[r + 1][1], '.6f'))
					excelCoor.append(format(geomList[r + 1][2], '.6f'))
					excelCoor.append(format(geomList[r + 1][3], '.6f'))
					excelCoors.append(excelCoor)

			elif len(geomList) % 2 == 1:
				excelCoorLineNO = int((len(geomList) + 1) / 2)
				for r in range(0, len(geomList) - 1, 2):
					excelCoor = []
					excelCoor.append(geomList[r][0])
					excelCoor.append(format(geomList[r][1], '.6f'))
					excelCoor.append(format(geomList[r][2], '.6f'))
					excelCoor.append(format(geomList[r][3], '.6f'))
					excelCoor.append(geomList[r + 1][0])
					excelCoor.append(format(geomList[r + 1][1], '.6f'))
					excelCoor.append(format(geomList[r + 1][2], '.6f'))
					excelCoor.append(format(geomList[r + 1][3], '.6f'))
					excelCoors.append(excelCoor)
				excelCoor = []
				excelCoor.append(geomList[-1][0])
				excelCoor.append(format(geomList[-1][1], '.6f'))
				excelCoor.append(format(geomList[-1][2], '.6f'))
				excelCoor.append(format(geomList[-1][3], '.6f'))
				excelCoors.append(excelCoor)

			for line in excelCoors:
				outWS.append(line)

			cellRange2 = outWS['A12:H4001']
			for cellNum in cellRange2:
				for cellNum1 in cellNum:
					cellNum1.font = Font(name = 'Times New Roman', size = 10)
					cellNum1.alignment = Alignment(horizontal = 'center', vertical = 'center')

			lastLine = excelCoorLineNO + 11

			cellRange7 = outWS[f'A{lastLine}:H{lastLine}']
			for cellNum7 in cellRange7:
				for cellNum8 in cellNum7:
					cellNum8.border = Border(bottom = mediumB)
			cellRange9 = outWS[f'E10:E{lastLine}']
			for cellNum9 in cellRange9:
				for cellNum10 in cellNum9:
					cellNum10.border = Border(left = dashB)
			outWS['E11'].border = Border(left = dashB, bottom = thinB)
			outWS[f'E{lastLine}'].border = Border(left = dashB, bottom = mediumB)

			outWB.save(f'{os.path.splitext(fileName)[0]}_CSI.xlsx')

'''
*************************************************************************

            TYPE-B WINDOW: for .log files without Bq atoms

*************************************************************************
'''
class TypeBWindow(QMainWindow):
	def __init__(self):
		super().__init__()

		self.setWindowTitle(fileName.split('/')[-1])

		with open(fileName, 'r') as outFile:
			outFileLine = outFile.readlines()

		nmrFlag = 0

		for jj in list(range(len(outFileLine)-2)):
			if '-' in outFileLine[jj].strip() and '#' in outFileLine[jj+1].strip() \
			and '-' in outFileLine[jj+2].strip():
				if 'nmr' in outFileLine[jj+1].strip().lower():
					nmrFlag = 1
				else:
					pass
				break

		window_toolbar = QToolBar('Type B Toolbar')
		self.addToolBar(window_toolbar)
		window_toolbar.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)

		self.bla_window = None
		blaAction = QAction(QIcon('./assets/bla.png'), 'BLA', self)
		blaAction.setStatusTip('Plot bond length alternation graph for specified bonds.')
		blaAction.triggered.connect(self.blaFunction)

		self.homa_window = None
		homaAction = QAction(QIcon('./assets/homa.png'), 'HOMA', self)
		homaAction.setStatusTip('Compute HOMA value for monocycles.')
		homaAction.triggered.connect(self.homaFunction)

		self.pova_window = None
		povaAction = QAction(QIcon('./assets/poav.png'), 'POAV', self)
		povaAction.setStatusTip('Compute POAV1 and POAV2 for specified atom.')
		povaAction.triggered.connect(self.povaFunction)

		self.sp_window = None
		spAction = QAction(QIcon('./assets/sp.png'), 'NICS', self)
		spAction.setStatusTip('Create input file for single point NICS calculations.')
		spAction.triggered.connect(self.spFunction)

		self.scan_window = None
		nicsScanAction = QAction(QIcon('./assets/scan.png'), 'NICS Scan', self)
		nicsScanAction.setStatusTip('Create input file for 1D NICS scan calculations.')
		nicsScanAction.triggered.connect(self.scanFunction)

		self.nics2d_window = None
		nics2DAction = QAction(QIcon('./assets/2d.png'), '2D NICS', self)
		nics2DAction.setStatusTip('Create input file for 2D NICS (ICSS) calculations.')
		nics2DAction.triggered.connect(self.nics2DFunction)

		self.nics3d_window = None
		nics3DAction = QAction(QIcon('./assets/3d.png'), '3D NICS', self)
		nics3DAction.setStatusTip('Create input file for 3D NICS (ICSS) calculations.')
		nics3DAction.triggered.connect(self.nics3DFunction)

		self.inics_window = None
		inicsAction = QAction(QIcon('./assets/inics.png'), 'INICS', self)
		inicsAction.setStatusTip('Create input file for INICS calculations.')
		inicsAction.triggered.connect(self.inicsFunction)

		if nmrFlag == 1:
			self.nmr_window = None
			nmrAction = QAction(QIcon('./assets/nmr.png'), 'NMR', self)
			nmrAction.setStatusTip('Process NMR result.')
			nmrAction.triggered.connect(self.nmrFunction)

		self.csi_window = None
		csi3DAction = QAction(QIcon('./assets/csi.png'), 'CSI', self)
		csi3DAction.setStatusTip('Generate computational supporting informations.')
		csi3DAction.triggered.connect(self.csiFunction)

		window_toolbar.addAction(blaAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(homaAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(povaAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(spAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(nicsScanAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(nics2DAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(nics3DAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(inicsAction)
		if nmrFlag == 1:
			window_toolbar.addSeparator()
			window_toolbar.addAction(nmrAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(csi3DAction)

		window_toolbar.setMovable(False)

		self.setStatusBar(QStatusBar(self))

		molView = Mpl3DCanvas()
		molView.ax.axis('off')

		if xMax == xMin:
			molView.ax.set_xlim(xMin - 1, xMin + 1)
			molView.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView.ax.set_ylim(yMin - 1, yMin + 1)
			molView.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView.ax.set_zlim(zMin - 1, zMin + 1)
			molView.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_i in bndAtom:
			molView.ax.plot([xCoor[bndAtom_i[0]], xCoor[bndAtom_i[1]]], [yCoor[bndAtom_i[0]], yCoor[bndAtom_i[1]]], \
				[zCoor[bndAtom_i[0]], zCoor[bndAtom_i[1]]], '0.5')
		molView.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		viewLayout = QVBoxLayout()
		viewLayout.addWidget(molView)

		viewer = QWidget()
		viewer.setLayout(viewLayout)

		self.setCentralWidget(viewer)

	def blaFunction(self):
		if self.bla_window is None:
			self.bla_window = BLAWindow()
		self.bla_window.show()

	def homaFunction(self):
		if self.homa_window is None:
			self.homa_window = HOMAWindow()
		self.homa_window.show()

	def povaFunction(self):
		if self.pova_window is None:
			self.pova_window = PovaWindow()
		self.pova_window.show()

	def spFunction(self):
		if self.sp_window is None:
			self.sp_window = SpWindow()
		self.sp_window.show()

	def scanFunction(self):
		if self.scan_window is None:
			self.scan_window = ScanWindow()
		self.scan_window.show()

	def nics2DFunction(self):
		if self.nics2d_window is None:
			self.nics2d_window = Nics2dWindow()
		self.nics2d_window.show()

	def nics3DFunction(self):
		if self.nics3d_window is None:
			self.nics3d_window = Nics3dWindow()
		self.nics3d_window.show()

	def inicsFunction(self):
		if self.inics_window is None:
			self.inics_window = InicsWindow()
		self.inics_window.show()

	def nmrFunction(self):
		if self.nmr_window is None:
			self.nmr_window = NmrWindow()
		self.nmr_window.show()

	def csiFunction(self):
		if self.csi_window is None:
			self.csi_window = CsiWindow()
		self.csi_window.show()

'''
*************************************************************************

                          SP NICS OUTPUT WINDOW

*************************************************************************
'''
class NICSoutWindow(QMainWindow):
	def __init__(self):
		super().__init__()

		self.setWindowTitle('NICS Output: ' + fileName.split('/')[-1])

		global logBqList
		logBqList = atomList[len(geomList) - len(atomList):]
		global xBqCoor, yBqCoor, zBqCoor, colorBqList
		xBqCoor, yBqCoor, zBqCoor, colorBqList = geomAnalyzer.save_coor_list(logBqList)

		global ringList, ringNums2
		ringList = geomAnalyzer.find_monocycle(bndAtom)
		ringNums2 = len(ringList)

		global isoTen, aniTen, xxTen, yxTen, zxTen, xyTen, yyTen, zyTen, xzTen, yzTen, zzTen
		isoTen, aniTen, xxTen, yxTen, zxTen, xyTen, yyTen, zyTen, xzTen, yzTen, zzTen = NICSout.save_tensor(fileName, len(logBqList), -1)

		spOutMainWidget = QWidget()
		spOutMainLayout = QHBoxLayout()

		global molView8
		molView8 = Mpl3DCanvas()
		molView8.ax.axis('off')

		if xMax2 == xMin2:
			molView8.ax.set_xlim(xMin2 - 1, xMin2 + 1)
			molView8.ax.set_box_aspect([2, yMax2 - yMin2, zMax2 - zMin2])
		elif yMax2 == yMin2:
			molView8.ax.set_ylim(yMin2 - 1, yMin2 + 1)
			molView8.ax.set_box_aspect([xMax2 - xMin2, 2, zMax2 - zMin2])
		elif zMax2 == zMin2:
			molView8.ax.set_zlim(zMin2 - 1, zMin2 + 1)
			molView8.ax.set_box_aspect([xMax2 - xMin2, yMax2 - yMin2, 2])
		else:
			molView8.ax.set_box_aspect([xMax2 - xMin2, yMax2 - yMin2, zMax2 - zMin2])
		
		for bndAtom_p in bndAtom:
			molView8.ax.plot([xCoor[bndAtom_p[0]], xCoor[bndAtom_p[1]]], [yCoor[bndAtom_p[0]], yCoor[bndAtom_p[1]]], \
				[zCoor[bndAtom_p[0]], zCoor[bndAtom_p[1]]], '0.5')
		molView8.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)
		molView8.ax.scatter(np.array([xBqCoor]), np.array([yBqCoor]), np.array([zBqCoor]), color = colorBqList, edgecolors = colorBqList, s = 40, depthshade = 0, alpha = 0.5)

		for bqNo in range(len(logBqList)):
			molView8.ax.text(logBqList[bqNo][1], logBqList[bqNo][2], logBqList[bqNo][3], bqNo + 1, color = 'darkred')

		toolBar8 = NavigationToolbar(molView8, self)

		viewLayout8 = QVBoxLayout()
		viewLayout8.addWidget(molView8)
		viewLayout8.addWidget(toolBar8)
		viewLayout8.setContentsMargins(0, 0, 0, 0)

		viewer8 = QWidget()
		viewer8.setLayout(viewLayout8)

		spOutResult = QWidget()
		spOutResultLayout = QVBoxLayout()
		spOutResult.setLayout(spOutResultLayout)

		global spOutText
		spOutText = QTextEdit()
		spOutResultLayout.addWidget(spOutText)

		radioBtnGroup = QGroupBox('Component of Shielding Tensors')
		radioBtnLayout = QGridLayout()
		radioBtnGroup.setLayout(radioBtnLayout)

		radion3Group = QButtonGroup()
		isoBtn = QRadioButton('Isotropy', self)
		isoBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(isoBtn, 0, 0)
		aniBtn = QRadioButton('Anisotropy', self)
		aniBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(aniBtn, 0, 1)
		xxBtn = QRadioButton('XX', self)
		xxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xxBtn, 1, 0)
		xyBtn = QRadioButton('XY', self)
		xyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xyBtn, 2, 0)
		xzBtn = QRadioButton('XZ', self)
		xzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xzBtn, 3, 0)
		yxBtn = QRadioButton('YX', self)
		yxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yxBtn, 1, 1)
		yyBtn = QRadioButton('YY', self)
		yyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yyBtn, 2, 1)
		yzBtn = QRadioButton('YZ', self)
		yzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yzBtn, 3, 1)
		zxBtn = QRadioButton('ZX', self)
		zxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zxBtn, 1, 2)
		zyBtn = QRadioButton('ZY', self)
		zyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zyBtn, 2, 2)
		zzBtn = QRadioButton('ZZ', self)
		zzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zzBtn, 3, 2)

		radion3Group.addButton(isoBtn, 1)
		radion3Group.addButton(aniBtn, 2)
		radion3Group.addButton(xxBtn, 3)
		radion3Group.addButton(yzBtn, 4)
		radion3Group.addButton(zxBtn, 5)
		radion3Group.addButton(xyBtn, 6)
		radion3Group.addButton(yyBtn, 7)
		radion3Group.addButton(zyBtn, 8)
		radion3Group.addButton(xzBtn, 9)
		radion3Group.addButton(yzBtn, 10)
		radion3Group.addButton(zzBtn, 11)

		zzBtn.setChecked(True)

		spOutResultLayout.addWidget(radioBtnGroup)

		disorderBtnGroup = QGroupBox('Obtain NICS_ZZ for Non-Planar or Tilted Rings')
		disorderBtnLayout = QGridLayout()
		disorderBtnGroup.setLayout(disorderBtnLayout)

		disorderBtnLayout.addWidget(QLabel('	Ring No.: '), 0, 0)

		global ringCombo_2
		ringCombo_2 = QComboBox()
		ringCombo_2.addItems([f'{iiii}' for iiii in (range(1, ringNums2+1))])
		ringCombo_2.setEditable(False)
		ringCombo_2.setCurrentText('1')

		disorderBtnLayout.addWidget(ringCombo_2, 0, 1)

		disorderBtnLayout.addWidget(QLabel('	  Bq No.: '), 1, 0)

		global bqCombo
		bqCombo = QComboBox()
		bqCombo.addItems([f'{jjj}' for jjj in (range(1, len(logBqList)+1))])
		bqCombo.setEditable(False)
		bqCombo.setCurrentText('1')

		disorderBtnLayout.addWidget(bqCombo, 1, 1)

		computeBtn = QPushButton('Compute ->')
		computeBtn.clicked.connect(self.computeNICSZZ)
		disorderBtnLayout.addWidget(computeBtn, 2,0)

		global nicszzOutText
		nicszzOutText = QLineEdit()
		nicszzOutText.setText('Click \"Compute\"!')
		disorderBtnLayout.addWidget(nicszzOutText, 2, 1)

		spOutResultLayout.addWidget(disorderBtnGroup)


		spOutBtnGroup = QWidget()
		spOutBtnLayout = QHBoxLayout()
		spOutBtnGroup.setLayout(spOutBtnLayout)

		self.bqLabelChk = QCheckBox('Show Bq No.')
		self.bqLabelChk.setCheckState(Qt.CheckState.Checked)
		self.bqLabelChk.stateChanged.connect(self.reDrawSpOutGeom)
		spOutBtnLayout.addWidget(self.bqLabelChk)

		self.ringLabelChk = QCheckBox('Show Ring No.')
		self.ringLabelChk.setCheckState(Qt.CheckState.Unchecked)
		self.ringLabelChk.stateChanged.connect(self.reDrawSpOutGeom)
		spOutBtnLayout.addWidget(self.ringLabelChk)

		saveTxtSpOutBtn = QPushButton('Save .txt')
		saveTxtSpOutBtn.clicked.connect(self.saveTxtSpOut)
		spOutBtnLayout.addWidget(saveTxtSpOutBtn)
		
		spOutResultLayout.addWidget(spOutBtnGroup)

		spOutMainLayout.addWidget(viewer8)
		spOutMainLayout.addWidget(spOutResult)
		spOutMainWidget.setLayout(spOutMainLayout)

		self.setCentralWidget(spOutMainWidget)

	def reDrawSpOutGeom(self):
		molView8.ax.cla()
		molView8.ax.grid(False)
		molView8.ax.axis('off')

		if xMax2 == xMin2:
			molView8.ax.set_xlim(xMin2 - 1, xMin2 + 1)
			molView8.ax.set_box_aspect([2, yMax2 - yMin2, zMax2 - zMin2])
		elif yMax2 == yMin2:
			molView8.ax.set_ylim(yMin2 - 1, yMin2 + 1)
			molView8.ax.set_box_aspect([xMax2 - xMin2, 2, zMax2 - zMin2])
		elif zMax2 == zMin2:
			molView8.ax.set_zlim(zMin2 - 1, zMin2 + 1)
			molView8.ax.set_box_aspect([xMax2 - xMin2, yMax2 - yMin2, 2])
		else:
			molView8.ax.set_box_aspect([xMax2 - xMin2, yMax2 - yMin2, zMax2 - zMin2])
		
		for bndAtom_p in bndAtom:
			molView8.ax.plot([xCoor[bndAtom_p[0]], xCoor[bndAtom_p[1]]], [yCoor[bndAtom_p[0]], yCoor[bndAtom_p[1]]], \
				[zCoor[bndAtom_p[0]], zCoor[bndAtom_p[1]]], '0.5')
		molView8.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)
		molView8.ax.scatter(np.array([xBqCoor]), np.array([yBqCoor]), np.array([zBqCoor]), color = colorBqList, edgecolors = colorBqList, s = 40, depthshade = 0, alpha = 0.5)

		if self.bqLabelChk.isChecked():
			for bqNo in range(len(logBqList)):
				molView8.ax.text(logBqList[bqNo][1], logBqList[bqNo][2], logBqList[bqNo][3], bqNo + 1, color = 'darkred')

		if self.ringLabelChk.isChecked():
			for ringNo in range(ringNums2):
				aveX = 0.0
				aveY = 0.0
				aveZ = 0.0
				for atom_m in ringList[ringNo]:
					aveX += atomList[atom_m][1]/len(ringList[ringNo])
					aveY += atomList[atom_m][2]/len(ringList[ringNo])
					aveZ += atomList[atom_m][3]/len(ringList[ringNo])
				
				molView8.ax.text(aveX, aveY, aveZ, ringNo+1, \
					bbox = dict(boxstyle = 'round', ec = (0., 0.5, 0.5), fc = (0., 0.8, 0.8)))

		molView8.draw()

	def tensorClicked(self):
		tensorType = ''
		if self.sender().isChecked():
			tensorType = self.sender().text()
		if tensorType == 'Isotropy':
			spOutText.setText(' No.     NICS(Iso) \n-------------------')
			for bq_l in range(len(logBqList)):
				spOutText.append(f'{str(bq_l + 1).rjust(4)}      {isoTen[bq_l]:.4f}')
		elif tensorType == 'Anisotropy':
			spOutText.setText(' No.     NICS(Ani) \n-------------------')
			for bq_l in range(len(logBqList)):
				spOutText.append(f'{str(bq_l + 1).rjust(4)}      {aniTen[bq_l]:.4f}')
		elif tensorType == 'XX':
			spOutText.setText(' No.     NICS(xx)  \n-------------------')
			for bq_l in range(len(logBqList)):
				spOutText.append(f'{str(bq_l + 1).rjust(4)}      {xxTen[bq_l]:.4f}')
		elif tensorType == 'YX':
			spOutText.setText(' No.     NICS(yx)  \n-------------------')
			for bq_l in range(len(logBqList)):
				spOutText.append(f'{str(bq_l + 1).rjust(4)}      {yxTen[bq_l]:.4f}')
		elif tensorType == 'ZX':
			spOutText.setText(' No.     NICS(zx)  \n-------------------')
			for bq_l in range(len(logBqList)):
				spOutText.append(f'{str(bq_l + 1).rjust(4)}      {zxTen[bq_l]:.4f}')
		elif tensorType == 'XY':
			spOutText.setText(' No.     NICS(xy)  \n-------------------')
			for bq_l in range(len(logBqList)):
				spOutText.append(f'{str(bq_l + 1).rjust(4)}      {xyTen[bq_l]:.4f}')
		elif tensorType == 'YY':
			spOutText.setText(' No.     NICS(yy)  \n-------------------')
			for bq_l in range(len(logBqList)):
				spOutText.append(f'{str(bq_l + 1).rjust(4)}      {yyTen[bq_l]:.4f}')
		elif tensorType == 'ZY':
			spOutText.setText(' No.     NICS(zy)  \n-------------------')
			for bq_l in range(len(logBqList)):
				spOutText.append(f'{str(bq_l + 1).rjust(4)}      {zyTen[bq_l]:.4f}')
		elif tensorType == 'XZ':
			spOutText.setText(' No.     NICS(xz)  \n-------------------')
			for bq_l in range(len(logBqList)):
				spOutText.append(f'{str(bq_l + 1).rjust(4)}      {xzTen[bq_l]:.4f}')
		elif tensorType == 'YZ':
			spOutText.setText(' No.     NICS(yz)  \n-------------------')
			for bq_l in range(len(logBqList)):
				spOutText.append(f'{str(bq_l + 1).rjust(4)}      {yzTen[bq_l]:.4f}')
		elif tensorType == 'ZZ':
			spOutText.setText(' No.     NICS(zz)  \n-------------------')
			for bq_l in range(len(logBqList)):
				spOutText.append(f'{str(bq_l + 1).rjust(4)}      {zzTen[bq_l]:.4f}')

	def saveTxtSpOut(self):
		spOutTxt = open(f'{os.path.splitext(fileName)[0]}_NICS_output.txt', 'w')
		spOutTxt.write('#\n#   File Created by py.Aroma 4\n#   Author: Zhe Wang (Ph.D.)\n')
		spOutTxt.write('#   https://wongzit.github.com/program/pyaroma/\n#\n\n')
		spOutTxt.write(spOutText.toPlainText())
		spOutTxt.write('\n\n')
		spOutTxt.close()

	def computeNICSZZ(self):
		ringNo_zz = int(ringCombo_2.currentText())
		bqNo_zz = int(bqCombo.currentText())
		norx, nory, norz = geomAnalyzer.normal_vector(ringList[ringNo_zz-1], geomList)
		
		comp_x = norx*norx*xxTen[bqNo_zz-1]+nory*norx*yxTen[bqNo_zz-1]+norz*norx*zxTen[bqNo_zz-1]
		comp_y = norx*nory*xyTen[bqNo_zz-1]+nory*nory*yyTen[bqNo_zz-1]+norz*nory*zyTen[bqNo_zz-1]
		comp_z = norx*norz*xzTen[bqNo_zz-1]+nory*norz*yzTen[bqNo_zz-1]+norz*norz*zzTen[bqNo_zz-1]
		new_nics = comp_x+comp_y+comp_z
		
		nicszzOutText.setText(f'{new_nics:.4f}')

'''
*************************************************************************

                        NICS SCAN OUTPUT WINDOW

*************************************************************************
'''
class NICSscanWindow(QMainWindow):
	def __init__(self):
		super().__init__()

		self.setWindowTitle('NICS Scan Output: ' + fileName.split('/')[-1])
		global logBqList
		logBqList = atomList[len(geomList) - len(atomList):]
		global xBqCoor, yBqCoor, zBqCoor, colorBqList
		xBqCoor, yBqCoor, zBqCoor, colorBqList = geomAnalyzer.save_coor_list(logBqList)

		global isoTen, aniTen, xxTen, yxTen, zxTen, xyTen, yyTen, zyTen, xzTen, yzTen, zzTen
		isoTen, aniTen, xxTen, yxTen, zxTen, xyTen, yyTen, zyTen, xzTen, yzTen, zzTen = NICSout.save_tensor(fileName, len(logBqList), -1)

		scanOutMainWidget = QWidget()
		scanOutMainLayout = QHBoxLayout()

		global molView9
		molView9 = Mpl3DCanvas()
		molView9.ax.axis('off')

		if xMax2 == xMin2:
			molView9.ax.set_xlim(xMin2 - 1, xMin2 + 1)
			molView9.ax.set_box_aspect([2, yMax2 - yMin2, zMax2 - zMin2])
		elif yMax2 == yMin2:
			molView9.ax.set_ylim(yMin2 - 1, yMin2 + 1)
			molView9.ax.set_box_aspect([xMax2 - xMin2, 2, zMax2 - zMin2])
		elif zMax2 == zMin2:
			molView9.ax.set_zlim(zMin2 - 1, zMin2 + 1)
			molView9.ax.set_box_aspect([xMax2 - xMin2, yMax2 - yMin2, 2])
		else:
			molView9.ax.set_box_aspect([xMax2 - xMin2, yMax2 - yMin2, zMax2 - zMin2])
		
		for bndAtom_q in bndAtom:
			molView9.ax.plot([xCoor[bndAtom_q[0]], xCoor[bndAtom_q[1]]], [yCoor[bndAtom_q[0]], yCoor[bndAtom_q[1]]], \
				[zCoor[bndAtom_q[0]], zCoor[bndAtom_q[1]]], '0.5')
		molView9.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)
		molView9.ax.scatter(np.array([xBqCoor]), np.array([yBqCoor]), np.array([zBqCoor]), color = colorBqList, edgecolors = colorBqList, s = 40, depthshade = 0, alpha = 0.5)

		toolBar9 = NavigationToolbar(molView9, self)

		viewLayout9 = QVBoxLayout()
		viewLayout9.addWidget(molView9)
		viewLayout9.addWidget(toolBar9)
		viewLayout9.setContentsMargins(0, 0, 0, 0)

		viewer9 = QWidget()
		viewer9.setLayout(viewLayout9)

		scanPreview = QWidget()
		previewLayout = QVBoxLayout()
		scanPreview.setLayout(previewLayout)

		global scanOutMpl
		scanOutMpl = MplScanOutCanvas()
		toolBar10 = NavigationToolbar(scanOutMpl, self)

		scanPreWidget = QWidget()
		scanPreviewLayout = QVBoxLayout()
		scanPreWidget.setLayout(scanPreviewLayout)

		scanPreviewLayout.addWidget(scanOutMpl)
		scanPreviewLayout.addWidget(toolBar10)
		scanPreviewLayout.setContentsMargins(0, 0, 0, 0)

		previewLayout.addWidget(scanPreWidget)

		scanOutResult = QWidget()
		scanOutResultLayout = QVBoxLayout()
		scanOutResult.setLayout(scanOutResultLayout)

		global scanOutText
		scanOutText = QTextEdit()
		scanOutResultLayout.addWidget(scanOutText)

		radioBtnGroup = QGroupBox('Component of Shielding Tensors')
		radioBtnLayout = QGridLayout()
		radioBtnGroup.setLayout(radioBtnLayout)

		radion3Group = QButtonGroup()
		isoBtn = QRadioButton('Isotropy', self)
		isoBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(isoBtn, 0, 0)
		aniBtn = QRadioButton('Anisotropy', self)
		aniBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(aniBtn, 0, 1)
		xxBtn = QRadioButton('XX', self)
		xxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xxBtn, 1, 0)
		xyBtn = QRadioButton('XY', self)
		xyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xyBtn, 2, 0)
		xzBtn = QRadioButton('XZ', self)
		xzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xzBtn, 3, 0)
		yxBtn = QRadioButton('YX', self)
		yxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yxBtn, 1, 1)
		yyBtn = QRadioButton('YY', self)
		yyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yyBtn, 2, 1)
		yzBtn = QRadioButton('YZ', self)
		yzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yzBtn, 3, 1)
		zxBtn = QRadioButton('ZX', self)
		zxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zxBtn, 1, 2)
		zyBtn = QRadioButton('ZY', self)
		zyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zyBtn, 2, 2)
		zzBtn = QRadioButton('ZZ', self)
		zzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zzBtn, 3, 2)

		radion3Group.addButton(isoBtn, 1)
		radion3Group.addButton(aniBtn, 2)
		radion3Group.addButton(xxBtn, 3)
		radion3Group.addButton(yzBtn, 4)
		radion3Group.addButton(zxBtn, 5)
		radion3Group.addButton(xyBtn, 6)
		radion3Group.addButton(yyBtn, 7)
		radion3Group.addButton(zyBtn, 8)
		radion3Group.addButton(xzBtn, 9)
		radion3Group.addButton(yzBtn, 10)
		radion3Group.addButton(zzBtn, 11)

		zzBtn.setChecked(True)

		scanOutResultLayout.addWidget(radioBtnGroup)

		scanOutBtnGroup = QWidget()
		scanOutBtnLayout = QHBoxLayout()
		scanOutBtnGroup.setLayout(scanOutBtnLayout)

		self.bqLabelChk = QCheckBox('Show Bq No.')
		self.bqLabelChk.setCheckState(Qt.CheckState.Unchecked)
		self.bqLabelChk.stateChanged.connect(self.reDrawScanOutGeom)
		scanOutBtnLayout.addWidget(self.bqLabelChk)

		savePngScanOutBtn = QPushButton('Save .png')
		savePngScanOutBtn.clicked.connect(self.savePngScanOutFunc)
		scanOutBtnLayout.addWidget(savePngScanOutBtn)

		saveXlsxScanOutBtn = QPushButton('Save .xlsx')
		saveXlsxScanOutBtn.clicked.connect(self.saveXlsxScanOut)
		scanOutBtnLayout.addWidget(saveXlsxScanOutBtn)
		
		scanOutResultLayout.addWidget(scanOutBtnGroup)

		scanOutMainLayout.addWidget(viewer9)
		scanOutMainLayout.addWidget(scanPreview)
		scanOutMainLayout.addWidget(scanOutResult)
		scanOutMainWidget.setLayout(scanOutMainLayout)

		self.setCentralWidget(scanOutMainWidget)
		self.setMinimumSize(1500, 600)

	def reDrawScanOutGeom(self):
		molView9.ax.cla()
		molView9.ax.grid(False)
		molView9.ax.axis('off')

		if xMax2 == xMin2:
			molView9.ax.set_xlim(xMin2 - 1, xMin2 + 1)
			molView9.ax.set_box_aspect([2, yMax2 - yMin2, zMax2 - zMin2])
		elif yMax2 == yMin2:
			molView9.ax.set_ylim(yMin2 - 1, yMin2 + 1)
			molView9.ax.set_box_aspect([xMax2 - xMin2, 2, zMax2 - zMin2])
		elif zMax2 == zMin2:
			molView9.ax.set_zlim(zMin2 - 1, zMin2 + 1)
			molView9.ax.set_box_aspect([xMax2 - xMin2, yMax2 - yMin2, 2])
		else:
			molView9.ax.set_box_aspect([xMax2 - xMin2, yMax2 - yMin2, zMax2 - zMin2])
	
		for bndAtom_q in bndAtom:
			molView9.ax.plot([xCoor[bndAtom_q[0]], xCoor[bndAtom_q[1]]], [yCoor[bndAtom_q[0]], yCoor[bndAtom_q[1]]], \
				[zCoor[bndAtom_q[0]], zCoor[bndAtom_q[1]]], '0.5')
		molView9.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)
		molView9.ax.scatter(np.array([xBqCoor]), np.array([yBqCoor]), np.array([zBqCoor]), color = colorBqList, edgecolors = colorBqList, s = 40, depthshade = 0, alpha = 0.5)

		if self.bqLabelChk.isChecked():
			for bqNo in range(len(logBqList)):
				molView9.ax.text(logBqList[bqNo][1], logBqList[bqNo][2], logBqList[bqNo][3], bqNo + 1, color = 'darkred')
		molView9.draw()

	def tensorClicked(self):
		scanOutMpl.ax.cla()
		global tensorType
		tensorType = ''
		if self.sender().isChecked():
			tensorType = self.sender().text()
		if tensorType == 'Isotropy':
			scanOutMpl.ax.plot(np.array(range(1, len(logBqList) + 1)), np.array(isoTen), c = '0.')
			scanOutText.setText(' No.     NICS(Iso) \n-------------------')
			for bq_l in range(len(logBqList)):
				scanOutText.append(f'{str(bq_l + 1).rjust(4)}      {isoTen[bq_l]:.4f}')
		elif tensorType == 'Anisotropy':
			scanOutMpl.ax.plot(np.array(range(1, len(logBqList) + 1)), np.array(aniTen), c = '0.')
			scanOutText.setText(' No.     NICS(Ani) \n-------------------')
			for bq_l in range(len(logBqList)):
				scanOutText.append(f'{str(bq_l + 1).rjust(4)}      {aniTen[bq_l]:.4f}')
		elif tensorType == 'XX':
			scanOutMpl.ax.plot(np.array(range(1, len(logBqList) + 1)), np.array(xxTen), c = '0.')
			scanOutText.setText(' No.     NICS(xx)  \n-------------------')
			for bq_l in range(len(logBqList)):
				scanOutText.append(f'{str(bq_l + 1).rjust(4)}      {xxTen[bq_l]:.4f}')
		elif tensorType == 'YX':
			scanOutMpl.ax.plot(np.array(range(1, len(logBqList) + 1)), np.array(yxTen), c = '0.')
			scanOutText.setText(' No.     NICS(yx)  \n-------------------')
			for bq_l in range(len(logBqList)):
				scanOutText.append(f'{str(bq_l + 1).rjust(4)}      {yxTen[bq_l]:.4f}')
		elif tensorType == 'ZX':
			scanOutMpl.ax.plot(np.array(range(1, len(logBqList) + 1)), np.array(zxTen), c = '0.')
			scanOutText.setText(' No.     NICS(zx)  \n-------------------')
			for bq_l in range(len(logBqList)):
				scanOutText.append(f'{str(bq_l + 1).rjust(4)}      {zxTen[bq_l]:.4f}')
		elif tensorType == 'XY':
			scanOutMpl.ax.plot(np.array(range(1, len(logBqList) + 1)), np.array(xyTen), c = '0.')
			scanOutText.setText(' No.     NICS(xy)  \n-------------------')
			for bq_l in range(len(logBqList)):
				scanOutText.append(f'{str(bq_l + 1).rjust(4)}      {xyTen[bq_l]:.4f}')
		elif tensorType == 'YY':
			scanOutMpl.ax.plot(np.array(range(1, len(logBqList) + 1)), np.array(yyTen), c = '0.')
			scanOutText.setText(' No.     NICS(yy)  \n-------------------')
			for bq_l in range(len(logBqList)):
				scanOutText.append(f'{str(bq_l + 1).rjust(4)}      {yyTen[bq_l]:.4f}')
		elif tensorType == 'ZY':
			scanOutMpl.ax.plot(np.array(range(1, len(logBqList) + 1)), np.array(zyTen), c = '0.')
			scanOutText.setText(' No.     NICS(zy)  \n-------------------')
			for bq_l in range(len(logBqList)):
				scanOutText.append(f'{str(bq_l + 1).rjust(4)}      {zyTen[bq_l]:.4f}')
		elif tensorType == 'XZ':
			scanOutMpl.ax.plot(np.array(range(1, len(logBqList) + 1)), np.array(xzTen), c = '0.')
			scanOutText.setText(' No.     NICS(xz)  \n-------------------')
			for bq_l in range(len(logBqList)):
				scanOutText.append(f'{str(bq_l + 1).rjust(4)}      {xzTen[bq_l]:.4f}')
		elif tensorType == 'YZ':
			scanOutMpl.ax.plot(np.array(range(1, len(logBqList) + 1)), np.array(yzTen), c = '0.')
			scanOutText.setText(' No.     NICS(yz)  \n-------------------')
			for bq_l in range(len(logBqList)):
				scanOutText.append(f'{str(bq_l + 1).rjust(4)}      {yzTen[bq_l]:.4f}')
		elif tensorType == 'ZZ':
			scanOutMpl.ax.plot(np.array(range(1, len(logBqList) + 1)), np.array(zzTen), c = '0.')
			scanOutText.setText(' No.     NICS(zz)  \n-------------------')
			for bq_l in range(len(logBqList)):
				scanOutText.append(f'{str(bq_l + 1).rjust(4)}      {zzTen[bq_l]:.4f}')

		scanOutMpl.ax.set_xlabel('Bq No.')
		scanOutMpl.ax.set_ylabel(f'NICS({tensorType.lower()}) / ppm')
		figScanOut.tight_layout()
		scanOutMpl.draw()

	def savePngScanOutFunc(self):
		figScanOut.savefig(f'{os.path.splitext(fileName)[0]}_NICS_Scan_output.png', dpi = 300)

	def saveXlsxScanOut(self):
		scanWB = openpyxl.Workbook()
		scanWS = scanWB.active
		scanWS['A1'] = 'Bq No.'
		scanWS['B1'] = f'NICS({tensorType.lower()}) / ppm'
		for j in range(len(logBqList)):
			if tensorType == 'Isotropy':
				scanWS.append([j + 1, isoTen[j]])
			elif tensorType == 'Anisotropy':
				scanWS.append([j + 1, aniTen[j]])
			elif tensorType == 'XX':
				scanWS.append([j + 1, xxTen[j]])
			elif tensorType == 'YX':
				scanWS.append([j + 1, yxTen[j]])
			elif tensorType == 'ZX':
				scanWS.append([j + 1, zxTen[j]])
			elif tensorType == 'XY':
				scanWS.append([j + 1, xyTen[j]])
			elif tensorType == 'YY':
				scanWS.append([j + 1, yyTen[j]])
			elif tensorType == 'ZY':
				scanWS.append([j + 1, zyTen[j]])
			elif tensorType == 'XZ':
				scanWS.append([j + 1, xzTen[j]])
			elif tensorType == 'YZ':
				scanWS.append([j + 1, yzTen[j]])
			elif tensorType == 'ZZ':
				scanWS.append([j + 1, zzTen[j]])

		scanChart = ScatterChart()
		xValue = Reference(scanWS, min_col = 1, min_row = 2, max_row = len(logBqList) + 1)
		yValue = Reference(scanWS, min_col = 2, min_row = 2, max_row = len(logBqList) + 1)
		scanChart.series.append(Series(yValue, xValue))
		scanChart.legend = None
		scanChart.x_axis.title = 'Bq No.'
		scanChart.y_axis.title = f'NICS({tensorType.lower()}) / ppm'
		scanWS.add_chart(scanChart, 'E2')
		scanWB.save(f'{os.path.splitext(fileName)[0]}_NICS_Scan.xlsx')

'''
*************************************************************************

                         2D NICS OUTPUT WINDOW

*************************************************************************
'''
class NICS2DWindow(QMainWindow):
	def __init__(self):
		super().__init__()

		self.setWindowTitle('2D NICS Output: ' + fileName.split('/')[-1])

		logBqList = atomList[len(geomList) - len(atomList):]
		xBqCoor, yBqCoor, zBqCoor, colorBqList = geomAnalyzer.save_coor_list(logBqList)

		global xBqSet, yBqSet, zBqSet, n2OutPlane
		xBqSet = sorted(set(xBqCoor))
		yBqSet = sorted(set(yBqCoor))
		zBqSet = sorted(set(zBqCoor))
		n2OutPlane = ''
		if len(xBqSet) == 1:
			n2OutPlane = 'yz'
		elif len(yBqSet) == 1:
			n2OutPlane = 'xz'
		elif len(zBqSet) == 1:
			n2OutPlane = 'xy'

		global isoTen, aniTen, xxTen, yxTen, zxTen, xyTen, yyTen, zyTen, xzTen, yzTen, zzTen
		isoTen, aniTen, xxTen, yxTen, zxTen, xyTen, yyTen, zyTen, xzTen, yzTen, zzTen = NICSout.save_tensor(fileName, len(logBqList), int(configFile.get('general', 'icss')))

		n2OutMainWidget = QWidget()
		n2OutMainLayout = QHBoxLayout()

		molView10 = Mpl3DCanvas()

		if xMax == xMin:
			molView10.ax.set_xlim(xMin - 1, xMin + 1)
			molView10.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView10.ax.set_ylim(yMin - 1, yMin + 1)
			molView10.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView10.ax.set_zlim(zMin - 1, zMin + 1)
			molView10.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView10.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_r in bndAtom:
			molView10.ax.plot([xCoor[bndAtom_r[0]], xCoor[bndAtom_r[1]]], [yCoor[bndAtom_r[0]], yCoor[bndAtom_r[1]]], \
				[zCoor[bndAtom_r[0]], zCoor[bndAtom_r[1]]], '0.5')
		molView10.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		toolBar11 = NavigationToolbar(molView10, self)

		viewLayout10 = QVBoxLayout()
		viewLayout10.addWidget(molView10)
		viewLayout10.addWidget(toolBar11)
		viewLayout10.setContentsMargins(0, 0, 0, 0)

		viewer10 = QWidget()
		viewer10.setLayout(viewLayout10)

		n2Preview = QWidget()
		n2preLayout = QVBoxLayout()
		n2Preview.setLayout(n2preLayout)

		global n2OutMpl
		n2OutMpl = MplN2OutCanvas()
		toolBar12 = NavigationToolbar(n2OutMpl, self)

		global cmapCombo
		cmapCombo = QComboBox()
		cmapCombo.addItems(CONSTANT.mplCmap)
		cmapCombo.setEditable(True)
		cmapCombo.setCurrentText('viridis')

		global lcolorCombo
		lcolorCombo = QComboBox()
		lcolorCombo.addItems(CONSTANT.molColor)
		lcolorCombo.setEditable(True)
		lcolorCombo.setCurrentText('white')

		icssArray = []
		global icssFig, cB
		if n2OutPlane == 'xy':
			for n_y in range(len(yBqSet)):
				icssArrayLine = []
				for n_x in range(len(xBqSet)):
					icssArrayLine.append(zzTen[n_y + len(yBqSet) * n_x])
				icssArray.append(icssArrayLine)
			x2Dvalue, y2Dvalue = np.meshgrid(xBqSet, yBqSet)
			icssFig = n2OutMpl.ax.pcolormesh(x2Dvalue, y2Dvalue, np.array(icssArray), shading = 'gouraud', cmap = cmapCombo.currentText())
			cB = figN2Out.colorbar(icssFig, cax = make_axes_locatable(n2OutMpl.ax).append_axes('right', size = '5%', pad = 0.05))
			cB.mappable.set_clim(min(zzTen), max(zzTen))
			n2OutMpl.ax.set_xlabel('X / Å')
			n2OutMpl.ax.set_ylabel('Y / Å')

		elif n2OutPlane == 'xz':
			for n_z in range(len(zBqSet)):
				icssArrayLine = []
				for n_x in range(len(xBqSet)):
					icssArrayLine.append(zzTen[n_z + len(zBqSet) * n_x])
				icssArray.append(icssArrayLine)
			x2Dvalue, z2Dvalue = np.meshgrid(xBqSet, zBqSet)
			icssFig = n2OutMpl.ax.pcolormesh(x2Dvalue, z2Dvalue, np.array(icssArray), shading = 'gouraud', cmap = cmapCombo.currentText())
			cB = figN2Out.colorbar(icssFig, cax = make_axes_locatable(n2OutMpl.ax).append_axes('right', size = '5%', pad = 0.05))
			cB.mappable.set_clim(min(zzTen), max(zzTen))
			n2OutMpl.ax.set_xlabel('X / Å')
			n2OutMpl.ax.set_ylabel('Z / Å')

		elif n2OutPlane == 'yz':
			for n_z in range(len(zBqSet)):
				icssArrayLine = []
				for n_y in range(len(yBqSet)):
					icssArrayLine.append(zzTen[n_z + len(zBqSet) * n_y])
				icssArray.append(icssArrayLine)
			y2Dvalue, z2Dvalue = np.meshgrid(yBqSet, zBqSet)
			icssFig = n2OutMpl.ax.pcolormesh(y2Dvalue, z2Dvalue, np.array(icssArray), shading = 'gouraud', cmap = cmapCombo.currentText())
			cB = figN2Out.colorbar(icssFig, cax = make_axes_locatable(n2OutMpl.ax).append_axes('right', size = '5%', pad = 0.05))
			cB.mappable.set_clim(min(zzTen), max(zzTen))
			n2OutMpl.ax.set_xlabel('Y / Å')
			n2OutMpl.ax.set_ylabel('Z / Å')
	
		if n2OutPlane == 'xy':
			polyXn2 = [min(xBqSet), max(xBqSet), max(xBqSet), min(xBqSet)]
			polyYn2 = [min(yBqSet), min(yBqSet), max(yBqSet), max(yBqSet)]
			polyZn2 = [zBqSet[0], zBqSet[0], zBqSet[0], zBqSet[0]]
			polyN2 = list(zip(polyXn2, polyYn2, polyZn2))
			molView10.ax.add_collection3d(Poly3DCollection([polyN2], facecolor = '#000080', alpha = 0.3))
		elif n2OutPlane == 'xz':
			polyXn2 = [min(xBqSet), max(xBqSet), max(xBqSet), min(xBqSet)]
			polyYn2 = [yBqSet[0], yBqSet[0], yBqSet[0], yBqSet[0]]
			polyZn2 = [min(zBqSet), min(zBqSet), max(zBqSet), max(zBqSet)]
			polyN2 = list(zip(polyXn2, polyYn2, polyZn2))
			molView10.ax.add_collection3d(Poly3DCollection([polyN2], facecolor = '#000080', alpha = 0.3))
		elif n2OutPlane == 'yz':
			polyXn2 = [xBqSet[0], xBqSet[0], xBqSet[0], xBqSet[0]]
			polyYn2 = [min(yBqSet), max(yBqSet), max(yBqSet), min(yBqSet)]
			polyZn2 = [min(zBqSet), min(zBqSet), max(zBqSet), max(zBqSet)]
			polyN2 = list(zip(polyXn2, polyYn2, polyZn2))
			molView10.ax.add_collection3d(Poly3DCollection([polyN2], facecolor = '#000080', alpha = 0.3))

		n2PreWidget = QWidget()
		n2PreviewLayout = QVBoxLayout()
		n2PreWidget.setLayout(n2PreviewLayout)

		global emphasisLine, cLabelChk
		emphasisLine = QLineEdit()
		cLabelChk = QCheckBox('Show Contour Label')

		n2PreviewLayout.addWidget(n2OutMpl)
		n2PreviewLayout.addWidget(toolBar12)
		n2PreviewLayout.setContentsMargins(0, 0, 0, 0)

		n2preLayout.addWidget(n2PreWidget)

		n2OutResult = QWidget()
		n2OutResultLayout = QVBoxLayout()
		n2OutResult.setLayout(n2OutResultLayout)

		radioBtnGroup = QGroupBox('Component of Shielding Tensors')
		radioBtnLayout = QGridLayout()
		radioBtnGroup.setLayout(radioBtnLayout)

		radion4Group = QButtonGroup()
		isoBtn = QRadioButton('Isotropy', self)
		isoBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(isoBtn, 0, 0)
		aniBtn = QRadioButton('Anisotropy', self)
		aniBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(aniBtn, 0, 1)
		xxBtn = QRadioButton('XX', self)
		xxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xxBtn, 1, 0)
		xyBtn = QRadioButton('XY', self)
		xyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xyBtn, 2, 0)
		xzBtn = QRadioButton('XZ', self)
		xzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xzBtn, 3, 0)
		yxBtn = QRadioButton('YX', self)
		yxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yxBtn, 1, 1)
		yyBtn = QRadioButton('YY', self)
		yyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yyBtn, 2, 1)
		yzBtn = QRadioButton('YZ', self)
		yzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yzBtn, 3, 1)
		zxBtn = QRadioButton('ZX', self)
		zxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zxBtn, 1, 2)
		zyBtn = QRadioButton('ZY', self)
		zyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zyBtn, 2, 2)
		zzBtn = QRadioButton('ZZ', self)
		zzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zzBtn, 3, 2)

		radion4Group.addButton(isoBtn, 1)
		radion4Group.addButton(aniBtn, 2)
		radion4Group.addButton(xxBtn, 3)
		radion4Group.addButton(yzBtn, 4)
		radion4Group.addButton(zxBtn, 5)
		radion4Group.addButton(xyBtn, 6)
		radion4Group.addButton(yyBtn, 7)
		radion4Group.addButton(zyBtn, 8)
		radion4Group.addButton(xzBtn, 9)
		radion4Group.addButton(yzBtn, 10)
		radion4Group.addButton(zzBtn, 11)

		zzBtn.setChecked(True)

		n2OutResultLayout.addWidget(radioBtnGroup)

		counterGroup = QGroupBox('Graph Style')
		counterLayout = QGridLayout()
		counterGroup.setLayout(counterLayout)

		global conLineChk
		conLineChk = QCheckBox('Show Contour Line')
		conLineChk.setCheckState(Qt.CheckState.Unchecked)
		conLineChk.stateChanged.connect(self.conChkFunc)
		counterLayout.addWidget(conLineChk, 0, 0)

		global stepsizeLine
		counterLayout.addWidget(QLabel('	Stepsize: '), 1, 0)
		stepsizeLine = QLineEdit()
		stepsizeLine.setText('10')
		counterLayout.addWidget(stepsizeLine, 1, 1)

		counterLayout.addWidget(QLabel('	Emphasize: '), 2, 0)
		counterLayout.addWidget(emphasisLine, 2, 1)

		counterLayout.addWidget(cLabelChk, 0, 1)

		counterLayout.addWidget(QLabel('Heatmap Color Style: '), 3, 0)
		counterLayout.addWidget(QLabel('Contour Line Color: '), 4, 0)

		counterLayout.addWidget(cmapCombo, 3, 1)
		counterLayout.addWidget(lcolorCombo, 4, 1)

		emphasisBtn = QPushButton('Update')
		emphasisBtn.clicked.connect(self.updateHmap)
		counterLayout.addWidget(emphasisBtn, 5, 0)

		clrBtn = QPushButton('Clear')
		clrBtn.clicked.connect(self.clrHmap)
		counterLayout.addWidget(clrBtn, 5, 1)

		n2OutResultLayout.addWidget(counterGroup)

		n2OutBtnGroup = QWidget()
		n2OutBtnLayout = QHBoxLayout()
		n2OutBtnGroup.setLayout(n2OutBtnLayout)

		savePngn2OutBtn = QPushButton('Save .png')
		savePngn2OutBtn.clicked.connect(self.savePngn2OutFunc)
		n2OutBtnLayout.addWidget(savePngn2OutBtn)

		saveCsvN2OutBtn = QPushButton('Save .csv')
		saveCsvN2OutBtn.clicked.connect(self.saveCsvn2Out)
		n2OutBtnLayout.addWidget(saveCsvN2OutBtn)
		
		n2OutResultLayout.addWidget(n2OutBtnGroup)

		statusBox = QTextEdit()
		statusBox.setEnabled(False)
		n2OutResultLayout.addWidget(statusBox)

		n2OutMainLayout.addWidget(viewer10)
		n2OutMainLayout.addWidget(n2Preview)
		n2OutMainLayout.addWidget(n2OutResult)
		n2OutMainWidget.setLayout(n2OutMainLayout)

		self.setCentralWidget(n2OutMainWidget)
		self.setMinimumSize(1500, 600)

	def reDraw2DOut(self, map_value):
		n2OutMpl.ax.cla()
		icssArray = []
		if n2OutPlane == 'xy':
			for n_y in range(len(yBqSet)):
				icssArrayLine = []
				for n_x in range(len(xBqSet)):
					icssArrayLine.append(map_value[n_y + len(yBqSet) * n_x])
				icssArray.append(icssArrayLine)
			x2Dvalue, y2Dvalue = np.meshgrid(xBqSet, yBqSet)
			icssFig = n2OutMpl.ax.pcolormesh(x2Dvalue, y2Dvalue, np.array(icssArray), shading = 'gouraud', cmap = cmapCombo.currentText())
			cB.mappable.set_clim(min(map_value), max(map_value))
			cB.mappable.set_cmap(cmapCombo.currentText())
			conX, conY = np.meshgrid(xBqSet, yBqSet)
			if configFile.get('heatmap', 'cb') == 'true':
				n2OutMpl.ax.contour(conX, conY, np.array(icssArray), colors = lcolorCombo.currentText(), linewidths = 0.3, levels = int(stepsizeLine.text()))
			if emphasisLine.text() != '':
				em = n2OutMpl.ax.contour(conX, conY, np.array(icssArray), colors = lcolorCombo.currentText(), linestyles = '-', linewidths = 1.0, levels = [float(emphasisLine.text())])
				if cLabelChk.isChecked():
					n2OutMpl.ax.clabel(em, inline = True)
			n2OutMpl.ax.set_xlabel('X / Å')
			n2OutMpl.ax.set_ylabel('Y / Å')

		elif n2OutPlane == 'xz':
			for n_z in range(len(zBqSet)):
				icssArrayLine = []
				for n_x in range(len(xBqSet)):
					icssArrayLine.append(map_value[n_z + len(zBqSet) * n_x])
				icssArray.append(icssArrayLine)
			x2Dvalue, z2Dvalue = np.meshgrid(xBqSet, zBqSet)
			icssFig = n2OutMpl.ax.pcolormesh(x2Dvalue, z2Dvalue, np.array(icssArray), shading = 'gouraud', cmap = cmapCombo.currentText())
			cB.mappable.set_clim(min(map_value), max(map_value))
			cB.mappable.set_cmap(cmapCombo.currentText())
			conX, conZ = np.meshgrid(xBqSet, zBqSet)
			if configFile.get('heatmap', 'cb') == 'true':
				n2OutMpl.ax.contour(conX, conZ, np.array(icssArray), colors = lcolorCombo.currentText(), linewidths = 0.3, levels = int(stepsizeLine.text()))
			if emphasisLine.text() != '':
				em = n2OutMpl.ax.contour(conX, conZ, np.array(icssArray), colors = lcolorCombo.currentText(), linestyles = '-', linewidths = 1.0, levels = [float(emphasisLine.text())])
				if cLabelChk.isChecked():
					n2OutMpl.ax.clabel(em, inline = True)
			n2OutMpl.ax.set_xlabel('X / Å')
			n2OutMpl.ax.set_ylabel('Z / Å')

		elif n2OutPlane == 'yz':
			for n_z in range(len(zBqSet)):
				icssArrayLine = []
				for n_y in range(len(yBqSet)):
					icssArrayLine.append(map_value[n_z + len(zBqSet) * n_y])
				icssArray.append(icssArrayLine)
			y2Dvalue, z2Dvalue = np.meshgrid(yBqSet, zBqSet)
			icssFig = n2OutMpl.ax.pcolormesh(y2Dvalue, z2Dvalue, np.array(icssArray), shading = 'gouraud', cmap = cmapCombo.currentText())
			cB.mappable.set_clim(min(map_value), max(map_value))
			cB.mappable.set_cmap(cmapCombo.currentText())
			conY, conZ = np.meshgrid(yBqSet, zBqSet)
			if configFile.get('heatmap', 'cb') == 'true':
				n2OutMpl.ax.contour(conY, conZ, np.array(icssArray), colors = lcolorCombo.currentText(), linewidths = 0.3, levels = int(stepsizeLine.text()))
			if emphasisLine.text() != '':
				em = n2OutMpl.ax.contour(conY, conZ, np.array(icssArray), colors = lcolorCombo.currentText(), linestyles = '-', linewidths = 1.0, levels = [float(emphasisLine.text())])
				if cLabelChk.isChecked():
					n2OutMpl.ax.clabel(em, inline = True)
			n2OutMpl.ax.set_xlabel('Y / Å')
			n2OutMpl.ax.set_ylabel('Z / Å')

		n2OutMpl.draw()

	def tensorClicked(self):
		global tensorType
		tensorType = ''
		if self.sender().isChecked():
			tensorType = self.sender().text()
		global mapValue
		mapValue = zzTen
		
		if tensorType == 'Isotropy':
			mapValue = isoTen
		elif tensorType == 'Anisotropy':
			mapValue = aniTen
		elif tensorType == 'XX':
			mapValue = xxTen
		elif tensorType == 'YX':
			mapValue = yxTen
		elif tensorType == 'ZX':
			mapValue = zxTen
		elif tensorType == 'XY':
			mapValue = xyTen
		elif tensorType == 'YY':
			mapValue = yyTen
		elif tensorType == 'ZY':
			mapValue = zyTen
		elif tensorType == 'XZ':
			mapValue = xzTen
		elif tensorType == 'YZ':
			mapValue = yzTen
		elif tensorType == 'ZZ':
			mapValue = zzTen

		self.reDraw2DOut(mapValue)

	def updateHmap(self):
		self.reDraw2DOut(mapValue)

	def conChkFunc(self):
		if conLineChk.isChecked():
			configFile.set('heatmap', 'cb', 'true')
			self.reDraw2DOut(mapValue)
		else:
			configFile.set('heatmap', 'cb', 'false')
			self.reDraw2DOut(mapValue)

	def savePngn2OutFunc(self):
		if configFile.get('general', 'icss') == '-1':
			figN2Out.savefig(f'{os.path.splitext(fileName)[0]}_2D_NICS_output_{tensorType.upper()}.png', dpi = 300)
		else:
			figN2Out.savefig(f'{os.path.splitext(fileName)[0]}_2D_ICSS_output_{tensorType.upper()}.png', dpi = 300)

	def saveCsvn2Out(self):
		if configFile.get('general', 'icss') == '-1':
			icssCsvOutput = open(f'{os.path.splitext(fileName)[0]}_2D_NICS_output_{tensorType.upper()}.csv', 'w')
		else:
			icssCsvOutput = open(f'{os.path.splitext(fileName)[0]}_2D_ICSS_output_{tensorType.upper()}.csv', 'w')

		icssCsvOutput.write('#\n#   File Created by py.Aroma 4\n')
		icssCsvOutput.write('#   Author: Zhe Wang (Ph.D.)\n')
		icssCsvOutput.write('#   https://wongzit.github.com/program/pyaroma/\n#\n\n')

		if n2OutPlane == 'xy':
			icssCsvOutput.write('XY,')
			for xStep in xBqSet:
				icssCsvOutput.write(f'{xStep},')
			icssCsvOutput.write('\n')
			yCount = 0
			for yNo in range(len(yBqSet)):
				if yCount < len(yBqSet):
					icssCsvOutput.write(f'{yBqSet[yCount]},')
				for xNo in range(len(xBqSet)):
					icssCsvOutput.write(f'{mapValue[yNo + len(yBqSet) * xNo]},')
				yCount += 1
				icssCsvOutput.write('\n')

		elif n2OutPlane == 'xz':
			icssCsvOutput.write('XZ,')
			for xStep in xBqSet:
				icssCsvOutput.write(f'{xStep},')
			icssCsvOutput.write('\n')
			zCount = 0
			for zNo in range(len(zBqSet)):
				if zCount < len(zBqSet):
					icssCsvOutput.write(f'{zBqSet[zCount]},')
				for xNo in range(len(xBqSet)):
					icssCsvOutput.write(f'{mapValue[zNo + len(zBqSet) * xNo]},')
				zCount += 1
				icssCsvOutput.write('\n')

		elif n2OutPlane == 'yz':
			icssCsvOutput.write('YZ,')
			for yStep in yBqSet:
				icssCsvOutput.write(f'{yStep},')
			icssCsvOutput.write('\n')
			zCount = 0
			for zNo in range(len(zBqSet)):
				if zCount < len(zBqSet):
					icssCsvOutput.write(f'{zBqSet[zCount]},')
				for yNo in range(len(yBqSet)):
					icssCsvOutput.write(f'{mapValue[zNo + len(zBqSet) * yNo]},')
				zCount += 1
				icssCsvOutput.write('\n')

		icssCsvOutput.close()

	def clrHmap(self):
		stepsizeLine.setText('10')
		emphasisLine.setText('')
		cLabelChk.setCheckState(Qt.CheckState.Unchecked)
		conLineChk.setCheckState(Qt.CheckState.Unchecked)
		self.reDraw2DOut(mapValue)

'''
*************************************************************************

                         3D NICS OUTPUT WINDOW

*************************************************************************
'''
class NICS3DWindow(QMainWindow):
	def __init__(self):
		super().__init__()

		self.setWindowTitle('3D NICS Output: ' + fileName.split('/')[-1])

		global isoTen, aniTen, xxTen, yxTen, zxTen, xyTen, yyTen, zyTen, xzTen, yzTen, zzTen, bqXList, bqYList, bqZList
		isoTen, aniTen, xxTen, yxTen, zxTen, xyTen, yyTen, zyTen, xzTen, yzTen, zzTen, bqXList, bqYList, bqZList = NICSout.save_tensor_3d(fileName, int(configFile.get('general', 'icss')))
		
		global xSet3d, ySet3d, zSet3d
		xSet3d = sorted(set(bqXList))
		ySet3d = sorted(set(bqYList))
		zSet3d = sorted(set(bqZList))

		n3OutMainWidget = QWidget()
		n3OutMainLayout = QHBoxLayout()
		n3OutMainWidget.setLayout(n3OutMainLayout)

		global n3cmapCombo
		n3cmapCombo = QComboBox()
		n3cmapCombo.addItems(CONSTANT.mplCmap)
		n3cmapCombo.setEditable(True)
		n3cmapCombo.setCurrentText('viridis')

		global n3lcolorCombo
		n3lcolorCombo = QComboBox()
		n3lcolorCombo.addItems(CONSTANT.molColor)
		n3lcolorCombo.setEditable(True)
		n3lcolorCombo.setCurrentText('white')

		global n3StepsizeLine, n3emphasisLine
		n3StepsizeLine = QLineEdit()
		n3emphasisLine = QLineEdit()

		global n3ClabelChk, n3ContourChk
		n3ClabelChk = QCheckBox('Show Contour Label')
		n3ContourChk = QCheckBox('Show Contour Line')
		n3ContourChk.setCheckState(Qt.CheckState.Unchecked)
		n3ContourChk.stateChanged.connect(self.conN3ChkFunc)

		global molView12
		molView12 = Mpl3DCanvas()

		if xMax == xMin:
			molView12.ax.set_xlim(xMin - 1, xMin + 1)
			molView12.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView12.ax.set_ylim(yMin - 1, yMin + 1)
			molView12.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView12.ax.set_zlim(zMin - 1, zMin + 1)
			molView12.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView12.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_s in bndAtom:
			molView12.ax.plot([xCoor[bndAtom_s[0]], xCoor[bndAtom_s[1]]], [yCoor[bndAtom_s[0]], yCoor[bndAtom_s[1]]], \
				[zCoor[bndAtom_s[0]], zCoor[bndAtom_s[1]]], '0.5')
		molView12.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		# Poly surface 1
		polyX1 = [max(bqXList), max(bqXList), max(bqXList), max(bqXList)]
		polyY1 = [min(bqYList), max(bqYList), max(bqYList), min(bqYList)]
		polyZ1 = [min(bqZList), min(bqZList), max(bqZList), max(bqZList)]
		poly1 = list(zip(polyX1, polyY1, polyZ1))
		molView12.ax.add_collection3d(Poly3DCollection([poly1], facecolor = '#000080', alpha = 0.1))
		# Poly surface 2
		polyX2 = [max(bqXList), min(bqXList), min(bqXList), max(bqXList)]
		polyY2 = [max(bqYList), max(bqYList), max(bqYList), max(bqYList)]
		polyZ2 = [min(bqZList), min(bqZList), max(bqZList), max(bqZList)]
		poly2 = list(zip(polyX2, polyY2, polyZ2))
		molView12.ax.add_collection3d(Poly3DCollection([poly2], facecolor = '#55559e', alpha = 0.1))
		# Poly surface 3
		polyX3 = [min(bqXList), min(bqXList), min(bqXList), min(bqXList)]
		polyY3 = [min(bqYList), max(bqYList), max(bqYList), min(bqYList)]
		polyZ3 = [min(bqZList), min(bqZList), max(bqZList), max(bqZList)]
		poly3 = list(zip(polyX3, polyY3, polyZ3))
		molView12.ax.add_collection3d(Poly3DCollection([poly3], facecolor = '#000080', alpha = 0.1))
		# Poly surface 4
		polyX4 = [max(bqXList), min(bqXList), min(bqXList), max(bqXList)]
		polyY4 = [min(bqYList), min(bqYList), min(bqYList), min(bqYList)]
		polyZ4 = [min(bqZList), min(bqZList), max(bqZList), max(bqZList)]
		poly4 = list(zip(polyX4, polyY4, polyZ4))
		molView12.ax.add_collection3d(Poly3DCollection([poly4], facecolor = '#55559e', alpha = 0.1))
		# Poly surface 5
		polyX5 = [max(bqXList), max(bqXList), min(bqXList), min(bqXList)]
		polyY5 = [min(bqYList), max(bqYList), max(bqYList), min(bqYList)]
		polyZ5 = [min(bqZList), min(bqZList), min(bqZList), min(bqZList)]
		poly5 = list(zip(polyX5, polyY5, polyZ5))
		molView12.ax.add_collection3d(Poly3DCollection([poly5], facecolor = '#3939e3', alpha = 0.1))
		# Poly surface 6
		polyX6 = [max(bqXList), max(bqXList), min(bqXList), min(bqXList)]
		polyY6 = [min(bqYList), max(bqYList), max(bqYList), min(bqYList)]
		polyZ6 = [max(bqZList), max(bqZList), max(bqZList), max(bqZList)]
		poly6 = list(zip(polyX6, polyY6, polyZ6))
		molView12.ax.add_collection3d(Poly3DCollection([poly6], facecolor = '#3939e3', alpha = 0.1))

		polyXn3_1 = [min(xSet3d), max(xSet3d), max(xSet3d), min(xSet3d)]
		polyYn3_1 = [min(ySet3d), min(ySet3d), max(ySet3d), max(ySet3d)]
		polyZn3_1 = [zSet3d[len(zSet3d)//2], zSet3d[len(zSet3d)//2], zSet3d[len(zSet3d)//2], zSet3d[len(zSet3d)//2]]
		polyN3_1 = list(zip(polyXn3_1, polyYn3_1, polyZn3_1))
		molView12.ax.add_collection3d(Poly3DCollection([polyN3_1], facecolor = '#000080', alpha = 0.3))

		toolBar13 = NavigationToolbar(molView12, self)

		viewLayout12 = QVBoxLayout()
		viewLayout12.addWidget(molView12)
		viewLayout12.addWidget(toolBar13)
		viewLayout12.setContentsMargins(0, 0, 0, 0)

		viewer12 = QWidget()
		viewer12.setLayout(viewLayout12)

		global n3n2OutMpl
		n3n2OutMpl = MplN2OutCanvas()
		toolBar14 = NavigationToolbar(n3n2OutMpl, self)

		global n3HeiXCombo, n3HeiYCombo, n3HeiZCombo
		n3HeiXCombo = QComboBox()
		n3HeiXCombo.addItems([str(v) for v in xSet3d])
		n3HeiXCombo.setEditable(True)
		n3HeiXCombo.setEnabled(False)
		n3HeiXCombo.setCurrentText(str(xSet3d[len(xSet3d)//2]))

		n3HeiYCombo = QComboBox()
		n3HeiYCombo.addItems([str(v) for v in ySet3d])
		n3HeiYCombo.setEditable(True)
		n3HeiYCombo.setEnabled(False)
		n3HeiYCombo.setCurrentText(str(ySet3d[len(ySet3d)//2]))

		n3HeiZCombo = QComboBox()
		n3HeiZCombo.addItems([str(v) for v in zSet3d])
		n3HeiZCombo.setEditable(True)
		n3HeiZCombo.setEnabled(True)
		n3HeiZCombo.setCurrentText(str(zSet3d[len(zSet3d)//2]))

		icssn3Array = []
		global icssn3Fig, cBn3
		mapn3Value = NICSout.extract_2d(zzTen, 'XY', float(n3HeiZCombo.currentText()), xSet3d, ySet3d, zSet3d)
		for n_y in range(len(ySet3d)):
			icssn3ArrayLine = []
			for n_x in range(len(xSet3d)):
				icssn3ArrayLine.append(mapn3Value[n_y + len(ySet3d) * n_x])
			icssn3Array.append(icssn3ArrayLine)
		x2Dn3value, y2Dn3value = np.meshgrid(xSet3d, ySet3d)
		icssn3Fig = n3n2OutMpl.ax.pcolormesh(x2Dn3value, y2Dn3value, np.array(icssn3Array), shading = 'gouraud', cmap = n3cmapCombo.currentText())
		cBn3 = figN2Out.colorbar(icssn3Fig, cax = make_axes_locatable(n3n2OutMpl.ax).append_axes('right', size = '5%', pad = 0.05))
		cBn3.mappable.set_clim(min(mapn3Value), max(mapn3Value))
		n3n2OutMpl.ax.set_xlabel('X / Å')
		n3n2OutMpl.ax.set_ylabel('Y / Å')

		n3n2PreWidget = QWidget()
		n3n2PreviewLayout = QVBoxLayout()
		n3n2PreWidget.setLayout(n3n2PreviewLayout)

		n3n2PreviewLayout.addWidget(n3n2OutMpl)
		n3n2PreviewLayout.addWidget(toolBar14)
		n3n2PreviewLayout.setContentsMargins(0, 0, 0, 0)

		n3OutResult = QWidget()
		n3OutResultLayout = QVBoxLayout()
		n3OutResult.setLayout(n3OutResultLayout)

		radioBtnGroup = QGroupBox('Component of Shielding Tensors')
		radioBtnLayout = QGridLayout()
		radioBtnGroup.setLayout(radioBtnLayout)

		radion5Group = QButtonGroup()
		isoBtn = QRadioButton('Isotropy', self)
		isoBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(isoBtn, 0, 0)
		aniBtn = QRadioButton('Anisotropy', self)
		aniBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(aniBtn, 0, 1)
		xxBtn = QRadioButton('XX', self)
		xxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xxBtn, 1, 0)
		xyBtn = QRadioButton('XY', self)
		xyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xyBtn, 2, 0)
		xzBtn = QRadioButton('XZ', self)
		xzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xzBtn, 3, 0)
		yxBtn = QRadioButton('YX', self)
		yxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yxBtn, 1, 1)
		yyBtn = QRadioButton('YY', self)
		yyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yyBtn, 2, 1)
		yzBtn = QRadioButton('YZ', self)
		yzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yzBtn, 3, 1)
		zxBtn = QRadioButton('ZX', self)
		zxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zxBtn, 1, 2)
		zyBtn = QRadioButton('ZY', self)
		zyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zyBtn, 2, 2)
		zzBtn = QRadioButton('ZZ', self)
		zzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zzBtn, 3, 2)

		radion5Group.addButton(isoBtn, 1)
		radion5Group.addButton(aniBtn, 2)
		radion5Group.addButton(xxBtn, 3)
		radion5Group.addButton(yzBtn, 4)
		radion5Group.addButton(zxBtn, 5)
		radion5Group.addButton(xyBtn, 6)
		radion5Group.addButton(yyBtn, 7)
		radion5Group.addButton(zyBtn, 8)
		radion5Group.addButton(xzBtn, 9)
		radion5Group.addButton(yzBtn, 10)
		radion5Group.addButton(zzBtn, 11)

		zzBtn.setChecked(True)

		n3OutResultLayout.addWidget(radioBtnGroup)

		n3OutBtnWidget = QWidget()
		n3OutBtnLayout = QHBoxLayout()
		n3OutBtnWidget.setLayout(n3OutBtnLayout)

		saveCubBtn = QPushButton('Save .cube')
		saveCubBtn.clicked.connect(self.saveCubFunc)
		n3OutBtnLayout.addWidget(saveCubBtn)

		n3OutResultLayout.addWidget(n3OutBtnWidget)

		n2PlaneGroup = QGroupBox('2D NICS Parameters')
		n2PlaneLayout = QGridLayout()
		n2PlaneGroup.setLayout(n2PlaneLayout)

		n2PlaneLayout.addWidget(n3HeiXCombo, 1, 1)
		n2PlaneLayout.addWidget(n3HeiYCombo, 1, 3)
		n2PlaneLayout.addWidget(n3HeiZCombo, 2, 1)

		n2PlaneBtnGourp = QButtonGroup()
		n3XYBtn = QRadioButton('XY', self)
		n3XYBtn.toggled.connect(self.planeClicked)
		n2PlaneLayout.addWidget(n3XYBtn, 0, 1)
		n3YZBtn = QRadioButton('YZ', self)
		n3YZBtn.toggled.connect(self.planeClicked)
		n2PlaneLayout.addWidget(n3YZBtn, 0, 2)
		n3XZBtn = QRadioButton('XZ', self)
		n3XZBtn.toggled.connect(self.planeClicked)
		n2PlaneLayout.addWidget(n3XZBtn, 0, 3)
		n2PlaneBtnGourp.addButton(n3XYBtn, 1)
		n2PlaneBtnGourp.addButton(n3YZBtn, 2)
		n2PlaneBtnGourp.addButton(n3XZBtn, 3)
		n3XYBtn.setChecked(True)

		n2PlaneLayout.addWidget(QLabel('Plane: '), 0, 0)
		n2PlaneLayout.addWidget(QLabel('Height (X): '), 1, 0)
		n2PlaneLayout.addWidget(QLabel('Height (Y): '), 1, 2)
		n2PlaneLayout.addWidget(QLabel('Height (Z): '), 2, 0)

		n3OutResultLayout.addWidget(n2PlaneGroup)

		n3n2Group = QGroupBox('2D NICS Style')
		n3n2Layout = QGridLayout()
		n3n2Group.setLayout(n3n2Layout)

		n3n2Layout.addWidget(n3ContourChk, 2, 0)

		n3n2Layout.addWidget(n3ClabelChk, 2, 1)

		n3n2Layout.addWidget(QLabel('	Stepsize: '), 3, 0)
		n3n2Layout.addWidget(QLabel('	Emphasize: '), 4, 0)
		n3n2Layout.addWidget(QLabel('Heatmap Color Style: '), 5, 0)
		n3n2Layout.addWidget(QLabel('Contour Line Color: '), 6, 0)

		n3n2Layout.addWidget(n3StepsizeLine, 3, 1)
		n3StepsizeLine.setText('10')
		n3n2Layout.addWidget(n3emphasisLine, 4, 1)
		n3n2Layout.addWidget(n3cmapCombo, 5, 1)
		n3n2Layout.addWidget(n3lcolorCombo, 6, 1)

		n3n2Update2 = QPushButton('Update')
		n3n2Update2.clicked.connect(self.updateHmapN3)
		n3n2Layout.addWidget(n3n2Update2, 7, 0)

		n3n2Clr = QPushButton('Clear')
		n3n2Clr.clicked.connect(self.clrDraw3DPre)
		n3n2Layout.addWidget(n3n2Clr, 7, 1)

		n3n2Png = QPushButton('Save .png')
		n3n2Png.clicked.connect(self.n3SavePng)
		n3n2Layout.addWidget(n3n2Png, 8, 0)

		n3n2Csv = QPushButton('Save .csv')
		n3n2Csv.clicked.connect(self.n3SaveCsv)
		n3n2Layout.addWidget(n3n2Csv, 8, 1)

		n3OutResultLayout.addWidget(n3n2Group)

		status3DBox = QTextEdit()
		status3DBox.setEnabled(False)
		n3OutResultLayout.addWidget(status3DBox)

		n3OutMainLayout.addWidget(viewer12)
		n3OutMainLayout.addWidget(n3n2PreWidget)
		n3OutMainLayout.addWidget(n3OutResult)
		n3OutMainWidget.setLayout(n3OutMainLayout)

		self.setCentralWidget(n3OutMainWidget)
		self.setMinimumSize(1500, 600)

	def reDrawn32DOut(self, map_value):
		n3n2OutMpl.ax.cla()
		icssn3Array = []
		if n3HeiZCombo.isEnabled() == True:
			for n_y in range(len(ySet3d)):
				icssn3ArrayLine = []
				for n_x in range(len(xSet3d)):
					icssn3ArrayLine.append(map_value[n_y + len(ySet3d) * n_x])
				icssn3Array.append(icssn3ArrayLine)
			x2Dn3value, y2Dn3value = np.meshgrid(xSet3d, ySet3d)
			icssn3Fig = n3n2OutMpl.ax.pcolormesh(x2Dn3value, y2Dn3value, np.array(icssn3Array), shading = 'gouraud', cmap = n3cmapCombo.currentText())
			cBn3.mappable.set_clim(min(map_value), max(map_value))
			cBn3.mappable.set_cmap(n3cmapCombo.currentText())
			conn3X, conn3Y = np.meshgrid(xSet3d, ySet3d)
			if configFile.get('heatmap', 'cb') == 'true':
				n3n2OutMpl.ax.contour(conn3X, conn3Y, np.array(icssn3Array), colors = n3lcolorCombo.currentText(), linewidths = 0.3, levels = int(n3StepsizeLine.text()))
			if n3emphasisLine.text() != '':
				emn3 = n3n2OutMpl.ax.contour(conn3X, conn3Y, np.array(icssn3Array), colors = n3lcolorCombo.currentText(), linestyles = '-', linewidths = 1.0, levels = [float(n3emphasisLine.text())])
				if n3ClabelChk.isChecked():
					n3n2OutMpl.ax.clabel(emn3, inline = True)
			n3n2OutMpl.ax.set_xlabel('X / Å')
			n3n2OutMpl.ax.set_ylabel('Y / Å')

		elif n3HeiYCombo.isEnabled() == True:
			for n_z in range(len(zSet3d)):
				icssn3ArrayLine = []
				for n_x in range(len(xSet3d)):
					icssn3ArrayLine.append(map_value[n_z + len(zSet3d) * n_x])
				icssn3Array.append(icssn3ArrayLine)
			x2Dn3value, z2Dn3value = np.meshgrid(xSet3d, zSet3d)
			icssn3Fig = n3n2OutMpl.ax.pcolormesh(x2Dn3value, z2Dn3value, np.array(icssn3Array), shading = 'gouraud', cmap = n3cmapCombo.currentText())
			cBn3.mappable.set_clim(min(map_value), max(map_value))
			cBn3.mappable.set_cmap(n3cmapCombo.currentText())
			conn3X, conn3Z = np.meshgrid(xSet3d, zSet3d)
			if configFile.get('heatmap', 'cb') == 'true':
				n3n2OutMpl.ax.contour(conn3X, conn3Z, np.array(icssn3Array), colors = n3lcolorCombo.currentText(), linewidths = 0.3, levels = int(n3StepsizeLine.text()))
			if n3emphasisLine.text() != '':
				emn3 = n3n2OutMpl.ax.contour(conn3X, conn3Z, np.array(icssn3Array), colors = n3lcolorCombo.currentText(), linestyles = '-', linewidths = 1.0, levels = [float(n3emphasisLine.text())])
				if n3ClabelChk.isChecked():
					n3n2OutMpl.ax.clabel(emn3, inline = True)
			n3n2OutMpl.ax.set_xlabel('X / Å')
			n3n2OutMpl.ax.set_ylabel('Z / Å')

		elif n3HeiXCombo.isEnabled() == True:
			for n_z in range(len(zSet3d)):
				icssn3ArrayLine = []
				for n_y in range(len(ySet3d)):
					icssn3ArrayLine.append(map_value[n_z + len(zSet3d) * n_y])
				icssn3Array.append(icssn3ArrayLine)
			y2Dn3value, z2Dn3value = np.meshgrid(ySet3d, zSet3d)
			icssn3Fig = n3n2OutMpl.ax.pcolormesh(y2Dn3value, z2Dn3value, np.array(icssn3Array), shading = 'gouraud', cmap = n3cmapCombo.currentText())
			cBn3.mappable.set_clim(min(map_value), max(map_value))
			cBn3.mappable.set_cmap(n3cmapCombo.currentText())
			conn3Y, conn3Z = np.meshgrid(ySet3d, zSet3d)
			if configFile.get('heatmap', 'cb') == 'true':
				n3n2OutMpl.ax.contour(conn3Y, conn3Z, np.array(icssn3Array), colors = n3lcolorCombo.currentText(), linewidths = 0.3, levels = int(n3StepsizeLine.text()))
			if n3emphasisLine.text() != '':
				emn3 = n3n2OutMpl.ax.contour(conn3Y, conn3Z, np.array(icssn3Array), colors = n3lcolorCombo.currentText(), linestyles = '-', linewidths = 1.0, levels = [float(n3emphasisLine.text())])
				if n3ClabelChk.isChecked():
					n3n2OutMpl.ax.clabel(emn3, inline = True)
			n3n2OutMpl.ax.set_xlabel('Y / Å')
			n3n2OutMpl.ax.set_ylabel('Z / Å')

		n3n2OutMpl.draw()

	def reDraw3DPre(self):
		molView12.ax.cla()
		molView12.ax.set_xlabel('X (Å)')
		molView12.ax.set_ylabel('Y (Å)')
		molView12.ax.set_zlabel('Z (Å)')
		molView12.ax.grid(False)

		if xMax == xMin:
			molView12.ax.set_xlim(xMin - 1, xMin + 1)
			molView12.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView12.ax.set_ylim(yMin - 1, yMin + 1)
			molView12.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView12.ax.set_zlim(zMin - 1, zMin + 1)
			molView12.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView12.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])

		for bndAtom_s in bndAtom:
			molView12.ax.plot([xCoor[bndAtom_s[0]], xCoor[bndAtom_s[1]]], [yCoor[bndAtom_s[0]], yCoor[bndAtom_s[1]]], \
				[zCoor[bndAtom_s[0]], zCoor[bndAtom_s[1]]], '0.5')
		molView12.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		# Poly surface 1
		polyX1 = [max(bqXList), max(bqXList), max(bqXList), max(bqXList)]
		polyY1 = [min(bqYList), max(bqYList), max(bqYList), min(bqYList)]
		polyZ1 = [min(bqZList), min(bqZList), max(bqZList), max(bqZList)]
		poly1 = list(zip(polyX1, polyY1, polyZ1))
		molView12.ax.add_collection3d(Poly3DCollection([poly1], facecolor = '#000080', alpha = 0.1))
		# Poly surface 2
		polyX2 = [max(bqXList), min(bqXList), min(bqXList), max(bqXList)]
		polyY2 = [max(bqYList), max(bqYList), max(bqYList), max(bqYList)]
		polyZ2 = [min(bqZList), min(bqZList), max(bqZList), max(bqZList)]
		poly2 = list(zip(polyX2, polyY2, polyZ2))
		molView12.ax.add_collection3d(Poly3DCollection([poly2], facecolor = '#55559e', alpha = 0.1))
		# Poly surface 3
		polyX3 = [min(bqXList), min(bqXList), min(bqXList), min(bqXList)]
		polyY3 = [min(bqYList), max(bqYList), max(bqYList), min(bqYList)]
		polyZ3 = [min(bqZList), min(bqZList), max(bqZList), max(bqZList)]
		poly3 = list(zip(polyX3, polyY3, polyZ3))
		molView12.ax.add_collection3d(Poly3DCollection([poly3], facecolor = '#000080', alpha = 0.1))
		# Poly surface 4
		polyX4 = [max(bqXList), min(bqXList), min(bqXList), max(bqXList)]
		polyY4 = [min(bqYList), min(bqYList), min(bqYList), min(bqYList)]
		polyZ4 = [min(bqZList), min(bqZList), max(bqZList), max(bqZList)]
		poly4 = list(zip(polyX4, polyY4, polyZ4))
		molView12.ax.add_collection3d(Poly3DCollection([poly4], facecolor = '#55559e', alpha = 0.1))
		# Poly surface 5
		polyX5 = [max(bqXList), max(bqXList), min(bqXList), min(bqXList)]
		polyY5 = [min(bqYList), max(bqYList), max(bqYList), min(bqYList)]
		polyZ5 = [min(bqZList), min(bqZList), min(bqZList), min(bqZList)]
		poly5 = list(zip(polyX5, polyY5, polyZ5))
		molView12.ax.add_collection3d(Poly3DCollection([poly5], facecolor = '#3939e3', alpha = 0.1))
		# Poly surface 6
		polyX6 = [max(bqXList), max(bqXList), min(bqXList), min(bqXList)]
		polyY6 = [min(bqYList), max(bqYList), max(bqYList), min(bqYList)]
		polyZ6 = [max(bqZList), max(bqZList), max(bqZList), max(bqZList)]
		poly6 = list(zip(polyX6, polyY6, polyZ6))
		molView12.ax.add_collection3d(Poly3DCollection([poly6], facecolor = '#3939e3', alpha = 0.1))

		if n3HeiZCombo.isEnabled() == True:
			polyXn3 = [min(xSet3d), max(xSet3d), max(xSet3d), min(xSet3d)]
			polyYn3 = [min(ySet3d), min(ySet3d), max(ySet3d), max(ySet3d)]
			polyZn3 = [float(n3HeiZCombo.currentText()), float(n3HeiZCombo.currentText()), float(n3HeiZCombo.currentText()), float(n3HeiZCombo.currentText())]
			polyN3 = list(zip(polyXn3, polyYn3, polyZn3))
			molView12.ax.add_collection3d(Poly3DCollection([polyN3], facecolor = '#000080', alpha = 0.3))
		elif n3HeiYCombo.isEnabled() == True:
			polyXn3 = [min(xSet3d), max(xSet3d), max(xSet3d), min(xSet3d)]
			polyYn3 = [float(n3HeiYCombo.currentText()), float(n3HeiYCombo.currentText()), float(n3HeiYCombo.currentText()), float(n3HeiYCombo.currentText())]
			polyZn3 = [min(zSet3d), min(zSet3d), max(zSet3d), max(zSet3d)]
			polyN3 = list(zip(polyXn3, polyYn3, polyZn3))
			molView12.ax.add_collection3d(Poly3DCollection([polyN3], facecolor = '#000080', alpha = 0.3))
		elif n3HeiXCombo.isEnabled() == True:
			polyXn3 = [float(n3HeiXCombo.currentText()), float(n3HeiXCombo.currentText()), float(n3HeiXCombo.currentText()), float(n3HeiXCombo.currentText())]
			polyYn3 = [min(ySet3d), max(ySet3d), max(ySet3d), min(ySet3d)]
			polyZn3 = [min(zSet3d), min(zSet3d), max(zSet3d), max(zSet3d)]
			polyN3 = list(zip(polyXn3, polyYn3, polyZn3))
			molView12.ax.add_collection3d(Poly3DCollection([polyN3], facecolor = '#000080', alpha = 0.3))
		
		molView12.draw()

	def updateHmapN3(self):
		if n3n2Plane == 'XY':
			mapn3Value = NICSout.extract_2d(cubValue, n3n2Plane, float(n3HeiZCombo.currentText()), xSet3d, ySet3d, zSet3d)
		elif n3n2Plane == 'XZ':
			mapn3Value = NICSout.extract_2d(cubValue, n3n2Plane, float(n3HeiYCombo.currentText()), xSet3d, ySet3d, zSet3d)
		elif n3n2Plane == 'YZ':
			mapn3Value = NICSout.extract_2d(cubValue, n3n2Plane, float(n3HeiXCombo.currentText()), xSet3d, ySet3d, zSet3d)
		self.reDrawn32DOut(mapn3Value)
		self.reDraw3DPre()

	def tensorClicked(self):
		global tensorType
		tensorType = ''
		if self.sender().isChecked():
			tensorType = self.sender().text()
		global cubValue
		cubValue = zzTen
		
		if tensorType == 'Isotropy':
			cubValue = isoTen
		elif tensorType == 'Anisotropy':
			cubValue = aniTen
		elif tensorType == 'XX':
			cubValue = xxTen
		elif tensorType == 'YX':
			cubValue = yxTen
		elif tensorType == 'ZX':
			cubValue = zxTen
		elif tensorType == 'XY':
			cubValue = xyTen
		elif tensorType == 'YY':
			cubValue = yyTen
		elif tensorType == 'ZY':
			cubValue = zyTen
		elif tensorType == 'XZ':
			cubValue = xzTen
		elif tensorType == 'YZ':
			cubValue = yzTen
		elif tensorType == 'ZZ':
			cubValue = zzTen

		if n3HeiZCombo.isEnabled() == True:
			mapn3Value = NICSout.extract_2d(cubValue, 'XY', float(n3HeiZCombo.currentText()), xSet3d, ySet3d, zSet3d)
		elif n3HeiYCombo.isEnabled() == True:
			mapn3Value = NICSout.extract_2d(cubValue, 'XZ', float(n3HeiYCombo.currentText()), xSet3d, ySet3d, zSet3d)
		elif n3HeiXCombo.isEnabled() == True:
			mapn3Value = NICSout.extract_2d(cubValue, 'YZ', float(n3HeiXCombo.currentText()), xSet3d, ySet3d, zSet3d)

		self.reDrawn32DOut(mapn3Value)

	def saveCubFunc(self):
		if configFile.get('general', 'icss') == '-1':
			cubeFile = open(f'{os.path.splitext(fileName)[0][:-4]}3D_NICS_output_{tensorType.upper()}.cub', 'w')
		else:
			cubeFile = open(f'{os.path.splitext(fileName)[0][:-4]}3D_ICSS_output_{tensorType.upper()}.cub', 'w')

		numCount = 0

		cubeFile.write('File Created by py.Aroma 4, Developed by Zhe Wang (Ph.D.)\n')
		cubeFile.write(f'Sheilding tensors from {tensorType.upper()} component, totally {len(cubValue)} grid points.\n')

		cubeFile.write(f' {len(geomList)}   {min(bqXList) * 1.88973 :.6f}   {min(bqYList) * 1.88973 :.6f}   {min(bqZList) * 1.88973 :.6f}\n')
		cubeFile.write(f' {len(xSet3d)}   {(xSet3d[1] - xSet3d[0]) * 1.88973 :.6f}    0.000000   0.000000\n')
		cubeFile.write(f' {len(ySet3d)}    0.000000   {(ySet3d[1] - ySet3d[0]) * 1.88973 :.6f}   0.000000\n')
		cubeFile.write(f' {len(zSet3d)}    0.000000   0.000000   {(zSet3d[1] - zSet3d[0]) * 1.88973 :.6f}\n')

		for atom_l in geomList:
			cubeFile.write(f'{CONSTANT.period_table.index(atom_l[0].upper())}    {CONSTANT.period_table.index(atom_l[0].upper()) :.6f}    ')
			cubeFile.write(f'{atom_l[1] * 1.88973 :.6f}    {atom_l[2] * 1.88973 :.6f}    {atom_l[3] * 1.88973 :.6f}\n')

		for num in cubValue:
			cubeFile.write('%e' % num)
			cubeFile.write('  ')
			numCount += 1
			if numCount % 6 == 0:
				cubeFile.write('\n')

		cubeFile.close()

	def clrDraw3DPre(self):
		self.reDraw3DPre()

		n3ContourChk.setCheckState(Qt.CheckState.Unchecked)
		n3ClabelChk.setCheckState(Qt.CheckState.Unchecked)
		n3StepsizeLine.setText('10')
		n3emphasisLine.setText('')
		n3cmapCombo.setCurrentText('viridis')
		n3lcolorCombo.setCurrentText('white')

		if n3n2Plane == 'XY':
			mapn3Value = NICSout.extract_2d(cubValue, n3n2Plane, float(n3HeiZCombo.currentText()), xSet3d, ySet3d, zSet3d)
		elif n3n2Plane == 'XZ':
			mapn3Value = NICSout.extract_2d(cubValue, n3n2Plane, float(n3HeiYCombo.currentText()), xSet3d, ySet3d, zSet3d)
		elif n3n2Plane == 'YZ':
			mapn3Value = NICSout.extract_2d(cubValue, n3n2Plane, float(n3HeiXCombo.currentText()), xSet3d, ySet3d, zSet3d)
		self.reDrawn32DOut(mapn3Value)

	def planeClicked(self):
		global n3n2Plane
		n3n2Plane = ''
		if self.sender().isChecked():
			n3n2Plane = self.sender().text()

		if n3n2Plane == 'XY':
			n3HeiXCombo.setEnabled(False)
			n3HeiYCombo.setEnabled(False)
			n3HeiZCombo.setEnabled(True)
		elif n3n2Plane == 'YZ':
			n3HeiXCombo.setEnabled(True)
			n3HeiYCombo.setEnabled(False)
			n3HeiZCombo.setEnabled(False)
		elif n3n2Plane == 'XZ':
			n3HeiXCombo.setEnabled(False)
			n3HeiYCombo.setEnabled(True)
			n3HeiZCombo.setEnabled(False)

	def n3SaveCsv(self):
		if configFile.get('general', 'icss') == '-1':
			icssn3CsvOutput = open(f'{os.path.splitext(fileName)[0]}_2D_NICS_output_from_3D_NICS_output_{tensorType.upper()}.csv', 'w')
		else:
			icssn3CsvOutput = open(f'{os.path.splitext(fileName)[0]}_2D_ICSS_output_from_3D_ICSS_output_{tensorType.upper()}.csv', 'w')

		icssn3CsvOutput.write('#\n#   File Created by py.Aroma 4\n')
		icssn3CsvOutput.write('#   Author: Zhe Wang (Ph.D.)\n')
		icssn3CsvOutput.write('#   https://wongzit.github.com/program/pyaroma/\n#\n\n')

		if n3n2Plane == 'XY':
			mapn3Value = NICSout.extract_2d(cubValue, n3n2Plane, float(n3HeiZCombo.currentText()), xSet3d, ySet3d, zSet3d)
			icssn3CsvOutput.write('XY,')
			for xStep in xSet3d:
				icssn3CsvOutput.write(f'{xStep},')
			icssn3CsvOutput.write('\n')
			yCount = 0
			for yNo in range(len(ySet3d)):
				if yCount < len(ySet3d):
					icssn3CsvOutput.write(f'{ySet3d[yCount]},')
				for xNo in range(len(xSet3d)):
					icssn3CsvOutput.write(f'{mapn3Value[yNo + len(ySet3d) * xNo]},')
				yCount += 1
				icssn3CsvOutput.write('\n')

		elif n3n2Plane == 'XZ':
			mapn3Value = NICSout.extract_2d(cubValue, n3n2Plane, float(n3HeiYCombo.currentText()), xSet3d, ySet3d, zSet3d)
			icssn3CsvOutput.write('XZ,')
			for xStep in xSet3d:
				icssn3CsvOutput.write(f'{xStep},')
			icssn3CsvOutput.write('\n')
			zCount = 0
			for zNo in range(len(zSet3d)):
				if zCount < len(zSet3d):
					icssn3CsvOutput.write(f'{zSet3d[zCount]},')
				for xNo in range(len(xSet3d)):
					icssn3CsvOutput.write(f'{mapn3Value[zNo + len(zSet3d) * xNo]},')
				zCount += 1
				icssn3CsvOutput.write('\n')

		elif n3n2Plane == 'YZ':
			mapn3Value = NICSout.extract_2d(cubValue, n3n2Plane, float(n3HeiXCombo.currentText()), xSet3d, ySet3d, zSet3d)
			icssn3CsvOutput.write('YZ,')
			for yStep in ySet3d:
				icssn3CsvOutput.write(f'{yStep},')
			icssn3CsvOutput.write('\n')
			zCount = 0
			for zNo in range(len(zSet3d)):
				if zCount < len(zSet3d):
					icssn3CsvOutput.write(f'{zSet3d[zCount]},')
				for yNo in range(len(ySet3d)):
					icssn3CsvOutput.write(f'{mapn3Value[zNo + len(zSet3d) * yNo]},')
				zCount += 1
				icssn3CsvOutput.write('\n')

		icssn3CsvOutput.close()

	def conN3ChkFunc(self):
		if n3n2Plane == 'XY':
			mapn3Value = NICSout.extract_2d(cubValue, n3n2Plane, float(n3HeiZCombo.currentText()), xSet3d, ySet3d, zSet3d)
		elif n3n2Plane == 'XZ':
			mapn3Value = NICSout.extract_2d(cubValue, n3n2Plane, float(n3HeiYCombo.currentText()), xSet3d, ySet3d, zSet3d)
		elif n3n2Plane == 'YZ':
			mapn3Value = NICSout.extract_2d(cubValue, n3n2Plane, float(n3HeiXCombo.currentText()), xSet3d, ySet3d, zSet3d)

		if n3ContourChk.isChecked():
			configFile.set('heatmap', 'cb', 'true')
			self.reDrawn32DOut(mapn3Value)
		else:
			configFile.set('heatmap', 'cb', 'false')
			self.reDrawn32DOut(mapn3Value)

	def n3SavePng(self):
		if configFile.get('general', 'icss') == '-1':
			figN2Out.savefig(f'{os.path.splitext(fileName)[0]}_2D_NICS_from_3D_NICS_output_{tensorType.upper()}.png', dpi = 300)
		else:
			figN2Out.savefig(f'{os.path.splitext(fileName)[0]}_2D_ICSS_from_3D_ICSS_output_{tensorType.upper()}.png', dpi = 300)

'''
*************************************************************************

                          INICS OUTPUT WINDOW

*************************************************************************
'''
class INICSWindow(QMainWindow):
	def __init__(self):
		super().__init__()

		self.setWindowTitle('INICS Output: ' + fileName.split('/')[-1])

		logBqList = atomList[len(geomList) - len(atomList):]
		xBqCoor, yBqCoor, zBqCoor, colorBqList = geomAnalyzer.save_coor_list(logBqList)

		global ringNums
		ringNums = len(geomAnalyzer.find_monocycle(bndAtom))

		inicsOutMainWidget = QWidget()
		inicsOutMainLayout = QHBoxLayout()

		molView20 = Mpl3DCanvas()

		if xMax == xMin:
			molView20.ax.set_xlim(xMin - 1, xMin + 1)
			molView20.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView20.ax.set_ylim(yMin - 1, yMin + 1)
			molView20.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView20.ax.set_zlim(zMin - 1, zMin + 1)
			molView20.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView20.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_r in bndAtom:
			molView20.ax.plot([xCoor[bndAtom_r[0]], xCoor[bndAtom_r[1]]], [yCoor[bndAtom_r[0]], yCoor[bndAtom_r[1]]], \
				[zCoor[bndAtom_r[0]], zCoor[bndAtom_r[1]]], '0.5')
		molView20.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		toolBar20 = NavigationToolbar(molView20, self)

		viewLayout20 = QVBoxLayout()
		viewLayout20.addWidget(molView20)
		viewLayout20.addWidget(toolBar20)
		viewLayout20.setContentsMargins(0, 0, 0, 0)

		viewer20 = QWidget()
		viewer20.setLayout(viewLayout20)

		inicsPreview = QWidget()
		inicspreLayout = QVBoxLayout()
		inicsPreview.setLayout(inicspreLayout)

		with open(fileName, 'r') as inicsFile:
			inicsLine = inicsFile.readlines()

		inicsRange = 0.0
		inicsSplit = 0.0
		numDots = 1
		for inicsline in inicsLine:
			if '// py.Aroma 4' in inicsline:
				inicsRange = float(inicsline.split()[1])
				inicsSplit = float(inicsline.split()[2])
				numDots = int(inicsline.split()[3])*2+1
				break
		global isoInt, aniInt, xxInt, yxInt, zxInt, xyInt, yyInt, zyInt, xzInt, yzInt, zzInt
		isoInt, aniInt, xxInt, yxInt, zxInt, xyInt, yyInt, zyInt, xzInt, yzInt, zzInt = NICSout.cal_inics(fileName, len(logBqList), len(geomAnalyzer.find_monocycle(bndAtom)), inicsSplit, numDots)

		global inicsHeightList
		inicsHeightList = [0]

		for inics_n in range(1, 999):
			if inics_n*inicsSplit <= inicsRange:
				inicsHeightList.append(float(inics_n*inicsSplit))
				inicsHeightList.insert(0, -float(inics_n*inicsSplit))
			else:
				break

		isoShiel, aniShiel, xxShiel, yxShiel, zxShiel, xyShiel, yyShiel, zyShiel, xzShiel, yzShiel, zzShiel = NICSout.save_tensor(fileName, len(logBqList), -1)

		global ringisoList, ringaniList, ringxxList, ringyxList, ringzxList, ringxyList, ringyyList, ringzyList, ringxzList, ringyzList, ringzzList
		ringisoList = []
		ringaniList = []
		ringxxList = []
		ringyxList = []
		ringzxList = []
		ringxyList = []
		ringyyList = []
		ringzyList = []
		ringxzList = []
		ringyzList = []
		ringzzList = []
		for nRing_2 in range(ringNums):
			monoIsoTensor = isoShiel[nRing_2*numDots:nRing_2*numDots+numDots]
			ringisoList.append(monoIsoTensor)
			monoAniTensor = aniShiel[nRing_2*numDots:nRing_2*numDots+numDots]
			ringaniList.append(monoAniTensor)
			monoxxTensor = xxShiel[nRing_2*numDots:nRing_2*numDots+numDots]
			ringxxList.append(monoxxTensor)
			monoyxTensor = yxShiel[nRing_2*numDots:nRing_2*numDots+numDots]
			ringyxList.append(monoyxTensor)
			monozxTensor = zxShiel[nRing_2*numDots:nRing_2*numDots+numDots]
			ringzxList.append(monozxTensor)
			monoxyTensor = xyShiel[nRing_2*numDots:nRing_2*numDots+numDots]
			ringxyList.append(monoxyTensor)
			monoyyTensor = yyShiel[nRing_2*numDots:nRing_2*numDots+numDots]
			ringyyList.append(monoyyTensor)
			monozyTensor = zyShiel[nRing_2*numDots:nRing_2*numDots+numDots]
			ringzyList.append(monozyTensor)
			monoxzTensor = xzShiel[nRing_2*numDots:nRing_2*numDots+numDots]
			ringxzList.append(monoxzTensor)
			monoyzTensor = yzShiel[nRing_2*numDots:nRing_2*numDots+numDots]
			ringyzList.append(monoyzTensor)
			monozzTensor = zzShiel[nRing_2*numDots:nRing_2*numDots+numDots]
			ringzzList.append(monozzTensor)

		global inicsOutMpl
		inicsOutMpl = MplInicsOutCanvas()
		toolBar21 = NavigationToolbar(inicsOutMpl, self)

		inicsOutMpl.ax.plot(np.array(inicsHeightList), np.array(ringzzList[0]))
		inicsOutMpl.ax.set_xlabel('Distance (Å)')
		inicsOutMpl.ax.set_ylabel('NICS (ppm)')
		inicsOutMpl.draw()

		inicsPreWidget = QWidget()
		inicsPreviewLayout = QVBoxLayout()
		inicsPreWidget.setLayout(inicsPreviewLayout)

		inicsPreviewLayout.addWidget(inicsOutMpl)
		inicsPreviewLayout.addWidget(toolBar21)
		inicsPreviewLayout.setContentsMargins(0, 0, 0, 0)

		inicspreLayout.addWidget(inicsPreWidget)

		inicsOutResult = QWidget()
		inicsOutResultLayout = QVBoxLayout()
		inicsOutResult.setLayout(inicsOutResultLayout)
		
		global inicsOutText
		inicsOutText = QTextEdit()
		inicsOutText.setText('INICS Results (ppm)\n=================\n')
		inicsOutResultLayout.addWidget(inicsOutText)

		for iRing in range(ringNums):
			inicsOutText.append(f'< Ring No. {iRing+1} >')
			inicsOutText.append(f'Iso: {isoInt[iRing]:.3f}')
			inicsOutText.append(f'Ani: {aniInt[iRing]:.3f}')
			inicsOutText.append(f'XX: {xxInt[iRing]:.3f}')
			inicsOutText.append(f'YX: {yxInt[iRing]:.3f}')
			inicsOutText.append(f'ZX: {zxInt[iRing]:.3f}')
			inicsOutText.append(f'XY: {xyInt[iRing]:.3f}')
			inicsOutText.append(f'YY: {yyInt[iRing]:.3f}')
			inicsOutText.append(f'ZY: {zyInt[iRing]:.3f}')
			inicsOutText.append(f'XZ: {xzInt[iRing]:.3f}')
			inicsOutText.append(f'YZ: {yzInt[iRing]:.3f}')
			inicsOutText.append(f'ZZ: {zzInt[iRing]:.3f}\n')
			molView20.ax.text(atomList[int(len(geomList)+(numDots-1)/2+numDots*iRing)][1], \
				atomList[int(len(geomList)+(numDots-1)/2+numDots*iRing)][2], \
				atomList[int(len(geomList)+(numDots-1)/2+numDots*iRing)][3], iRing+1, \
				bbox = dict(boxstyle = 'round', ec = (0., 0.5, 0.5), fc = (0., 0.8, 0.8)))

		radioBtnGroup = QGroupBox('Component of Shielding Tensors')
		radioBtnLayout = QGridLayout()
		radioBtnGroup.setLayout(radioBtnLayout)

		global ringCombo
		ringCombo = QComboBox()
		ringCombo.addItems([f'{iii}' for iii in (range(1, ringNums+1))])
		ringCombo.setEditable(False)
		ringCombo.setCurrentText('1')
		ringCombo.currentTextChanged.connect(self.ringChanged)

		radion6Group = QButtonGroup()
		isoBtn = QRadioButton('Isotropy', self)
		isoBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(isoBtn, 0, 0)
		aniBtn = QRadioButton('Anisotropy', self)
		aniBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(aniBtn, 0, 1)
		xxBtn = QRadioButton('XX', self)
		xxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xxBtn, 1, 0)
		xyBtn = QRadioButton('XY', self)
		xyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xyBtn, 2, 0)
		xzBtn = QRadioButton('XZ', self)
		xzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(xzBtn, 3, 0)
		yxBtn = QRadioButton('YX', self)
		yxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yxBtn, 1, 1)
		yyBtn = QRadioButton('YY', self)
		yyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yyBtn, 2, 1)
		yzBtn = QRadioButton('YZ', self)
		yzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(yzBtn, 3, 1)
		zxBtn = QRadioButton('ZX', self)
		zxBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zxBtn, 1, 2)
		zyBtn = QRadioButton('ZY', self)
		zyBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zyBtn, 2, 2)
		zzBtn = QRadioButton('ZZ', self)
		zzBtn.toggled.connect(self.tensorClicked)
		radioBtnLayout.addWidget(zzBtn, 3, 2)

		radion6Group.addButton(isoBtn, 1)
		radion6Group.addButton(aniBtn, 2)
		radion6Group.addButton(xxBtn, 3)
		radion6Group.addButton(yzBtn, 4)
		radion6Group.addButton(zxBtn, 5)
		radion6Group.addButton(xyBtn, 6)
		radion6Group.addButton(yyBtn, 7)
		radion6Group.addButton(zyBtn, 8)
		radion6Group.addButton(xzBtn, 9)
		radion6Group.addButton(yzBtn, 10)
		radion6Group.addButton(zzBtn, 11)
		global tensorType
		tensorType = 'ZZ'
		zzBtn.setChecked(True)

		inicsOutResultLayout.addWidget(radioBtnGroup)

		ringGroup = QGroupBox('Ring Selection')
		ringLayout = QGridLayout()
		ringGroup.setLayout(ringLayout)

		ringLayout.addWidget(QLabel(' Ring No.: '), 0, 0)

		ringLayout.addWidget(ringCombo, 0, 1)

		inicsOutResultLayout.addWidget(ringGroup)

		inicsOutBtnGroup = QWidget()
		inicsOutBtnLayout = QHBoxLayout()
		inicsOutBtnGroup.setLayout(inicsOutBtnLayout)

		savePnginicsOutBtn = QPushButton('Save .png')
		savePnginicsOutBtn.clicked.connect(self.savePnginicsOutFunc)
		inicsOutBtnLayout.addWidget(savePnginicsOutBtn)

		saveXlsxInicsOutBtn = QPushButton('Save .xlsx')
		saveXlsxInicsOutBtn.clicked.connect(self.saveXlsxinicsOut)
		inicsOutBtnLayout.addWidget(saveXlsxInicsOutBtn)
		
		inicsOutResultLayout.addWidget(inicsOutBtnGroup)

		inicsOutMainLayout.addWidget(viewer20)
		inicsOutMainLayout.addWidget(inicsPreview)
		inicsOutMainLayout.addWidget(inicsOutResult)
		inicsOutMainWidget.setLayout(inicsOutMainLayout)

		self.setCentralWidget(inicsOutMainWidget)
		self.setMinimumSize(1500, 600)

	def reDrawInicsOut(self, tensor_type, ring_number):
		inicsOutMpl.ax.cla()

		if tensor_type == 'Isotropy':
			inicsOutMpl.ax.plot(np.array(inicsHeightList), np.array(ringisoList[ring_number-1]))
		elif tensor_type == 'Anisotropy':
			inicsOutMpl.ax.plot(np.array(inicsHeightList), np.array(ringaniList[ring_number-1]))
		elif tensor_type == 'XX':
			inicsOutMpl.ax.plot(np.array(inicsHeightList), np.array(ringxxList[ring_number-1]))
		elif tensor_type == 'YX':
			inicsOutMpl.ax.plot(np.array(inicsHeightList), np.array(ringyxList[ring_number-1]))
		elif tensor_type == 'ZX':
			inicsOutMpl.ax.plot(np.array(inicsHeightList), np.array(ringzxList[ring_number-1]))
		elif tensor_type == 'XY':
			inicsOutMpl.ax.plot(np.array(inicsHeightList), np.array(ringxyList[ring_number-1]))
		elif tensor_type == 'YY':
			inicsOutMpl.ax.plot(np.array(inicsHeightList), np.array(ringyyList[ring_number-1]))
		elif tensor_type == 'ZY':
			inicsOutMpl.ax.plot(np.array(inicsHeightList), np.array(ringzyList[ring_number-1]))
		elif tensor_type == 'XZ':
			inicsOutMpl.ax.plot(np.array(inicsHeightList), np.array(ringxzList[ring_number-1]))
		elif tensor_type == 'YZ':
			inicsOutMpl.ax.plot(np.array(inicsHeightList), np.array(ringyzList[ring_number-1]))
		elif tensor_type == 'ZZ':
			inicsOutMpl.ax.plot(np.array(inicsHeightList), np.array(ringzzList[ring_number-1]))

		inicsOutMpl.ax.set_xlabel('Distance (Å)')
		inicsOutMpl.ax.set_ylabel('NICS (ppm)')
		inicsOutMpl.draw()

	def tensorClicked(self):
		global tensorType
		if self.sender().isChecked():
			tensorType = self.sender().text()
		self.reDrawInicsOut(tensorType, int(ringCombo.currentText()))

	def ringChanged(self):
		self.reDrawInicsOut(tensorType, int(ringCombo.currentText()))

	def savePnginicsOutFunc(self):
		figInicsOut.savefig(f'{os.path.splitext(fileName)[0]}_INICS_{tensorType.upper()}_Ring{int(ringCombo.currentText())}.png', dpi = 300)

	def saveXlsxinicsOut(self):
		inicsWB = openpyxl.Workbook()

		summaryWS = inicsWB.active
		summaryWS.title = 'Summary'
		summaryWS.append(['INICS (ppm)', 'Iso', 'Ani', 'XX', 'YX', 'ZX', 'XY', 'YY', 'ZY', 'XZ', 'YZ', 'ZZ'])

		for ringI in range(ringNums):
			summaryWS.append([f'{str(ringI+1)}', isoInt[ringI], aniInt[ringI], xxInt[ringI], yxInt[ringI], zxInt[ringI], xyInt[ringI], yyInt[ringI], zyInt[ringI], xzInt[ringI], yzInt[ringI], zzInt[ringI]])
			inicsWS = inicsWB.create_sheet(title = f'Ring {ringI+1}')
			inicsWS.append(['NICS (ppm)', 'Iso', 'Ani', 'XX', 'YX', 'ZX', 'XY', 'YY', 'ZY', 'XZ', 'YZ', 'ZZ'])
			for heiI in range(len(inicsHeightList)):
				inicsWS.append([f'{inicsHeightList[heiI]:.1f}', ringisoList[ringI][heiI], \
					ringaniList[ringI][heiI], ringxxList[ringI][heiI], ringyxList[ringI][heiI], \
					ringzxList[ringI][heiI], ringxyList[ringI][heiI], ringyyList[ringI][heiI], \
					ringzyList[ringI][heiI], ringxzList[ringI][heiI], ringyzList[ringI][heiI], \
					ringzzList[ringI][heiI]])

		inicsWB.save(f'{os.path.splitext(fileName)[0]}_INICS.xlsx')

'''
*************************************************************************

              TYPE-C WINDOW: for .log files with Bq atoms

*************************************************************************
'''
class TypeCWindow(QMainWindow):
	def __init__(self):
		super().__init__()

		self.setWindowTitle(fileName.split('/')[-1])

		window_toolbar = QToolBar('Type C Toolbar')
		self.addToolBar(window_toolbar)
		window_toolbar.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)

		self.bla_window = None
		blaAction = QAction(QIcon('./assets/bla.png'), 'BLA', self)
		blaAction.setStatusTip('Plot bond length alternation graph for specified bonds.')
		blaAction.triggered.connect(self.blaFunction)

		self.homa_window = None
		homaAction = QAction(QIcon('./assets/homa.png'), 'HOMA', self)
		homaAction.setStatusTip('Compute HOMA value for monocycles.')
		homaAction.triggered.connect(self.homaFunction)

		self.pova_window = None
		povaAction = QAction(QIcon('./assets//poav.png'), 'POAV', self)
		povaAction.setStatusTip('Compute POAV1 and POAV2 for specified atom.')
		povaAction.triggered.connect(self.povaFunction)

		self.nmr_window = None
		nmrAction = QAction(QIcon('./assets/nmr.png'), 'NMR', self)
		nmrAction.setStatusTip('Process NMR result.')
		nmrAction.triggered.connect(self.nmrFunction)

		outAction = QAction(QIcon('./assets/output.png'), 'NICS Output', self)
		outAction.setStatusTip('Process NICS output files.')
		if '_NICS_Scan' in fileName.split('/')[-1]:
			self.scan_out_window = None
			outAction.triggered.connect(self.nicsScanOutput)
		elif '_2D_NICS' in fileName.split('/')[-1] or '_2D_ICSS' in fileName.split('/')[-1] or '_2DICSS' in fileName.split('/')[-1] or '_2DNICS' in fileName.split('/')[-1]:
			self.n2d_out_window = None
			outAction.triggered.connect(self.nics2DOutput)
		elif '_3D_NICS' in fileName.split('/')[-1] or '_3D_ICSS' in fileName.split('/')[-1] or '_3DICSS' in fileName.split('/')[-1] or '_3DNICS' in fileName.split('/')[-1]:
			self.n3d_out_window = None
			outAction.triggered.connect(self.nics3DOutput)
		elif '_INICS' in fileName.split('/')[-1]:
			self.inics_out_window = None
			outAction.triggered.connect(self.inicsOutput)
		else:
			self.out_window = None
			outAction.triggered.connect(self.nicsOutput)

		window_toolbar.addAction(blaAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(homaAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(povaAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(nmrAction)
		window_toolbar.addSeparator()
		window_toolbar.addAction(outAction)

		window_toolbar.setMovable(False)

		self.setStatusBar(QStatusBar(self))

		molView = Mpl3DCanvas()
		molView.ax.axis('off')

		if xMax == xMin:
			molView.ax.set_xlim(xMin - 1, xMin + 1)
			molView.ax.set_box_aspect([2, yMax - yMin, zMax - zMin])
		elif yMax == yMin:
			molView.ax.set_ylim(yMin - 1, yMin + 1)
			molView.ax.set_box_aspect([xMax - xMin, 2, zMax - zMin])
		elif zMax == zMin:
			molView.ax.set_zlim(zMin - 1, zMin + 1)
			molView.ax.set_box_aspect([xMax - xMin, yMax - yMin, 2])
		else:
			molView.ax.set_box_aspect([xMax - xMin, yMax - yMin, zMax - zMin])
		
		for bndAtom_i in bndAtom:
			molView.ax.plot([xCoor[bndAtom_i[0]], xCoor[bndAtom_i[1]]], [yCoor[bndAtom_i[0]], yCoor[bndAtom_i[1]]], \
				[zCoor[bndAtom_i[0]], zCoor[bndAtom_i[1]]], '0.5')
		molView.ax.scatter(np.array([xCoor]), np.array([yCoor]), np.array([zCoor]), color = colorList, edgecolors = '0.0', s = 40, depthshade = 0)

		viewLayout = QVBoxLayout()
		viewLayout.addWidget(molView)

		viewer = QWidget()
		viewer.setLayout(viewLayout)

		self.setCentralWidget(viewer)

	def blaFunction(self):
		if self.bla_window is None:
			self.bla_window = BLAWindow()
		self.bla_window.show()

	def homaFunction(self):
		if self.homa_window is None:
			self.homa_window = HOMAWindow()
		self.homa_window.show()

	def povaFunction(self):
		if self.pova_window is None:
			self.pova_window = PovaWindow()
		self.pova_window.show()
		
	def nmrFunction(self):
		if self.nmr_window is None:
			self.nmr_window = NmrWindow()
		self.nmr_window.show()
	def nicsOutput(self):
		if self.out_window is None:
			self.out_window = NICSoutWindow()
		self.out_window.show()

	def nicsScanOutput(self):
		if self.scan_out_window is None:
			self.scan_out_window = NICSscanWindow()
		self.scan_out_window.show()
		
	def nics2DOutput(self):
		if self.n2d_out_window is None:
			self.n2d_out_window = NICS2DWindow()
		self.n2d_out_window.show()
		
	def nics3DOutput(self):
		if self.n3d_out_window is None:
			self.n3d_out_window = NICS3DWindow()
		self.n3d_out_window.show()

	def inicsOutput(self):
		if self.inics_out_window is None:
			self.inics_out_window = INICSWindow()
		self.inics_out_window.show()

'''
*************************************************************************

                             SETTING WINDOW

*************************************************************************
'''
class SetWindow(QMainWindow):
	def __init__(self):
		super().__init__()

		self.setWindowTitle('py.Aroma Preferences')

		set_widget = QWidget()
		mainLayout = QVBoxLayout()

		genGroup = QGroupBox('General')
		mainLayout.addWidget(genGroup)
		genLayout = QVBoxLayout()

		global icssCheckBox, connecCheckBox, inputCheckBox
		icssCheckBox = QCheckBox('Save ICSS instead of NICS for 2D and 3D processing.')
		if configFile.get('general', 'icss') == '1':
			icssCheckBox.setCheckState(Qt.CheckState.Checked)
		genLayout.addWidget(icssCheckBox)

		connecCheckBox = QCheckBox('(Recommend) Write connectivity when saving input files.')
		if configFile.get('general', 'connectivity') == 'true':
			connecCheckBox.setCheckState(Qt.CheckState.Checked)
		genLayout.addWidget(connecCheckBox)

		inputCheckBox = QCheckBox('Overwrite routine section when reading input files.')
		if configFile.get('general', 'input') == 'true':
			inputCheckBox.setCheckState(Qt.CheckState.Checked)
		genLayout.addWidget(inputCheckBox)

		genGroup.setLayout(genLayout)

		homaGroup = QGroupBox('HOMA Parameters')            # HOMA
		mainLayout.addWidget(homaGroup)

		homaLayout = QGridLayout()

		bondTypeLabel = QLabel('Bond Type')
		boldFont = QFont()
		boldFont.setBold(True)
		bondTypeLabel.setFont(boldFont)
		homaLayout.addWidget(bondTypeLabel, 0, 0)

		homaLayout.addWidget(QLabel('      '), 0, 1)

		roptLabel = QLabel('Optimal Bond Length (Å)')
		roptLabel.setFont(boldFont)
		homaLayout.addWidget(roptLabel, 0, 2)

		homaLayout.addWidget(QLabel('      '), 0, 3)

		alphaLabel = QLabel('Normalization Constant')
		alphaLabel.setFont(boldFont)
		homaLayout.addWidget(alphaLabel, 0, 4)

		ccLabel = QLabel('C-C')								 # HOMA for C-C
		ccLabel.setAlignment(Qt.AlignmentFlag.AlignCenter)
		homaLayout.addWidget(ccLabel, 1, 0)

		global ccRopt, cnRopt, coRopt, cpRopt, csRopt, nnRopt, noRopt, bnRopt, cseRopt
		global ccAlpha, cnAlpha, coAlpha, cpAlpha, csAlpha, nnAlpha, noAlpha, bnAlpha, cseAlpha
		ccRopt = QLineEdit()
		ccRopt.setText(configFile.get('homa', 'cc').split(',')[0])
		homaLayout.addWidget(ccRopt, 1, 2)

		ccAlpha = QLineEdit()
		ccAlpha.setText(configFile.get('homa', 'cc').split(',')[1])
		homaLayout.addWidget(ccAlpha, 1, 4)

		cnLabel = QLabel('C-N')								 # HOMA for C-N
		cnLabel.setAlignment(Qt.AlignmentFlag.AlignCenter)
		homaLayout.addWidget(cnLabel, 2, 0)

		cnRopt = QLineEdit()
		cnRopt.setText(configFile.get('homa', 'cn').split(',')[0])
		homaLayout.addWidget(cnRopt, 2, 2)

		cnAlpha = QLineEdit()
		cnAlpha.setText(configFile.get('homa', 'cn').split(',')[1])
		homaLayout.addWidget(cnAlpha, 2, 4)

		coLabel = QLabel('C-O')								 # HOMA for C-O
		coLabel.setAlignment(Qt.AlignmentFlag.AlignCenter)
		homaLayout.addWidget(coLabel, 3, 0)

		coRopt = QLineEdit()
		coRopt.setText(configFile.get('homa', 'co').split(',')[0])
		homaLayout.addWidget(coRopt, 3, 2)

		coAlpha = QLineEdit()
		coAlpha.setText(configFile.get('homa', 'co').split(',')[1])
		homaLayout.addWidget(coAlpha, 3, 4)

		cpLabel = QLabel('C-P')								 # HOMA for C-P
		cpLabel.setAlignment(Qt.AlignmentFlag.AlignCenter)
		homaLayout.addWidget(cpLabel, 4, 0)

		cpRopt = QLineEdit()
		cpRopt.setText(configFile.get('homa', 'cp').split(',')[0])
		homaLayout.addWidget(cpRopt, 4, 2)

		cpAlpha = QLineEdit()
		cpAlpha.setText(configFile.get('homa', 'cp').split(',')[1])
		homaLayout.addWidget(cpAlpha, 4, 4)

		csLabel = QLabel('C-S')								 # HOMA for C-S
		csLabel.setAlignment(Qt.AlignmentFlag.AlignCenter)
		homaLayout.addWidget(csLabel, 5, 0)

		csRopt = QLineEdit()
		csRopt.setText(configFile.get('homa', 'cs').split(',')[0])
		homaLayout.addWidget(csRopt, 5, 2)

		csAlpha = QLineEdit()
		csAlpha.setText(configFile.get('homa', 'cs').split(',')[1])
		homaLayout.addWidget(csAlpha, 5, 4)

		nnLabel = QLabel('N-N')								 # HOMA for N-N
		nnLabel.setAlignment(Qt.AlignmentFlag.AlignCenter)
		homaLayout.addWidget(nnLabel, 6, 0)

		nnRopt = QLineEdit()
		nnRopt.setText(configFile.get('homa', 'nn').split(',')[0])
		homaLayout.addWidget(nnRopt, 6, 2)

		nnAlpha = QLineEdit()
		nnAlpha.setText(configFile.get('homa', 'nn').split(',')[1])
		homaLayout.addWidget(nnAlpha, 6, 4)

		noLabel = QLabel('N-O')								 # HOMA for N-O
		noLabel.setAlignment(Qt.AlignmentFlag.AlignCenter)
		homaLayout.addWidget(noLabel, 7, 0)

		noRopt = QLineEdit()
		noRopt.setText(configFile.get('homa', 'no').split(',')[0])
		homaLayout.addWidget(noRopt, 7, 2)

		noAlpha = QLineEdit()
		noAlpha.setText(configFile.get('homa', 'no').split(',')[1])
		homaLayout.addWidget(noAlpha, 7, 4)

		bnLabel = QLabel('B-N')								 # HOMA for B-N
		bnLabel.setAlignment(Qt.AlignmentFlag.AlignCenter)
		homaLayout.addWidget(bnLabel, 8, 0)

		bnRopt = QLineEdit()
		bnRopt.setText(configFile.get('homa', 'bn').split(',')[0])
		homaLayout.addWidget(bnRopt, 8, 2)

		bnAlpha = QLineEdit()
		bnAlpha.setText(configFile.get('homa', 'bn').split(',')[1])
		homaLayout.addWidget(bnAlpha, 8, 4)

		cseLabel = QLabel('C-Se')								 # HOMA for C-Se
		cseLabel.setAlignment(Qt.AlignmentFlag.AlignCenter)
		homaLayout.addWidget(cseLabel, 9, 0)

		cseRopt = QLineEdit()
		cseRopt.setText(configFile.get('homa', 'cse').split(',')[0])
		homaLayout.addWidget(cseRopt, 9, 2)

		cseAlpha = QLineEdit()
		cseAlpha.setText(configFile.get('homa', 'cse').split(',')[1])
		homaLayout.addWidget(cseAlpha, 9, 4)

		homaGroup.setLayout(homaLayout)

		btnLayout = QHBoxLayout()          # Button Section
		mainLayout.addLayout(btnLayout)

		dftBtn = QPushButton('Default')
		btnLayout.addWidget(dftBtn)
		dftBtn.clicked.connect(self.goDefault)

		saveBtn = QPushButton('Save')
		btnLayout.addWidget(saveBtn)
		saveBtn.clicked.connect(self.saveSet)

		clsBtn = QPushButton('Close')
		btnLayout.addWidget(clsBtn)
		clsBtn.clicked.connect(lambda: self.close())

		set_widget.setLayout(mainLayout)

		self.setCentralWidget(set_widget)

	def goDefault(self):
		ccRopt.setText('1.388')
		ccAlpha.setText('257.7')
		cnRopt.setText('1.334')
		cnAlpha.setText('93.52')
		coRopt.setText('1.265')
		coAlpha.setText('157.38')
		cpRopt.setText('1.698')
		cpAlpha.setText('118.91')
		csRopt.setText('1.677')
		csAlpha.setText('94.09')
		nnRopt.setText('1.309')
		nnAlpha.setText('130.33')
		noRopt.setText('1.248')
		noAlpha.setText('57.21')
		bnRopt.setText('1.402')
		bnAlpha.setText('72.03')
		cseRopt.setText('1.8217')
		cseAlpha.setText('84.9144')
		icssCheckBox.setCheckState(Qt.CheckState.Unchecked)
		connecCheckBox.setCheckState(Qt.CheckState.Checked)
		inputCheckBox.setCheckState(Qt.CheckState.Checked)

	def saveSet(self):
		configFile.set('homa', 'cc', f'{ccRopt.text()},{ccAlpha.text()}')
		configFile.set('homa', 'cn', f'{cnRopt.text()},{cnAlpha.text()}')
		configFile.set('homa', 'co', f'{coRopt.text()},{coAlpha.text()}')
		configFile.set('homa', 'cp', f'{cpRopt.text()},{cpAlpha.text()}')
		configFile.set('homa', 'cs', f'{csRopt.text()},{csAlpha.text()}')
		configFile.set('homa', 'nn', f'{nnRopt.text()},{nnAlpha.text()}')
		configFile.set('homa', 'no', f'{noRopt.text()},{noAlpha.text()}')
		configFile.set('homa', 'bn', f'{bnRopt.text()},{bnAlpha.text()}')
		configFile.set('homa', 'cse', f'{cseRopt.text()},{cseAlpha.text()}')

		if icssCheckBox.isChecked():
			configFile.set('general', 'icss', '1')
		else:
			configFile.set('general', 'icss', '-1')

		if connecCheckBox.isChecked():
			configFile.set('general', 'connectivity', 'true')
		else:
			configFile.set('general', 'connectivity', 'false')

		if inputCheckBox.isChecked():
			configFile.set('general', 'input', 'true')
		else:
			configFile.set('general', 'input', 'false')

		with open(configFilePath, 'w') as configFileNew:
			configFile.write(configFileNew)
		self.close()

'''
*************************************************************************

                        ABOUT THIS PROGRAM WINDOW

*************************************************************************
'''
class AbtWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('About py.Aroma')

		abtBox = QWidget()
		abtLayout = QHBoxLayout()
		abtBox.setLayout(abtLayout)

		tab_1 = QLabel()
		tab_1.setPixmap(QPixmap('./assets/abt_window.png'))
		tab_1.setScaledContents(True)
		abtLayout.addWidget(tab_1)

		self.setCentralWidget(abtBox)
		self.setFixedSize(400, 600)

'''
*************************************************************************

                              START UP WINDOW

*************************************************************************
'''
class MainWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('py.Aroma 4')

		main_toolbar = QToolBar('Show Tool Bar')
		self.addToolBar(main_toolbar)

		open_action = QAction(QIcon('./assets/open_icon.png'), 'Open...', self)
		open_action.setStatusTip('Browse a file for processing.')
		open_action.triggered.connect(self.showOpenFileDialog)
		open_action.setShortcut(QKeySequence("Ctrl+o"))

		self.set_window = None
		set_action = QAction(QIcon('./assets/setting_icon.png'), 'Settings', self)
		set_action.setStatusTip('Open program setting panel.')
		set_action.triggered.connect(self.settingWindow)

		help_action = QAction(QIcon('./assets/help_icon.png'), 'Online Document', self)
		help_action.setStatusTip('Open online user manual.')
		help_action.triggered.connect(self.openManual)
		help_action.setShortcut(QKeySequence("Ctrl+h"))

		self.abt_window = None
		abt_action = QAction(QIcon('./assets/about_icon.png'), 'About py.Aroma', self)
		abt_action.setStatusTip('About this program.')
		abt_action.triggered.connect(self.aboutWindow)

		bug_action = QAction(QIcon('./assets/email_icon.png'), 'Contact Developer', self)
		bug_action.setStatusTip('Contact developer via e-mail.')
		bug_action.triggered.connect(self.bugReport)

		exit_action = QAction(QIcon('./assets/quit_icon.png'), 'Quit py.Aroma', self)
		exit_action.setStatusTip('Quit program.')
		exit_action.triggered.connect(sys.exit)
		exit_action.setShortcut(QKeySequence("Ctrl+q"))

		git_action = QAction(QIcon('./assets/git_icon.png'), 'GitHub Page', self)
		git_action.setStatusTip('Check other programs developed by me!')
		git_action.triggered.connect(self.gitHub)

		hp_action = QAction(QIcon('./assets/web.png'), 'Homepage', self)
		hp_action.setStatusTip('Open py.Aroma homepage.')
		hp_action.triggered.connect(self.hpWeb)

		main_toolbar.addAction(open_action)
		main_toolbar.addSeparator()
		main_toolbar.addAction(set_action)
		main_toolbar.addSeparator()
		main_toolbar.addAction(help_action)
		main_toolbar.addSeparator()
		main_toolbar.addAction(abt_action)
		main_toolbar.addSeparator()
		main_toolbar.addAction(bug_action)
		main_toolbar.addSeparator()
		main_toolbar.addAction(git_action)
		main_toolbar.addSeparator()
		main_toolbar.addAction(hp_action)
		main_toolbar.addSeparator()
		main_toolbar.addAction(exit_action)

		main_toolbar.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonIconOnly)
		main_toolbar.setMovable(False)

		self.setStatusBar(QStatusBar(self))

		main_menu = self.menuBar().addMenu('&py.Aroma')
		main_menu.addAction(open_action)
		main_menu.addSeparator()
		main_menu.addAction(help_action)
		main_menu.addAction(git_action)
		main_menu.addAction(hp_action)
		sub_menu = main_menu.addMenu('Program')
		sub_menu.addAction(set_action)
		sub_menu.addAction(abt_action)
		sub_menu.addAction(bug_action)
		sub_menu.addAction(exit_action)

		banner_label = QLabel()
		banner_label.setPixmap(QPixmap('./assets/main_window.png'))
		banner_label.setScaledContents(True)

		self.setCentralWidget(banner_label)
		self.setFixedSize(700, 766)

	def settingWindow(self):
		if self.set_window is None:
			self.set_window = SetWindow()
		self.set_window.show()

	def showOpenFileDialog(self):
		global fileName, fileType, atomList, geomList
		home_dir = str(Path.home())
		fileName = QFileDialog.getOpenFileName(self, 'Open file', home_dir)[0]
		fileType = os.path.splitext(fileName)[1][1:]
		errorFlag = 0

		if fileType.lower() == 'mol2':
			atomList, geomList = readFile.read_mol2(fileName)
		elif fileType.lower() == 'pdb':
			atomList, geomList = readFile.read_pdb(fileName)
		elif fileType.lower() == 'xyz':
			atomList, geomList = readFile.read_xyz(fileName)
		elif fileType.lower() == 'gjf' or fileType.lower() == 'com':
			atomList, geomList = readFile.read_gjf(fileName)
		elif fileType.lower() == 'log' or fileType.lower() == 'out':
			atomList, geomList = readFile.read_log(fileName)
		elif fileType.lower() == '':
			pass
			errorFlag = 1
		else:
			QMessageBox.critical(None, 'Unsupported File', f'py.Aroma could not process .{fileType.lower()} file.')
			errorFlag = 1

		if errorFlag == 0:
			global xMin, xMax, yMin, yMax, zMin, zMax
			xMin, xMax, yMin, yMax, zMin, zMax = geomAnalyzer.find_max_min(geomList)
			global xMin2, xMax2, yMin2, yMax2, zMin2, zMax2
			xMin2, xMax2, yMin2, yMax2, zMin2, zMax2 = geomAnalyzer.find_max_min(atomList)
			global xCoor, yCoor, zCoor, colorList
			xCoor, yCoor, zCoor, colorList = geomAnalyzer.save_coor_list(geomList)
			global xAllCoor, yAllCoor, zAllCoor, colorAllList
			xAllCoor, yAllCoor, zAllCoor, colorAllList = geomAnalyzer.save_coor_list(atomList)
			global conMatrix, bndAtom
			conMatrix, bndAtom = geomAnalyzer.get_connectivity(geomList)

			if len(atomList) == len(geomList) and (fileType.lower() == 'log'or fileType.lower() == 'out'):
				self.typeB_window = TypeBWindow()
				self.typeB_window.show()
			elif len(atomList) != len(geomList) and (fileType.lower() == 'log'or fileType.lower() == 'out'):
				self.typeC_window = TypeCWindow()
				self.typeC_window.show()
			else:
				self.typeA_window = TypeAWindow()
				self.typeA_window.show()

	def aboutWindow(self):
		if self.abt_window is None:
			self.abt_window = AbtWindow()
		self.abt_window.show()

	def openManual(self):
		webbrowser.open('https://wongzit.github.io/program/pyaroma/manual_v4.pdf')

	def bugReport(self):
		webbrowser.open('mailto:wang.zhe.dr@gmail.com')

	def gitHub(self):
		webbrowser.open('https://github.com/wongzit')

	def hpWeb(self):
		webbrowser.open('https://wongzit.github.io/program/pyaroma/')

# Read config.ini
global configFile, configFilePath
configFile = configparser.ConfigParser()
configFilePath = os.path.join(os.path.dirname(__file__), 'config.ini')
configFile.read(configFilePath)

app = QApplication([])

window = MainWindow()
window.show()

app.exec()
